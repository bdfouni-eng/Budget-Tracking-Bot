"""
BUDGET TRACKER BOT - ENHANCED VERSION WITH DOWNLOAD FEATURE
"""

import os
import re
import logging
import shutil
import sys
import json
import asyncio
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Tuple, Optional, List, Any
from difflib import get_close_matches
import threading
from dataclasses import dataclass
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor
from telegram import BotCommand

# ========== EXCEL IMPORT CHECK (PythonAnywhere compatible) ==========
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    # Optional: import pandas for advanced operations if needed
    EXCEL_SUPPORT = True
except ImportError as e:
    EXCEL_SUPPORT = False
    print(f"⚠️  Excel libraries not found: {e}")

# ========== TELEGRAM BOT IMPORTS ==========
try:
    from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackQueryHandler
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    print("⚠️  python-telegram-bot not found. Bot will not start.")

# ========== ONEDRIVE API FUNCTIONS ==========
try:
    import msal
    import requests
    ONEDRIVE_API_AVAILABLE = True
except ImportError:
    ONEDRIVE_API_AVAILABLE = False
    print("⚠️  OneDrive API libraries (msal, requests) not found. OneDrive sync will not work.")

def get_onedrive_access_token():
    """Get access token for OneDrive API"""
    client_id = os.getenv("ONEDRIVE_CLIENT_ID")
    client_secret = os.getenv("ONEDRIVE_CLIENT_SECRET")
    tenant_id = os.getenv("ONEDRIVE_TENANT_ID")
    
    if not all([client_id, client_secret, tenant_id]):
        return None
    
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    app = msal.ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=authority
    )
    
    scope = ["https://graph.microsoft.com/.default"]
    result = app.acquire_token_for_client(scopes=scope)
    
    if "access_token" in result:
        return result["access_token"]
    else:
        logger.error(f"Could not acquire token: {result.get('error_description')}")
        return None

def download_from_onedrive_api() -> Tuple[bool, str]:
    """Download file from OneDrive to local copy"""
    if not ONEDRIVE_API_AVAILABLE:
        return False, "❌ OneDrive API libraries not installed"
    
    access_token = get_onedrive_access_token()
    if not access_token:
        return False, "❌ Could not get OneDrive access token"
    
    file_path = os.getenv("ONEDRIVE_FILE_PATH", "/drive/root:/budget.xlsx")
    
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        url = f'https://graph.microsoft.com/v1.0/me{file_path}:/content'
        
        response = requests.get(url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            # Create backup of existing file if it exists
            if LOCAL_COPY_PATH.exists():
                create_backup()
            
            with open(LOCAL_COPY_PATH, 'wb') as f:
                f.write(response.content)
            
            logger.info(f"Downloaded file from OneDrive: {file_path}")
            return True, "✅ File downloaded from OneDrive"
        else:
            logger.error(f"Failed to download: {response.status_code} - {response.text[:200]}")
            return False, f"❌ Failed to download from OneDrive: {response.status_code}"
            
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return False, f"❌ Download error: {str(e)[:200]}"

def upload_to_onedrive_api() -> Tuple[bool, str]:
    """Upload local file back to OneDrive"""
    if not ONEDRIVE_API_AVAILABLE:
        return False, "❌ OneDrive API libraries not installed"
    
    if not LOCAL_COPY_PATH.exists():
        return False, "❌ Local file not found"
    
    access_token = get_onedrive_access_token()
    if not access_token:
        return False, "❌ Could not get OneDrive access token"
    
    file_path = os.getenv("ONEDRIVE_FILE_PATH", "/drive/root:/budget.xlsx")
    
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/octet-stream'
        }
        
        url = f'https://graph.microsoft.com/v1.0/me{file_path}:/content'
        
        with open(LOCAL_COPY_PATH, 'rb') as f:
            response = requests.put(url, headers=headers, data=f, timeout=30)
        
        if response.status_code in [200, 201]:
            logger.info(f"Uploaded file to OneDrive: {file_path}")
            return True, "✅ File uploaded to OneDrive"
        else:
            logger.error(f"Failed to upload: {response.status_code} - {response.text[:200]}")
            return False, f"❌ Failed to upload to OneDrive: {response.status_code}"
            
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        return False, f"❌ Upload error: {str(e)[:200]}"

# ========== CONFIGURATION ==========
from dotenv import load_dotenv
load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")

# Handle both ALLOWED_USER_ID (singular) and ALLOWED_USER_IDS (plural)
ALLOWED_USER_IDS = []
user_ids_str = os.getenv("ALLOWED_USER_IDS", os.getenv("ALLOWED_USER_ID", ""))
if user_ids_str:
    for id_str in user_ids_str.split(","):
        id_str = id_str.strip()
        if id_str:
            try:
                ALLOWED_USER_IDS.append(int(id_str))
            except ValueError:
                print(f"⚠️ Warning: Invalid user ID '{id_str}' in .env file")

# File paths
ONEDRIVE_PATH = Path(os.getenv("ONEDRIVE_PATH", r"/home/ubuntu/Tracking_Budget_Sheet_Python/temp_budget.xlsm"))
LOCAL_COPY_PATH = Path(os.getenv("LOCAL_COPY_PATH", r"/home/ubuntu/Tracking_Budget_Sheet_Python/temp_budget.xlsm"))
TRACKING_SHEET_NAME = "Budget Tracking"
DROPDOWN_SHEET_NAME = "Dropdown Data"
EXPORT_DIR = Path(os.getenv("EXPORT_DIR", "./exports"))

# Backup configuration
BACKUP_DIR = Path(os.getenv("BACKUP_DIR", "./backups"))
BACKUP_RETENTION_DAYS = 7

# Matching configuration
MINIMUM_CONFIDENCE = 0.5  # 50% minimum confidence to accept a match
FUZZY_MATCH_THRESHOLD = 0.6  # Minimum for fuzzy matching

# Common words to reject (these won't match any category)
COMMON_NON_CATEGORIES = [
    'ok', 'okay', 'yes', 'no', 'maybe', 'test', 'hello', 'hi', 
    'thanks', 'thank', 'please', 'help', 'what', 'how', 'why',
    'when', 'where', 'who', 'which', 'good', 'bad', 'nice',
    'great', 'fine', 'well', 'sorry', 'excuse', 'hey', 'yo'
]

# Currency configuration
@dataclass
class CurrencyConfig:
    symbols: List[str]
    column: str
    decimal_places: int = 2
    thousands_separator: bool = True

CURRENCIES = {
    'USD': CurrencyConfig(symbols=['$', 'usd', 'dollar'], column='H', decimal_places=2),
    'LBP': CurrencyConfig(symbols=['lbp', 'ليرة', 'lira', 'ل.ل'], column='I', decimal_places=0, thousands_separator=True),
    'EURO': CurrencyConfig(symbols=['€', 'euro', 'eur'], column='J', decimal_places=2)
}
DEFAULT_CURRENCY = 'USD'

# Payment types with additional options
PAYMENT_TYPES = {
    "Cash": "💵",
    "Card": "💳",
    "Bank Transfer": "🏦",
    "Digital Wallet": "📱"
}

CARD_KEYWORDS = ["card", "credit", "debit", "visa", "mastercard", "amex", "paypal", "bank"]
DIGITAL_KEYWORDS = ["paypal", "digital", "wallet", "apple pay", "google pay"]

# Setup logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.FileHandler('budget_bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Thread safety for Excel operations
excel_lock = threading.Lock()

# Cache for loaded tables
_table_cache = {}
_table_cache_timestamp = None

# ========== NEW: DOWNLOAD/EXPORT FUNCTIONS ==========

def setup_export_directory():
    """Create export directory if it doesn't exist"""
    EXPORT_DIR.mkdir(exist_ok=True)

def export_to_csv(time_range: str = "month") -> Tuple[bool, str]:
    """Export transactions to CSV format
    time_range: "day", "week", "month", "year", "all"
    """
    try:
        if not EXCEL_SUPPORT:
            return False, "❌ Excel libraries not installed"
        
        # Copy fresh file
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}"
        
        wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
        
        if TRACKING_SHEET_NAME not in wb.sheetnames:
            return False, f"❌ Sheet '{TRACKING_SHEET_NAME}' not found"
        
        sheet = wb[TRACKING_SHEET_NAME]
        
        # Find date range
        today = datetime.now()
        if time_range == "day":
            start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        elif time_range == "week":
            start_date = today - timedelta(days=today.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        elif time_range == "month":
            start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        elif time_range == "year":
            start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:  # "all"
            start_date = None
        
        # Find all transactions
        transactions = []
        row = 12
        
        while True:
            date_val = sheet.cell(row=row, column=3).value  # Column C
            if date_val is None or date_val == "":
                break
            
            # Check date if range specified
            if start_date:
                try:
                    if isinstance(date_val, datetime):
                        trans_date = date_val
                    else:
                        trans_date = datetime.strptime(str(date_val), "%d-%b-%y")
                    
                    if trans_date < start_date:
                        row += 1
                        continue
                except:
                    pass
            
            # Extract transaction data
            transaction = {
                'date': date_val,
                'payment': sheet.cell(row=row, column=4).value or "Cash",  # Column D
                'type': sheet.cell(row=row, column=5).value or "Expenses",  # Column E
                'category': sheet.cell(row=row, column=6).value or "",  # Column F
                'subcategory': sheet.cell(row=row, column=7).value or "",  # Column G
                'usd': sheet.cell(row=row, column=8).value,  # Column H
                'lbp': sheet.cell(row=row, column=9).value,  # Column I
                'euro': sheet.cell(row=row, column=10).value,  # Column J
                'notes': sheet.cell(row=row, column=11).value or ""  # Column K
            }
            transactions.append(transaction)
            row += 1
            
            if row > 1000:  # Safety limit
                break
        
        wb.close()
        
        # Create CSV content
        if not transactions:
            return False, "❌ No transactions found"
        
        csv_lines = []
        # Header
        csv_lines.append("Date,Payment Type,Transaction Type,Category,Subcategory,USD,LBP,EURO,Notes")
        
        # Data
        for trans in transactions:
            date_str = trans['date'].strftime("%Y-%m-%d") if hasattr(trans['date'], 'strftime') else str(trans['date'])
            
            # Escape notes properly
            if trans['notes']:
                # Replace double quotes with two double quotes for CSV
                notes_escaped = trans['notes'].replace('"', '""')
            else:
                notes_escaped = ""
            
            # Build CSV line
            csv_line = (
                f'"{date_str}",'
                f'"{trans["payment"]}",'
                f'"{trans["type"]}",'
                f'"{trans["category"]}",'
                f'"{trans["subcategory"]}",'
                f'{trans["usd"] or ""},'
                f'{trans["lbp"] or ""},'
                f'{trans["euro"] or ""},'
                f'"{notes_escaped}"'
            )
            csv_lines.append(csv_line)
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transactions_{time_range}_{timestamp}.csv"
        filepath = EXPORT_DIR / filename
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(csv_lines))
        
        logger.info(f"Exported {len(transactions)} transactions to {filepath}")
        return True, str(filepath)
        
    except Exception as e:
        logger.error(f"Export error: {str(e)}", exc_info=True)
        return False, f"❌ Export error: {str(e)[:200]}"

def export_summary() -> Tuple[bool, str]:
    """Export summary statistics"""
    try:
        if not EXCEL_SUPPORT:
            return False, "❌ Excel libraries not installed"
        
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}"
        
        wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
        
        if TRACKING_SHEET_NAME not in wb.sheetnames:
            return False, f"❌ Sheet '{TRACKING_SHEET_NAME}' not found"
        
        sheet = wb[TRACKING_SHEET_NAME]
        
        # Calculate totals
        usd_total = 0
        lbp_total = 0
        euro_total = 0
        transaction_count = 0
        row = 12
        
        # Category breakdown
        categories = {}
        
        while True:
            if sheet.cell(row=row, column=3).value is None:
                break
            
            # Get category
            category = sheet.cell(row=row, column=6).value  # Column F
            if category:
                if category not in categories:
                    categories[category] = {
                        'usd': 0,
                        'lbp': 0,
                        'euro': 0,
                        'count': 0
                    }
                
                # Add amounts
                usd_val = sheet.cell(row=row, column=8).value or 0  # Column H
                lbp_val = sheet.cell(row=row, column=9).value or 0  # Column I
                euro_val = sheet.cell(row=row, column=10).value or 0  # Column J
                
                categories[category]['usd'] += usd_val
                categories[category]['lbp'] += lbp_val
                categories[category]['euro'] += euro_val
                categories[category]['count'] += 1
                
                usd_total += usd_val
                lbp_total += lbp_val
                euro_total += euro_val
                transaction_count += 1
            
            row += 1
            if row > 1000:
                break
        
        wb.close()
        
        # Create summary JSON
        summary = {
            'export_date': datetime.now().isoformat(),
            'total_transactions': transaction_count,
            'totals': {
                'USD': usd_total,
                'LBP': lbp_total,
                'EURO': euro_total
            },
            'categories': categories,
            'file_info': {
                'source': str(LOCAL_COPY_PATH),
                'last_modified': datetime.fromtimestamp(LOCAL_COPY_PATH.stat().st_mtime).isoformat() if LOCAL_COPY_PATH.exists() else None
            }
        }
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"summary_{timestamp}.json"
        filepath = EXPORT_DIR / filename
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, default=str)
        
        logger.info(f"Exported summary to {filepath}")
        return True, str(filepath)
        
    except Exception as e:
        logger.error(f"Summary export error: {str(e)}", exc_info=True)
        return False, f"❌ Summary error: {str(e)[:200]}"

def create_backup_copy() -> Tuple[bool, str]:
    """Create a backup copy of the Excel file for download"""
    try:
        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local copy not found"
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"budget_backup_{timestamp}.xlsm"
        backup_path = BACKUP_DIR / backup_filename
        
        shutil.copy2(LOCAL_COPY_PATH, backup_path)
        
        logger.info(f"Created downloadable backup: {backup_path}")
        return True, str(backup_path)
        
    except Exception as e:
        logger.error(f"Backup copy error: {str(e)}")
        return False, f"❌ Backup copy error: {str(e)[:200]}"

# ========== UTILITY FUNCTIONS ==========

def is_authorized(user_id: int) -> bool:
    """Check if user is authorized to use the bot"""
    return user_id in ALLOWED_USER_IDS

def format_currency_amount(amount: float, currency: str) -> str:
    """Format amount according to currency rules - supports negative numbers"""
    if amount is None:
        return "No amount"
    
    # If no currency, just return the number
    if not currency:
        return f"{amount:,.2f}"
    
    is_negative = amount < 0
    abs_amount = abs(amount)
    
    if currency not in CURRENCIES:
        currency = DEFAULT_CURRENCY
    
    config = CURRENCIES[currency]
    
    if config.thousands_separator:
        formatted = f"{abs_amount:,.{config.decimal_places}f}"
    else:
        formatted = f"{abs_amount:.{config.decimal_places}f}"
    
    symbol = config.symbols[0]
    
    if is_negative:
        if currency == "USD":
            return f"-{symbol}{formatted}"
        elif currency == "EURO":
            return f"-{symbol}{formatted}"
        elif currency == "LBP":
            return f"-{formatted} LBP"
        else:
            return f"-{formatted}"
    else:
        if currency == "USD":
            return f"{symbol}{formatted}"
        elif currency == "EURO":
            return f"{symbol}{formatted}"
        elif currency == "LBP":
            return f"{formatted} LBP"
        return formatted

def format_transaction_response(transaction: Dict, action: str = "Added") -> str:
    """Format a transaction for display in messages"""
    date_val = transaction['date']
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
    
    # Format amounts
    amounts = []
    if transaction.get('usd') is not None:
        prefix = "🔄 " if transaction['usd'] < 0 else ""
        amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
    if transaction.get('lbp') is not None:
        prefix = "🔄 " if transaction['lbp'] < 0 else ""
        amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
    if transaction.get('euro') is not None:
        prefix = "🔄 " if transaction['euro'] < 0 else ""
        amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
    
    amount_str = " + ".join(amounts) if amounts else "No amount"
    
    payment_emoji = "💳" if str(transaction.get('payment', '')).lower() == "card" else "💵"
    
    import html
    safe_item = html.escape(transaction.get('subcategory') or transaction.get('category', ''))
    safe_category = html.escape(transaction.get('category', 'Unknown'))
    
    response = (
        f"✅ <b>Transaction {action}:</b>\n\n"
        f"{payment_emoji} {date_str}: {safe_item}\n"
        f"Amount: {amount_str}\n"
        f"Category: {safe_category}\n"
        f"Payment: {html.escape(transaction.get('payment', 'Cash'))}"
    )
    
    notes = transaction.get('notes')
    if notes:
        response += f"\nNotes: {html.escape(str(notes)[:100])}{'...' if len(str(notes)) > 100 else ''}"
    
    return response

def create_backup():
    """Create backup of the Excel file"""
    try:
        BACKUP_DIR.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"backup_{timestamp}_{LOCAL_COPY_PATH.name}"
        shutil.copy2(LOCAL_COPY_PATH, backup_path)
        
        cleanup_old_backups()
        logger.info(f"Backup created: {backup_path}")
        return True
    except Exception as e:
        logger.error(f"Backup creation failed: {e}")
        return False

def cleanup_old_backups():
    """Remove backups older than retention period"""
    try:
        cutoff_date = datetime.now().timestamp() - (BACKUP_RETENTION_DAYS * 24 * 3600)
        
        for backup_file in BACKUP_DIR.glob("backup_*"):
            if backup_file.stat().st_mtime < cutoff_date:
                backup_file.unlink()
                logger.info(f"Removed old backup: {backup_file}")
    except Exception as e:
        logger.error(f"Backup cleanup failed: {e}")

@contextmanager
def excel_operation():
    """Context manager for Excel operations with openpyxl"""
    wb = None
    try:
        with excel_lock:
            if not EXCEL_SUPPORT:
                raise ImportError("Excel libraries are not available")
            
            # Check if file exists
            if not LOCAL_COPY_PATH.exists():
                raise FileNotFoundError(f"Excel file not found: {LOCAL_COPY_PATH}")
            
            # Open the workbook with openpyxl
            # Note: keep_vba=True is important for .xlsm files to preserve macros
            try:
                wb = load_workbook(str(LOCAL_COPY_PATH), keep_vba=True)
            except Exception as e:
                logger.error(f"Failed to open workbook: {e}")
                # Try without keep_vba as fallback
                wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
            
            yield wb
            
    except Exception as e:
        logger.error(f"Excel operation error: {e}")
        raise
    finally:
        if wb:
            try:
                # Save if there were changes
                wb.save(str(LOCAL_COPY_PATH))
                wb.close()
            except Exception as e:
                logger.error(f"Error saving/closing Excel: {e}")

async def async_copy_excel_from_onedrive() -> Tuple[bool, str]:
    """Async version of file copy"""
    loop = asyncio.get_event_loop()
    
    try:
        # Run the synchronous function in a thread
        with ThreadPoolExecutor() as executor:
            result = await loop.run_in_executor(
                executor, 
                copy_excel_from_onedrive
            )
        return result
    except Exception as e:
        logger.error(f"Async copy error: {e}")
        return False, f"❌ Async copy error: {str(e)}"

async def async_save_excel_to_onedrive() -> Tuple[bool, str]:
    """Async version of save"""
    loop = asyncio.get_event_loop()
    
    try:
        with ThreadPoolExecutor() as executor:
            result = await loop.run_in_executor(
                executor, 
                save_excel_to_onedrive
            )
        return result
    except Exception as e:
        logger.error(f"Async save error: {e}")
        return False, f"❌ Async save error: {str(e)}"

# ========== TEXT PROCESSING ==========

def normalize_text_for_matching(text: str) -> str:
    """Normalize text for matching"""
    if not text:
        return ""
    
    text = text.lower()
    text_without_parentheses = re.sub(r'\([^)]*\)', ' ', text)
    content_in_parentheses = re.findall(r'\(([^)]+)\)', text)
    all_text = text_without_parentheses + ' ' + ' '.join(content_in_parentheses)
    all_text = re.sub(r'[^\w\s]', ' ', all_text)
    all_text = ' '.join(all_text.split())
    
    return all_text.strip()

def extract_payment_amount_currency(text: str) -> Tuple[str, Dict[str, Optional[float]], str]:
    """
    Extract payment type, multiple amounts with currencies, and clean subcategory
    Now supports multiple currencies in one input
    
    Returns: (subcategory, currency_amounts_dict, payment_type)
    where currency_amounts_dict is {'USD': amount, 'LBP': amount, 'EURO': amount}
    """
    if not text:
        return "", {}, "Cash"
    
    original_text = text
    text_lower = text.lower()
    
    # 1. Extract payment type
    payment_type = "Cash"
    
    for keyword in CARD_KEYWORDS:
        if keyword in text_lower:
            payment_type = "Card"
            break
    
    if payment_type == "Cash":
        for keyword in DIGITAL_KEYWORDS:
            if keyword in text_lower:
                payment_type = "Digital Wallet"
                break
    
    # 2. IMPROVED MULTI-CURRENCY EXTRACTION
    currency_amounts = {}  # {'USD': 10.0, 'LBP': -150000.0}
    
    # Patterns for all currency types
    patterns = [
        # LBP patterns
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(lbp|ليرة|lira|ل\.ل)', 'LBP'),
        (r'(lbp|ليرة|lira|ل\.ل)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'LBP'),
        
        # USD patterns
        (r'[$\$](-?\d+(?:,\d{3})*(?:\.\d+)?)', 'USD'),
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(usd|dollar)', 'USD'),
        (r'(usd|dollar)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'USD'),
        
        # EURO patterns
        (r'€(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'EURO'),
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(eur|euro)', 'EURO'),
        (r'(eur|euro)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'EURO'),
        
        # Symbol suffix patterns
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)[€$]', None),  # Will be handled separately
    ]
    
    # Process each pattern
    for pattern, currency_code in patterns:
        try:
            matches = list(re.finditer(pattern, text_lower, re.IGNORECASE))
            for match in matches:
                if currency_code is None:
                    # Handle symbol suffix patterns
                    amount_str = match.group(1).replace(',', '')
                    symbol = match.group(0)[-1]
                    currency_code = 'USD' if symbol == '$' else 'EURO'
                elif currency_code == 'LBP':
                    if 'amount_first' in match.groupdict():
                        amount_str = match.group(1).replace(',', '')
                    else:
                        amount_str = match.group(2).replace(',', '')
                else:
                    if match.group(0)[0] in ['$', '€']:
                        amount_str = match.group(1).replace(',', '')
                    elif 'amount_first' in match.groupdict():
                        amount_str = match.group(1).replace(',', '')
                    else:
                        amount_str = match.group(2).replace(',', '')
                
                try:
                    amount = float(amount_str)
                    
                    # Check if amount starts with - in original text
                    match_text = match.group(0)
                    if match_text.startswith('-') and amount > 0:
                        amount = -amount
                    
                    # Store the amount for this currency
                    currency_amounts[currency_code] = amount
                    
                    logger.info(f"Found amount: {amount} {currency_code}")
                    
                except ValueError:
                    continue
                    
        except re.error:
            continue
    
    # 3. Also look for standalone negative numbers near currency words
    # This handles cases like "KSC 10$, -150000 LBP" where the negative is separate
    if not currency_amounts.get('LBP'):
        # Look for negative number followed by LBP
        lbp_pattern = r'-\s*(\d+(?:,\d{3})*)\s*(lbp|ليرة|lira|ل\.ل)'
        matches = re.findall(lbp_pattern, text_lower, re.IGNORECASE)
        for match in matches:
            try:
                amount = -float(match[0].replace(',', ''))
                currency_amounts['LBP'] = amount
                logger.info(f"Found negative LBP amount: {amount}")
            except ValueError:
                pass
    
    # 4. If we found some but not all currencies, look for standalone amounts
    if currency_amounts:
        # Look for standalone numbers that aren't already part of currency patterns
        standalone_pattern = r'(?<!\S)(-?\d+(?:,\d{3})*(?:\.\d+)?)(?![\$\€\w])'
        matches = re.finditer(standalone_pattern, text_lower)
        
        for match in matches:
            try:
                amount = float(match.group(1).replace(',', ''))
                # Don't add if we already have this amount (might be duplicate)
                # Check if this amount already exists in currency_amounts
                if amount not in currency_amounts.values():
                    # If it's a large whole number, assume LBP
                    if amount >= 1000 and amount.is_integer() and 'LBP' not in currency_amounts:
                        currency_amounts['LBP'] = amount
                        logger.info(f"Assumed LBP for standalone: {amount}")
                    # If it has decimals and no USD yet, assume USD
                    elif not amount.is_integer() and 'USD' not in currency_amounts:
                        currency_amounts['USD'] = amount
                        logger.info(f"Assumed USD for standalone: {amount}")
            except ValueError:
                continue
    
    # 5. Clean subcategory - remove all amount patterns
    subcategory = original_text
    
    # Remove all detected amounts
    for currency_code, amount in currency_amounts.items():
        abs_amount = abs(amount)
        
        if currency_code == 'LBP':
            # Remove LBP formats
            formats_to_remove = [
                f"{abs_amount:,.0f}LBP", f"{abs_amount}LBP",
                f"LBP{abs_amount:,.0f}", f"LBP{abs_amount}",
                f"{abs_amount:,.0f} lbp", f"{abs_amount} lbp",
                f"lbp {abs_amount:,.0f}", f"lbp {abs_amount}",
                f"{abs_amount:,.0f}ليرة", f"{abs_amount}ليرة",
                f"ليرة{abs_amount:,.0f}", f"ليرة{abs_amount}",
                f"{abs_amount:,.0f} ليرة", f"{abs_amount} ليرة",
                f"ليرة {abs_amount:,.0f}", f"ليرة {abs_amount}",
            ]
        elif currency_code == 'USD':
            # Remove USD formats
            formats_to_remove = [
                f"${abs_amount:.2f}", f"${abs_amount:,.2f}",
                f"{abs_amount:.2f}$", f"{abs_amount:,.2f}$",
                f"{abs_amount:.2f}usd", f"{abs_amount:.2f} dollar",
                f"{abs_amount}usd", f"{abs_amount} dollar",
            ]
        elif currency_code == 'EURO':
            # Remove EURO formats
            formats_to_remove = [
                f"€{abs_amount:.2f}", f"€{abs_amount:,.2f}",
                f"{abs_amount:.2f}€", f"{abs_amount:,.2f}€",
                f"{abs_amount:.2f}eur", f"{abs_amount:.2f} euro",
                f"{abs_amount}eur", f"{abs_amount} euro",
            ]
        else:
            formats_to_remove = []
        
        # Add negative versions
        negative_formats = [f"-{fmt}" for fmt in formats_to_remove]
        formats_to_remove.extend(negative_formats)
        
        # Remove all formats
        for fmt in formats_to_remove:
            subcategory = subcategory.replace(fmt, ' ')
            subcategory = subcategory.replace(fmt.lower(), ' ')
            subcategory = subcategory.replace(fmt.upper(), ' ')
        
        # Also remove just the number (with and without commas)
        subcategory = subcategory.replace(str(abs_amount), ' ')
        if abs_amount >= 1000:
            # Remove with commas
            subcategory = subcategory.replace(f"{abs_amount:,.0f}", ' ')
            # Remove without commas
            subcategory = subcategory.replace(f"{int(abs_amount)}", ' ')
    
    # Remove standalone currency symbols
    subcategory = subcategory.replace('$', ' ').replace('€', ' ')
    
    # Remove payment keywords
    for keyword in CARD_KEYWORDS + DIGITAL_KEYWORDS:
        subcategory = subcategory.replace(keyword, ' ').replace(keyword.title(), ' ')
    
    # Remove currency words
    currency_words = ['usd', 'dollar', 'euro', 'eur', 'lbp', 'lira', 'ليرة', 'ل.ل']
    for word in currency_words:
        subcategory = subcategory.replace(word, ' ').replace(word.title(), ' ')
        subcategory = subcategory.replace(word.upper(), ' ')
    
    # Clean up
    subcategory = ' '.join(subcategory.split())
    subcategory = subcategory.title()
    
    if not subcategory:
        subcategory = original_text
    
    logger.info(f"Parsed: subcategory='{subcategory}', amounts={currency_amounts}, payment={payment_type}")
    return subcategory, currency_amounts, payment_type

# ========== EXCEL FUNCTIONS ==========

def copy_excel_from_onedrive() -> Tuple[bool, str]:
    """Simple version that doesn't delete the source file"""
    try:
        if not ONEDRIVE_PATH.exists():
            return False, f"❌ OneDrive file not found: {ONEDRIVE_PATH}"
        
        # If LOCAL_COPY_PATH doesn't exist, create it
        if not LOCAL_COPY_PATH.exists():
            shutil.copy2(ONEDRIVE_PATH, LOCAL_COPY_PATH)
            return True, "✅ Created local copy"
        
        # If they're the same file, just return success
        if str(ONEDRIVE_PATH) == str(LOCAL_COPY_PATH):
            return True, "✅ Using existing file"
        
        # Otherwise, update the local copy
        shutil.copy2(ONEDRIVE_PATH, LOCAL_COPY_PATH)
        return True, "✅ Updated local copy"
        
    except Exception as e:
        return False, f"❌ Copy error: {str(e)}"

def save_excel_to_onedrive() -> Tuple[bool, str]:
    """Save changes back to OneDrive - now supports API"""
    # Check if we should use OneDrive API
    use_api = os.getenv("USE_ONEDRIVE_API", "false").lower() == "true"
    
    if use_api and ONEDRIVE_API_AVAILABLE:
        # Use OneDrive API
        return upload_to_onedrive_api()
    
    # Fall back to local file copy (original logic)
    try:
        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local copy not found"
        
        create_backup()
        shutil.copy2(LOCAL_COPY_PATH, ONEDRIVE_PATH)
        return True, "✅ Changes saved to OneDrive"
    except Exception as e:
        return False, f"❌ Save error: {str(e)}"

def load_all_tables_with_details() -> Dict[str, Dict]:
    """Load ALL Excel Tables with detailed information and caching"""
    global _table_cache, _table_cache_timestamp
    
    if not EXCEL_SUPPORT:
        logger.error("Excel libraries are not available. Cannot load tables.")
        return {}
    
    try:
        if LOCAL_COPY_PATH.exists():
            current_mtime = LOCAL_COPY_PATH.stat().st_mtime
            if (_table_cache_timestamp and 
                _table_cache_timestamp == current_mtime and 
                _table_cache):
                return _table_cache
        
        # Open workbook in read-only mode for performance
        wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True, read_only=True)
        
        # Check if the dropdown sheet exists
        if DROPDOWN_SHEET_NAME not in wb.sheetnames:
            logger.error(f"Sheet '{DROPDOWN_SHEET_NAME}' not found in workbook")
            wb.close()
            return {}
        
        sheet = wb[DROPDOWN_SHEET_NAME]
        
        tables_dict = {}
        current_category = None
        current_items = []
        
        # Read all rows until we hit an empty row
        for row in sheet.iter_rows(min_row=1, max_row=1000, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            
            # Look for category headers - adjust this logic based on your Excel structure
            first_cell = row[0]  # Column A
            
            if first_cell:
                first_cell_str = str(first_cell).strip()
                
                # Heuristic: Categories might be in uppercase, bold, or have specific formatting
                # You might need to adjust this based on your actual Excel structure
                if (first_cell_str.isupper() or 
                    first_cell_str.endswith(':') or
                    len(first_cell_str) < 20 and not first_cell_str[0].islower()):
                    
                    # Save previous category if we have one
                    if current_category and current_items:
                        variations_dict = {}
                        for item in current_items:
                            original_text = str(item).strip()
                            if original_text:
                                normalized = normalize_text_for_matching(original_text)
                                variations_dict[normalized] = original_text
                                
                                # Also store individual words for better matching
                                words = set(normalized.split())
                                for word in words:
                                    if len(word) > 3:  # Only store words longer than 3 characters
                                        if word not in variations_dict:
                                            variations_dict[word] = original_text
                        
                        tables_dict[current_category] = {
                            'original_name': current_category,
                            'subcategories': current_items,
                            'variations': variations_dict,
                            'count': len(current_items)
                        }
                    
                    # Start new category
                    current_category = first_cell_str
                    current_items = []
                elif current_category and first_cell_str:
                    # Add item to current category
                    current_items.append(first_cell_str)
        
        # Don't forget the last category
        if current_category and current_items:
            variations_dict = {}
            for item in current_items:
                original_text = str(item).strip()
                if original_text:
                    normalized = normalize_text_for_matching(original_text)
                    variations_dict[normalized] = original_text
                    
                    words = set(normalized.split())
                    for word in words:
                        if len(word) > 3:
                            if word not in variations_dict:
                                variations_dict[word] = original_text
            
            tables_dict[current_category] = {
                'original_name': current_category,
                'subcategories': current_items,
                'variations': variations_dict,
                'count': len(current_items)
            }
        
        wb.close()
        
        _table_cache = tables_dict
        _table_cache_timestamp = current_mtime if LOCAL_COPY_PATH.exists() else None
        
        logger.info(f"Loaded {len(tables_dict)} categories from Excel")
        return tables_dict
        
    except Exception as e:
        logger.error(f"Error loading tables: {str(e)}", exc_info=True)
        return {}

def find_best_match_for_input(input_text: str, tables_dict: Dict) -> Tuple[str, str, str, float]:
    """Find the best match for input text with confidence score"""
    if not input_text:
        return "", "Unknown", "Empty input", 0.0
    
    normalized_input = normalize_text_for_matching(input_text)
    
    if not normalized_input:
        return "", "Unknown", "Empty after normalization", 0.0
    
    # Reject very short inputs
    if len(normalized_input) < 3:
        return "", "Unknown", "Input too short", 0.0
    
    # Reject common non-category words
    if normalized_input in COMMON_NON_CATEGORIES:
        return "", "Unknown", "Common non-category word", 0.0
    
    best_match = None
    best_category = None
    best_score = 0.0
    match_type = "No match"
    
    # SIMPLIFY: Just split by spaces, no complex regex
    input_words = set(normalized_input.split())
    
    for category, data in tables_dict.items():
        for variation, original in data['variations'].items():
            # SIMPLIFY: Split variation by spaces
            variation_words = set(variation.split())
            
            # Calculate intersection
            intersection = len(input_words.intersection(variation_words))
            union = len(input_words.union(variation_words))
            
            if union > 0:
                score = intersection / union
                
                # Boost for exact matches
                if normalized_input == variation:
                    score = 1.0
                elif normalized_input in variation:
                    score = max(score, 0.8)
                elif variation in normalized_input:
                    score = max(score, 0.8)
                
                if score > best_score:
                    best_score = score
                    best_match = original
                    best_category = category
    
    # Check if we have a good enough match
    if best_score >= MINIMUM_CONFIDENCE:
        if best_score >= 0.95:
            match_type = "Exact match"
        elif best_score >= 0.85:
            match_type = "Strong match"
        elif best_score >= MINIMUM_CONFIDENCE:
            match_type = "Good match"
        
        logger.info(f"Good match found: '{input_text}' → '{best_match}' (score: {best_score:.2f})")
        return best_match, best_category, match_type, best_score
    
    # Try fuzzy matching with simplified approach
    if best_score < FUZZY_MATCH_THRESHOLD:
        # Build simple lists for fuzzy matching
        all_variations = []
        all_originals = []
        all_categories = []
        
        for category, data in tables_dict.items():
            for variation, original in data['variations'].items():
                all_variations.append(variation)
                all_originals.append(original)
                all_categories.append(category)
        
        # Use simple fuzzy matching
        matches = get_close_matches(normalized_input, all_variations, n=1, cutoff=0.5)
        
        if matches:
            best_variation = matches[0]
            idx = all_variations.index(best_variation)
            fuzzy_score = 0.5  # Fixed score for fuzzy matches
            if fuzzy_score >= MINIMUM_CONFIDENCE:
                return all_originals[idx], all_categories[idx], "Fuzzy match", fuzzy_score
    
    logger.info(f"No good match found for '{input_text}'. Best score: {best_score:.2f}")
    return "", "Unknown", "No good match found", best_score

# ========== TRANSACTION FUNCTIONS ==========

def add_transaction_smart(subcategory_input: str, currency_amounts: Dict[str, Optional[float]], 
                         payment_type: str, optional_notes: str = "") -> Tuple[bool, str]:
    """Add transaction with smart matching - supports optional notes"""
    try:
        # FIXED: Always use "Expenses" as transaction type (even for negative amounts)
        transaction_type = "Expenses"
        
        logger.info(f"Processing transaction: '{subcategory_input}', amounts: {currency_amounts}, payment: {payment_type}, notes: '{optional_notes}'")
        
        # 1. Copy fresh file
        success, msg = copy_excel_from_onedrive()
        if not success:
            logger.error(f"Failed to copy file: {msg}")
            return False, msg
        
        # 2. Load tables and find best match
        tables_dict = load_all_tables_with_details()
        if not tables_dict:
            logger.error("No tables found in Dropdown Data sheet")
            return False, "❌ No tables found in Dropdown Data sheet"
        
        # Find best match with confidence
        matched_original, category, match_type, confidence = find_best_match_for_input(subcategory_input, tables_dict)
        
        # FIXED: Display confidence as percentage (e.g., 70% instead of 0.7)
        confidence_percentage = confidence * 100
        
        # REJECTION LOGIC: If confidence is below threshold, reject the transaction
        if confidence < MINIMUM_CONFIDENCE:
            # Generate helpful suggestions
            suggestions = []
            
            # Get top suggestions from all categories
            all_suggestions = []
            for cat_name, data in tables_dict.items():
                for item in data['subcategories'][:5]:  # Top 5 from each category
                    all_suggestions.append((cat_name, item))
            
            # Show most relevant suggestions (alphabetically sorted)
            all_suggestions.sort(key=lambda x: x[1])
            
            suggestions.append("\n<b>Available categories:</b>")
            for cat_name, _ in tables_dict.items():
                suggestions.append(f"  • {cat_name}")
            
            suggestions.append("\n<b>Common items:</b>")
            displayed_items = 0
            for cat_name, item in all_suggestions:
                if displayed_items >= 8:  # Limit to 8 items
                    break
                suggestions.append(f"  • {item} ({cat_name})")
                displayed_items += 1
            
            suggestions_text = "\n".join(suggestions)
            
            rejection_msg = (
                f"❌ <b>Transaction Rejected:</b>\n\n"
                f"<b>Input:</b> <code>{subcategory_input}</code>\n"
                f"<b>Confidence:</b> {confidence_percentage:.1f}% (minimum required: {MINIMUM_CONFIDENCE*100:.0f}%)\n"
                f"<b>Reason:</b> {match_type}\n\n"
                f"<b>No good match found. Did you mean one of these?</b>\n"
                f"{suggestions_text}\n\n"
                f"<i>Tip: Use more specific terms or check spelling.</i>"
            )
            logger.warning(f"Transaction rejected: {subcategory_input} (confidence: {confidence_percentage:.1f}%)")
            return False, rejection_msg
        
        logger.info(f"Match found: '{subcategory_input}' → '{matched_original}' → '{category}' ({match_type}, {confidence_percentage:.1f}%)")
        
        # 3. Open Excel and add transaction
        if not EXCEL_SUPPORT:
            return False, "❌ Excel libraries not installed. Please install: pip install openpyxl"
        
        with excel_operation() as wb:
            sheet = wb[TRACKING_SHEET_NAME]
            
            # Find first empty row in column C (date column)
            row = 12
            while sheet.cell(row=row, column=3).value not in [None, ""]:
                row += 1
                if row > 1000:
                    break
            
            logger.debug(f"Adding transaction at row {row}")
            
            # Add data
            today = datetime.now().strftime("%d-%b-%y")
            
            sheet.cell(row=row, column=3, value=today)  # Column C
            sheet.cell(row=row, column=4, value=payment_type)  # Column D
            sheet.cell(row=row, column=5, value=transaction_type)  # Column E
            sheet.cell(row=row, column=6, value=category)  # Column F
            sheet.cell(row=row, column=7, value=matched_original)  # Column G
            
            # Add amounts to correct currency columns
            for currency, amount in currency_amounts.items():
                if currency in CURRENCIES and amount is not None:
                    column_letter = CURRENCIES[currency].column  # 'H', 'I', 'J'
                    column_index = column_index_from_string(column_letter)
                    sheet.cell(row=row, column=column_index, value=float(amount))
            
            # If no currencies specified, add to USD as default
            if not currency_amounts:
                sheet.cell(row=row, column=8, value=0.0)  # Column H
            
            # Add optional notes
            if optional_notes:
                clean_notes = optional_notes.strip()
                if len(clean_notes) > 500:
                    clean_notes = clean_notes[:497] + "..."
                sheet.cell(row=row, column=11, value=clean_notes)  # Column K
            
            logger.debug("Transaction saved to local file")
        # Workbook is saved automatically by the context manager
        
        # 4. Save back to OneDrive
        save_success, save_msg = save_excel_to_onedrive()
        
        # Format response
        amount_display_parts = []
        total_amounts = len(currency_amounts)
        
        for currency, amount in currency_amounts.items():
            if amount is not None:
                amount_str = format_currency_amount(amount, currency)
                # Add visual indicator for negative amounts
                if amount < 0:
                    amount_display = f"🔄 {amount_str}"
                else:
                    amount_display = amount_str
                amount_display_parts.append(amount_display)
        
        if amount_display_parts:
            if total_amounts > 1:
                amount_display = " + ".join(amount_display_parts)
            else:
                amount_display = amount_display_parts[0]
        else:
            amount_display = "No amount"
        
        payment_emoji = PAYMENT_TYPES.get(payment_type, "💵")
        
        # Check if any amounts are negative
        has_negative = any(a is not None and a < 0 for a in currency_amounts.values())
        
        if has_negative:
            type_emoji = "🔄"  # Circular arrow for refund/income
            transaction_desc = "Expense (with refunds/corrections)"
        else:
            type_emoji = "📤"  # Outgoing arrow for expense
            transaction_desc = "Expense"
        
        # Add currency emojis for each currency
        currency_emojis = []
        for currency in currency_amounts.keys():
            if currency == "USD":
                currency_emojis.append("💵")
            elif currency == "LBP":
                currency_emojis.append("🇱🇧")
            elif currency == "EURO":
                currency_emojis.append("💶")
        
        currency_emoji_str = " ".join(currency_emojis)
        
        import html
        safe_input = html.escape(subcategory_input)
        safe_matched = html.escape(matched_original)
        safe_category = html.escape(category)
        safe_payment = html.escape(payment_type)
        safe_notes = html.escape(optional_notes) if optional_notes else ""
        safe_confidence = f"{confidence_percentage:.1f}%"
        
        # Add multi-currency indicator
        multi_currency_indicator = "🌐 " if total_amounts > 1 else ""
        
        # Add notes to response if provided
        notes_section = f"\n• <b>Notes:</b> {safe_notes}" if optional_notes else ""
        
        if save_success:
            message = (
                f"{multi_currency_indicator}✅ {type_emoji} {payment_emoji} {currency_emoji_str} <b>{transaction_desc} Added:</b>\n\n"
                f"• <b>You typed:</b> <code>{safe_input}</code>\n"
                f"• <b>Recorded as:</b> <code>{safe_matched}</code>\n"
                f"• <b>Category:</b> {safe_category}\n"
                f"• <b>Amount{'s' if total_amounts > 1 else ''}:</b> {amount_display}\n"
                f"• <b>Payment:</b> {safe_payment}"
                f"{notes_section}\n"
                f"• <b>Confidence:</b> {safe_confidence}\n"
                f"• <b>Match type:</b> {match_type}"
            )
            logger.info(f"Transaction added successfully: {transaction_desc}")
            return True, message
        else:
            message = (
                f"{multi_currency_indicator}⚠️ {type_emoji} {payment_emoji} {currency_emoji_str} <b>{transaction_desc} Added Locally:</b>\n\n"
                f"• <b>You typed:</b> <code>{safe_input}</code>\n"
                f"• <b>Recorded as:</b> <code>{safe_matched}</code>\n"
                f"• <b>Category:</b> {safe_category}\n"
                f"• <b>Amount{'s' if total_amounts > 1 else ''}:</b> {amount_display}\n"
                f"• <b>Payment:</b> {safe_payment}"
                f"{notes_section}\n\n"
                f"<b>⚠️ OneDrive Sync Failed:</b>\n{save_msg}"
            )
            logger.warning(f"Transaction added locally but OneDrive sync failed: {save_msg}")
            return True, message
        
    except Exception as e:
        logger.error(f"Error in add_transaction_smart: {str(e)}", exc_info=True)
        import html
        error_msg = html.escape(str(e)[:200])
        return False, f"❌ <b>Error:</b> {error_msg}"

def delete_transaction_at_row(row: int) -> Tuple[bool, str, Optional[Dict]]:
    """Delete a transaction at a specific row in the Excel file"""
    try:
        if not EXCEL_SUPPORT:
            return False, "❌ Excel libraries not installed", None
        
        # 1. Copy fresh file
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}", None
        
        with excel_operation() as wb:
            sheet = wb[TRACKING_SHEET_NAME]
            
            # 2. Validate row number
            if row < 12 or row > 1000:
                return False, f"❌ Invalid row number: {row}. Must be between 12 and 1000.", None
            
            # 3. Check if the row has data
            if sheet.cell(row=row, column=3).value is None:
                return False, f"❌ No transaction found at row {row}", None
            
            # 4. Get the transaction to be deleted (for confirmation message)
            deleted_transaction = {
                'row': row,
                'date': sheet.cell(row=row, column=3).value,
                'payment': sheet.cell(row=row, column=4).value or "Cash",
                'type': sheet.cell(row=row, column=5).value or "Expenses",
                'category': sheet.cell(row=row, column=6).value or "",
                'subcategory': sheet.cell(row=row, column=7).value or "",
                'usd': sheet.cell(row=row, column=8).value,
                'lbp': sheet.cell(row=row, column=9).value,
                'euro': sheet.cell(row=row, column=10).value,
                'notes': sheet.cell(row=row, column=11).value or ""
            }
            
            # 5. Clear the row
            for col in range(3, 12):  # Columns C to K
                sheet.cell(row=row, column=col, value=None)
            
            logger.info(f"Deleted transaction at row {row}")
        
        # 6. Save back to OneDrive
        save_success, save_msg = save_excel_to_onedrive()
        
        if save_success:
            return True, "✅ Transaction deleted successfully!", deleted_transaction
        else:
            return True, f"⚠️ Transaction deleted locally, but OneDrive sync failed:\n{save_msg}", deleted_transaction
        
    except Exception as e:
        logger.error(f"Error deleting transaction at row {row}: {str(e)}", exc_info=True)
        return False, f"❌ Error deleting transaction: {str(e)[:200]}", None

def delete_last_transaction() -> Tuple[bool, str, Optional[Dict]]:
    """Delete the last transaction from the Excel file (for backward compatibility)"""
    if not EXCEL_SUPPORT:
        return False, "❌ Excel libraries not installed", None
    
    try:
        # 1. Copy fresh file
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}", None
        
        with excel_operation() as wb:
            sheet = wb[TRACKING_SHEET_NAME]
            
            # 2. Find the last row with data
            last_row = 12
            while sheet.cell(row=last_row, column=3).value not in [None, ""]:
                last_row += 1
                if last_row > 1000:  # Safety limit
                    break
            
            # 3. Check if there are any transactions
            if last_row == 12:
                return False, "❌ No transactions found to delete", None
            
            # 4. Get the transaction to be deleted (for confirmation message)
            row_to_delete = last_row - 1  # Last row with data
        
        # Use the new function to delete at row
        return delete_transaction_at_row(row_to_delete)
        
    except Exception as e:
        logger.error(f"Error finding last transaction: {str(e)}", exc_info=True)
        return False, f"❌ Error: {str(e)[:200]}", None

def modify_transaction_at_row(row: int, 
                             new_subcategory: str = None, 
                             new_currency_amounts: Dict[str, Optional[float]] = None,
                             new_payment_type: str = None, 
                             new_notes: str = None) -> Tuple[bool, str, Dict]:
    """Modify specific fields of a transaction at row"""
    try:
        if not EXCEL_SUPPORT:
            return False, "❌ Excel libraries not installed", None
        
        # Validate row number
        if row < 12 or row > 1000:
            return False, f"❌ Invalid row number: {row}", None
        
        # Check if anything to modify
        if not any([new_subcategory, new_currency_amounts, new_payment_type is not None, new_notes is not None]):
            return False, "❌ Nothing to modify. Specify at least one field to change.", None
        
        # 1. Copy fresh file
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}", None
        
        original_transaction = None
        modified_transaction = None
        
        with excel_operation() as wb:
            sheet = wb[TRACKING_SHEET_NAME]
            
            # Check if the row has data
            if sheet.cell(row=row, column=3).value is None:
                return False, f"❌ No transaction found at row {row}", None
            
            # Store original transaction details
            original_transaction = {
                'row': row,
                'date': sheet.cell(row=row, column=3).value,
                'payment': sheet.cell(row=row, column=4).value or "Cash",
                'type': sheet.cell(row=row, column=5).value or "Expenses",
                'category': sheet.cell(row=row, column=6).value or "",
                'subcategory': sheet.cell(row=row, column=7).value or "",
                'usd': sheet.cell(row=row, column=8).value,
                'lbp': sheet.cell(row=row, column=9).value,
                'euro': sheet.cell(row=row, column=10).value,
                'notes': sheet.cell(row=row, column=11).value or ""
            }
            
            modified_transaction = original_transaction.copy()
            changes_made = []
            
            # 2. Modify subcategory if provided (with smart matching)
            if new_subcategory:
                tables_dict = load_all_tables_with_details()
                if not tables_dict:
                    return False, "❌ No tables found in Dropdown Data sheet", None
                
                # Find best match for new subcategory
                matched_original, category, match_type, confidence = find_best_match_for_input(new_subcategory, tables_dict)
                confidence_percentage = confidence * 100
                
                # REJECTION LOGIC: If confidence is below threshold
                if confidence < MINIMUM_CONFIDENCE:
                    suggestions = []
                    for cat_name, data in tables_dict.items():
                        for item in data['subcategories'][:3]:
                            suggestions.append((cat_name, item))
                    
                    suggestions.sort(key=lambda x: x[1])
                    suggestions_text = "\n".join([f"  • {item} ({cat})" for cat, item in suggestions[:6]])
                    
                    rejection_msg = (
                        f"❌ <b>Modification Rejected:</b>\n\n"
                        f"<b>New item:</b> <code>{new_subcategory}</code>\n"
                        f"<b>Confidence:</b> {confidence_percentage:.1f}% (minimum required: {MINIMUM_CONFIDENCE*100:.0f}%)\n"
                        f"<b>Reason:</b> {match_type}\n\n"
                        f"<b>No good match found. Try:</b>\n"
                        f"{suggestions_text}"
                    )
                    return False, rejection_msg, original_transaction
                
                # Update subcategory and category
                sheet.cell(row=row, column=6, value=category)
                sheet.cell(row=row, column=7, value=matched_original)
                modified_transaction['category'] = category
                modified_transaction['subcategory'] = matched_original
                changes_made.append(f"Item: '{original_transaction['subcategory']}' → '{matched_original}'")
            
            # 3. Modify payment type if provided
            if new_payment_type is not None:
                sheet.cell(row=row, column=4, value=new_payment_type)
                modified_transaction['payment'] = new_payment_type
                changes_made.append(f"Payment: '{original_transaction['payment']}' → '{new_payment_type}'")
            
            # 4. Modify currency amounts if provided
            if new_currency_amounts:
                # Clear all currency columns first if we're setting new amounts
                for col_letter in ['H', 'I', 'J']:
                    col_index = column_index_from_string(col_letter)
                    sheet.cell(row=row, column=col_index, value=None)
                
                # Update currency amounts
                for currency, amount in new_currency_amounts.items():
                    if currency in CURRENCIES and amount is not None:
                        column_letter = CURRENCIES[currency].column
                        column_index = column_index_from_string(column_letter)
                        sheet.cell(row=row, column=column_index, value=float(amount))
                
                # Update modified transaction
                modified_transaction['usd'] = new_currency_amounts.get('USD')
                modified_transaction['lbp'] = new_currency_amounts.get('LBP')
                modified_transaction['euro'] = new_currency_amounts.get('EURO')
                
                # Format amount changes
                old_amounts = []
                new_amounts = []
                
                for curr in ['USD', 'LBP', 'EURO']:
                    old_val = original_transaction.get(curr.lower())
                    new_val = modified_transaction.get(curr.lower())
                    if old_val is not None or new_val is not None:
                        old_str = format_currency_amount(old_val, curr) if old_val is not None else "None"
                        new_str = format_currency_amount(new_val, curr) if new_val is not None else "None"
                        old_amounts.append(old_str)
                        new_amounts.append(new_str)
                
                if old_amounts or new_amounts:
                    changes_made.append(f"Amount: {' + '.join(old_amounts) if old_amounts else 'None'} → {' + '.join(new_amounts) if new_amounts else 'None'}")
            
            # 5. Modify notes if provided
            if new_notes is not None:
                if new_notes.strip():
                    clean_notes = new_notes.strip()
                    if len(clean_notes) > 500:
                        clean_notes = clean_notes[:497] + "..."
                    sheet.cell(row=row, column=11, value=clean_notes)
                    modified_transaction['notes'] = clean_notes
                    
                    old_notes_preview = original_transaction['notes'][:50] + "..." if original_transaction['notes'] and len(original_transaction['notes']) > 50 else original_transaction['notes'] or "None"
                    new_notes_preview = clean_notes[:50] + "..." if len(clean_notes) > 50 else clean_notes
                    changes_made.append(f"Notes: '{old_notes_preview}' → '{new_notes_preview}'")
                else:
                    sheet.cell(row=row, column=11, value=None)
                    modified_transaction['notes'] = None
                    changes_made.append(f"Notes: '{original_transaction['notes'] or 'None'}' → 'None'")
            
            # 6. Add modification note to existing notes
            if changes_made:
                existing_notes = sheet.cell(row=row, column=11).value or ""
                modification_note = f"\n\n[Modified {datetime.now().strftime('%Y-%m-%d %H:%M')}: {', '.join(changes_made)}]"
                
                # Only add modification note if not already there
                if "[Modified" not in existing_notes:
                    new_notes_value = existing_notes + modification_note
                    if len(new_notes_value) > 500:
                        # Keep the modification note but truncate earlier content
                        new_notes_value = new_notes_value[-500:]
                    sheet.cell(row=row, column=11, value=new_notes_value)
            
            logger.info(f"Modified transaction at row {row}. Changes: {changes_made}")
        
        # 7. Save back to OneDrive
        save_success, save_msg = save_excel_to_onedrive()
        
        if save_success:
            success_msg = f"✅ Transaction modified successfully!\n\n<b>Changes made:</b>\n" + "\n".join([f"• {change}" for change in changes_made])
            return True, success_msg, modified_transaction
        else:
            warning_msg = f"⚠️ Transaction modified locally, but OneDrive sync failed:\n{save_msg}\n\n<b>Changes made:</b>\n" + "\n".join([f"• {change}" for change in changes_made])
            return True, warning_msg, modified_transaction
        
    except Exception as e:
        logger.error(f"Error modifying transaction: {str(e)}", exc_info=True)
        return False, f"❌ Error modifying transaction: {str(e)[:200]}", None

def get_recent_transactions(count: int = 10) -> List[Dict]:
    """Get multiple recent transactions from Excel"""
    transactions = []
    
    if not EXCEL_SUPPORT:
        return transactions
    
    try:
        # Copy fresh file first
        copy_excel_from_onedrive()
        
        with excel_operation() as wb:
            sheet = wb[TRACKING_SHEET_NAME]
            
            # Find the last row with data in column C
            last_row = 12
            while sheet.cell(row=last_row, column=3).value not in [None, ""]:
                last_row += 1
                if last_row > 1000:
                    break
            
            # Get the last 'count' transactions
            start_row = max(12, last_row - count)
            
            for row in range(start_row, last_row):
                date_val = sheet.cell(row=row, column=3).value
                if date_val:
                    transaction = {
                        'row': row,
                        'date': date_val,
                        'payment': sheet.cell(row=row, column=4).value or "Cash",
                        'type': sheet.cell(row=row, column=5).value or "Expenses",
                        'category': sheet.cell(row=row, column=6).value or "",
                        'subcategory': sheet.cell(row=row, column=7).value or "",
                        'usd': sheet.cell(row=row, column=8).value,
                        'lbp': sheet.cell(row=row, column=9).value,
                        'euro': sheet.cell(row=row, column=10).value,
                        'notes': sheet.cell(row=row, column=11).value or ""
                    }
                    transactions.append(transaction)
        
        # Return in reverse order (most recent first)
        return list(reversed(transactions))
        
    except Exception as e:
        logger.error(f"Error getting recent transactions: {str(e)}")
        return []

# ========== TELEGRAM BOT FUNCTIONS ==========

async def start_command(update, context):
    """Handle /start - UPDATED with enhanced delete feature"""
    logger.info(f"Received /start from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    excel_status = "✅ INSTALLED" if EXCEL_SUPPORT else "❌ NOT INSTALLED"
    
    await update.message.reply_text(
        f"💰 <b>Smart Budget Tracker Bot</b>\n\n"
        f"📊 <b>Status:</b> {excel_status}\n"
        f"🎯 <b>Minimum Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}&#37;\n\n" # Fixed: Changed % to &#37;
        
        "✅ <b>Smart Matching Examples:</b>\n"
        "• <code>Chamsin</code> → Bakery Products (Chamsin)\n"
        "• <code>Mazda</code> → Fuel (Mazda)\n"
        "• <code>KSC</code> → KSC (exact match)\n"
        "• <code>Fuel Opel</code> → Fuel (Opel)\n\n"
        
        "🔄 <b>Negative Amounts (Refunds/Corrections):</b>\n"
        "• <code>Chamsin -10$</code> → Refund from Chamsin\n"
        "• <code>Refund -200000 LBP</code> → Refund (no commas needed)\n"
        "• <code>Correction -50€</code> → Negative adjustment\n\n"
        
        "💱 <b>LBP Amounts:</b>\n"
        "• <code>200000 LBP</code> → Works without commas\n"
        "• <code>200,000 LBP</code> → Also works with commas\n\n"
        
        "❌ <b>Will Reject:</b>\n"
        f"• <code>OKAY</code>, <code>Test</code>, <code>Hello</code>\n"
        f"• Poor matches (&lt;{MINIMUM_CONFIDENCE*100:.0f}&#37; confidence)\n\n"  # Fixed: Changed < to &lt; and % to &#37;
        
        "<b>📝 How to Use:</b>\n"
        "Just send: <code>ITEM AMOUNT CURRENCY PAYMENT</code>\n\n"
        "<b>Examples:</b>\n"
        "• <code>Chamsin 10</code> (Expense)\n"
        "• <code>Chamsin -10$</code> (Refund)\n"
        "• <code>Fuel Mazda 200000 LBP</code> (No commas needed)\n"
        "• <code>Refund -500$ card</code> (Negative amount)\n"
        "• <code>KSC 15.50 Card</code>\n"
        "• <code>Daouk Sweets 20$ Card</code>\n\n"
        
        "<b>🗑️ ENHANCED DELETE FEATURE:</b>\n"
        "Now you can delete any of the last 10 transactions!\n"
        "• <code>/delete</code> - See list and select\n"
        "• <code>/delete 3</code> - Delete transaction #3\n"
        "• <code>/delete confirm</code> - Delete last transaction\n\n"
        
        "<b>✏️ ENHANCED MODIFY FEATURE:</b>\n"
        "Now you can modify any of the last 10 transactions!\n"
        "• <code>/modify</code> - See list and select\n"
        "• <code>/modify 3</code> - Select transaction #3\n"
        "• <code>/modify 3 20$</code> - Change amount to $20\n"
        "• <code>/modify 3 to card</code> - Change payment type\n\n"
        
        "<b>📥 DOWNLOAD FEATURES:</b>\n"
        "• /download csv [day/week/month/year/all] - Export to CSV\n"
        "• /download summary - Export summary JSON\n"
        "• /download backup - Download Excel backup\n\n"
        
        "<b>⚙️ Commands:</b>\n"
        "/start - This message\n"
        "/help - Detailed help\n"
        "/testparse [text] - Test parsing & matching\n"
        "/recent - Last 10 transactions\n"
        "/stats - Statistics\n"
        "/save - Force save to OneDrive\n"
        "/clearcache - Clear matching cache\n"
        "/download - Download data files\n"
        "/delete - Delete any transaction (with selection)\n"
        "/modify - Modify any transaction\n\n"
        
        "💡 <b>Tip:</b> Tap the menu icon (/) to see all commands",
        parse_mode='HTML'
    )

async def delete_command(update, context):
    """Delete a transaction - enhanced to allow deleting any of last 10 transactions"""
    logger.info(f"Received /delete from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("\u26d4 Unauthorized.")
        return
    
    # SCENARIO 1: No arguments - show recent transactions for selection
    if not context.args:
        await show_transaction_selection_for_delete(update, context)
        return
    
    args = context.args
    
    # SCENARIO 2: Direct deletion with row number and confirm (e.g., /delete 83 confirm)
    if len(args) >= 2 and args[0].isdigit() and args[1].lower() == "confirm":
        row_number = int(args[0])
        
        # Get recent transactions to validate this is one of them
        recent_transactions = get_recent_transactions(10)
        
        # Check if this row number is in recent transactions
        is_recent = any(t.get('row') == row_number for t in recent_transactions)
        
        if is_recent:
            await update.message.reply_text(f"\u1f5d1\ufe0f Deleting transaction at row {row_number}...")
            success, message, deleted_transaction = delete_transaction_at_row(row_number)
            await send_deletion_result(update, deleted_transaction, success, message)
        else:
            await update.message.reply_text(
                f"\u274c Row {row_number} is not in the recent 10 transactions.\n"
                f"Use <code>/delete</code> to see recent transactions first."
            )
        return
    
    # SCENARIO 3: Direct deletion with number only (e.g., /delete 1)
    if args[0].isdigit():
        transaction_num = int(args[0])
        
        # Get recent transactions
        recent_transactions = get_recent_transactions(10)
        
        if not recent_transactions:
            await update.message.reply_text("\u274c No transactions found.")
            return
        
        if transaction_num < 1 or transaction_num > len(recent_transactions):
            await update.message.reply_text(
                f"\u274c Invalid selection. Choose 1-{len(recent_transactions)}"
            )
            return
        
        # Get the selected transaction
        selected_transaction = recent_transactions[transaction_num - 1]
        
        # Ask for confirmation
        await show_deletion_confirmation(update, selected_transaction)
        return
    
    # SCENARIO 4: Delete last transaction (old behavior: /delete confirm)
    if args[0].lower() == "confirm":
        # If there's a second argument, it might be a transaction number
        if len(args) > 1 and args[1].isdigit():
            transaction_num = int(args[1])
            recent_transactions = get_recent_transactions(10)
            
            if not recent_transactions:
                await update.message.reply_text("\u274c No transactions found.")
                return
            
            if transaction_num < 1 or transaction_num > len(recent_transactions):
                await update.message.reply_text(
                    f"\u274c Invalid selection. Choose 1-{len(recent_transactions)}"
                )
                return
            
            selected_transaction = recent_transactions[transaction_num - 1]
            await process_deletion(update, selected_transaction)
        else:
            # Delete last transaction (backward compatibility)
            await update.message.reply_text("\u1f5d1\ufe0f Deleting last transaction...")
            success, message, deleted_transaction = delete_last_transaction()
            await send_deletion_result(update, deleted_transaction, success, message)
        return
    
    # SCENARIO 5: Show help if arguments don't match expected patterns
    await update.message.reply_text(
        "\u1f5d1\ufe0f <b>Delete Transaction Command</b>\n\n"
        "<b>To delete a transaction:</b>\n"
        "1. <code>/delete</code> - See recent transactions\n"
        "2. <code>/delete 1</code> - Select transaction #1\n"
        "3. <code>/delete 83 confirm</code> - Delete transaction at row 83 immediately\n\n"
        "<b>Example:</b>\n"
        "<code>/delete</code> → Shows list\n"
        "<code>/delete 3</code> → Selects transaction #3\n"
        "<code>/delete 83 confirm</code> → Deletes transaction at row 83\n\n"
        "<b>Old method (still works):</b>\n"
        "<code>/delete confirm</code> - Delete last transaction",
        parse_mode='HTML'
    )

async def show_transaction_selection_for_delete(update, context):
    """Show recent transactions for deletion selection menu"""
    await update.message.reply_text("📋 Loading last 10 transactions...")
    
    # Get recent transactions
    recent_transactions = get_recent_transactions(10)  # Show last 10
    
    if not recent_transactions:
        await update.message.reply_text("❌ No transactions found.")
        return
    
    response_lines = ["🗑️ <b>Recent Transactions (Select one to delete):</b>\n"]
    
    for i, transaction in enumerate(recent_transactions, 1):
        date_val = transaction['date']
        date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
        
        # Format amounts
        amounts = []
        if transaction['usd'] is not None:
            prefix = "🔄 " if transaction['usd'] < 0 else ""
            amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
        if transaction['lbp'] is not None:
            prefix = "🔄 " if transaction['lbp'] < 0 else ""
            amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
        if transaction['euro'] is not None:
            prefix = "🔄 " if transaction['euro'] < 0 else ""
            amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
        
        amount_str = " + ".join(amounts) if amounts else "No amount"
        
        payment_emoji = "💳" if str(transaction['payment']).lower() == "card" else "💵"
        item = transaction['subcategory'] if transaction['subcategory'] else transaction['category']
        
        import html
        safe_item = html.escape(item) if item else "Unknown"
        
        # Show row number for reference
        row_num = transaction.get('row', 'N/A')
        
        response_lines.append(
            f"\n<b>{i}.</b> {payment_emoji} {date_str}: <code>{safe_item}</code>\n"
            f"     Amount: {amount_str}\n"
            f"     Category: {html.escape(transaction['category'] or 'Unknown')}\n"
            f"     Row: {row_num}"
        )
    
    response_lines.append(
        "\n\n<b>To delete, reply with:</b>\n"
        "<code>/delete NUMBER</code>\n\n"
        "<b>Example:</b>\n"
        "<code>/delete 1</code> to delete the first transaction\n"
        "<code>/delete 3 confirm</code> to delete transaction #3 immediately\n\n"
        "⚠️ <b>Warning:</b> Deletion cannot be undone!"
    )
    
    await update.message.reply_text("\n".join(response_lines), parse_mode='HTML')

async def show_deletion_confirmation(update, transaction):
    """Show confirmation message before deleting a transaction"""
    date_val = transaction['date']
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
    
    # Format amounts
    amounts = []
    if transaction['usd'] is not None:
        prefix = "\u1f504 " if transaction['usd'] < 0 else ""
        amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
    if transaction['lbp'] is not None:
        prefix = "\u1f504 " if transaction['lbp'] < 0 else ""
        amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
    if transaction['euro'] is not None:
        prefix = "\u1f504 " if transaction['euro'] < 0 else ""
        amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
    
    amount_str = " + ".join(amounts) if amounts else "No amount"
    
    payment_emoji = "\u1f4b3" if str(transaction['payment']).lower() == "card" else "\u1f4b5"
    item = transaction['subcategory'] if transaction['subcategory'] else transaction['category']
    
    import html
    safe_item = html.escape(item) if item else "Unknown"
    
    response = (
        f"\u26a0\ufe0f <b>Confirm Deletion:</b>\n\n"
        f"{payment_emoji} {date_str}: <code>{safe_item}</code>\n"
        f"Amount: {amount_str}\n"
        f"Category: {html.escape(transaction['category'] or 'Unknown')}\n"
        f"Payment: {html.escape(transaction['payment'])}"
    )
    
    if transaction.get('notes'):
        response += f"\nNotes: {html.escape(transaction['notes'][:100])}{'...' if len(transaction['notes']) > 100 else ''}"
    
    # Get the row number from the transaction
    row_number = transaction.get('row', 'N/A')
    
    response += (
        f"\n\n<b>Are you sure you want to delete this transaction?</b>\n\n"
        f"\u2705 To confirm deletion, type:\n"
        f"<code>/delete {row_number} confirm</code>\n\n"
        f"\u274c To cancel, just ignore this message.\n"
        f"<i>Note: This action cannot be undone!</i>"
    )
    
    await update.message.reply_text(response, parse_mode='HTML')

async def process_deletion(update, transaction):
    """Process the deletion of a transaction"""
    row_number = transaction.get('row')
    
    if not row_number:
        await update.message.reply_text("\u274c Error: Could not determine row number.")
        return
    
    await update.message.reply_text(f"\u1f5d1\ufe0f Deleting transaction at row {row_number}...")
    success, message, deleted_transaction = delete_transaction_at_row(row_number)
    await send_deletion_result(update, deleted_transaction, success, message)

async def send_deletion_result(update, deleted_transaction, success, message):
    """Send the deletion result to user"""
    if success and deleted_transaction:
        # Format the deleted transaction details
        date_val = deleted_transaction['date']
        date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
        
        # Format amounts
        amounts = []
        if deleted_transaction.get('usd') is not None:
            prefix = "🔄 " if deleted_transaction['usd'] < 0 else ""
            amounts.append(f"{prefix}${abs(deleted_transaction['usd']):.2f}")
        if deleted_transaction.get('lbp') is not None:
            prefix = "🔄 " if deleted_transaction['lbp'] < 0 else ""
            amounts.append(f"{prefix}{abs(deleted_transaction['lbp']):,.0f} LBP")
        if deleted_transaction.get('euro') is not None:
            prefix = "🔄 " if deleted_transaction['euro'] < 0 else ""
            amounts.append(f"{prefix}€{abs(deleted_transaction['euro']):.2f}")
        
        amount_str = " + ".join(amounts) if amounts else "No amount"
        
        payment_emoji = "💳" if str(deleted_transaction.get('payment', '')).lower() == "card" else "💵"
        item = deleted_transaction.get('subcategory') or deleted_transaction.get('category', 'Unknown')
        
        import html
        safe_item = html.escape(item) if item else "Unknown"
        
        details = (
            f"🗑️ <b>Transaction Deleted:</b>\n\n"
            f"{payment_emoji} {date_str}: {safe_item}\n"
            f"Amount: {amount_str}\n"
            f"Category: {html.escape(deleted_transaction.get('category', 'Unknown'))}\n"
            f"Payment: {html.escape(deleted_transaction.get('payment', 'Cash'))}"
        )
        
        if deleted_transaction.get('notes'):
            details += f"\nNotes: {html.escape(str(deleted_transaction['notes'])[:100])}{'...' if len(str(deleted_transaction['notes'])) > 100 else ''}"
        
        await update.message.reply_text(f"{details}\n\n{message}", parse_mode='HTML')
    else:
        await update.message.reply_text(message, parse_mode='HTML')

async def modify_command(update, context):
    """Modify a transaction with flexible field updates"""
    logger.info(f"Received /modify from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    # Check if we're in selection mode
    if not context.args:
        await show_transaction_selection(update, context)
        return
    
    # Get all arguments
    args = context.args
    
    # SCENARIO 1: Direct modification with row number and details
    # Example: /modify 1 55$ or /modify 1 to card
    if args[0].isdigit() and len(args) > 1:
        transaction_num = int(args[0])
        modification_args = args[1:]  # Everything after the number
        
        # Get recent transactions
        recent_transactions = get_recent_transactions(10)
        
        if not recent_transactions:
            await update.message.reply_text("❌ No transactions found.")
            return
        
        if transaction_num < 1 or transaction_num > len(recent_transactions):
            await update.message.reply_text(
                f"❌ Invalid selection. Choose 1-{len(recent_transactions)}"
            )
            return
        
        # Get the selected transaction
        selected_transaction = recent_transactions[transaction_num - 1]
        row_number = selected_transaction['row']
        
        # Parse the modification
        modify_text = ' '.join(modification_args)
        await process_modification(update, row_number, selected_transaction, modify_text)
        return
    
    # SCENARIO 2: Just a number - show transaction details
    # Example: /modify 1 (without modification details)
    if args[0].isdigit() and len(args) == 1:
        transaction_num = int(args[0])
        
        recent_transactions = get_recent_transactions(10)
        
        if not recent_transactions:
            await update.message.reply_text("❌ No transactions found.")
            return
        
        if transaction_num < 1 or transaction_num > len(recent_transactions):
            await update.message.reply_text(
                f"❌ Invalid selection. Choose 1-{len(recent_transactions)}"
            )
            return
        
        selected_transaction = recent_transactions[transaction_num - 1]
        
        # Store in context for next modification
        context.user_data['modify_row'] = selected_transaction['row']
        context.user_data['original_transaction'] = selected_transaction
        
        await show_selected_transaction_flexible(update, selected_transaction)
        return
    
    # SCENARIO 3: Modification without number (using stored context)
    # Example: /modify 55$ or /modify to card
    if 'modify_row' in context.user_data:
        row_number = context.user_data['modify_row']
        original_transaction = context.user_data.get('original_transaction', {})
        
        modify_text = ' '.join(args)
        await process_modification(update, row_number, original_transaction, modify_text)
    else:
        # No transaction selected, show help
        await update.message.reply_text(
            "❌ <b>No transaction selected.</b>\n\n"
            "<b>Select from last 10 transactions:</b>\n"
            "1. <code>/modify</code> - See last 10 transactions\n"
            "2. <code>/modify 1</code> - Select transaction #1\n"
            "3. <code>/modify 1 55$</code> - Modify amount to $55\n"
            "4. <code>/modify 1 to card</code> - Change payment to card\n\n"
            "<b>Direct modification examples:</b>\n"
            "• <code>/modify 1 55$</code> - Change amount\n"
            "• <code>/modify 1 to card</code> - Change payment\n"
            "• <code>/modify 1 NOTES : new notes</code> - Add notes\n"
            "• <code>/modify 1 same 20$</code> - Keep item, change amount",
            parse_mode='HTML'
        )

async def process_modification(update, row_number, original_transaction, modify_text):
    """Process a modification request"""
    # Initialize modification variables
    new_subcategory = None
    new_currency_amounts = None
    new_payment_type = None
    new_notes = None
    
    # Parse special keywords
    text_lower = modify_text.lower()
    
    # Check for payment type changes
    payment_keywords = {
        'card': 'Card',
        'cash': 'Cash',
        'bank': 'Bank Transfer',
        'digital': 'Digital Wallet',
        'transfer': 'Bank Transfer',
        'wallet': 'Digital Wallet'
    }
    
    # Check if user wants to change payment only
    payment_only = any(keyword in text_lower for keyword in ['to card', 'to cash', 'to bank', 'to digital', 'payment:', 'pay:'])
    
    if payment_only:
        for keyword, payment in payment_keywords.items():
            if keyword in text_lower:
                new_payment_type = payment
                break
        # If we found a payment type, that's the only change
        if new_payment_type:
            await update.message.reply_text(f"🔧 Changing payment to {new_payment_type}...")
            success, message, modified_transaction = modify_transaction_at_row(
                row_number, None, None, new_payment_type, None
            )
            await send_modification_result(update, original_transaction, modified_transaction, success, message)
            return
    
    # Check for notes
    notes_pattern = r'\s+NOTES\s*:\s*(.+)'
    notes_match = re.search(notes_pattern, modify_text, re.IGNORECASE)
    
    if notes_match:
        modify_text = modify_text[:notes_match.start()].strip()
        new_notes = notes_match.group(1).strip()
    
    # Check if it's just an amount (starts with number or currency symbol)
    amount_only_pattern = r'^[\$\€-]?\d'
    is_amount_only = re.match(amount_only_pattern, modify_text.strip()) is not None
    
    if is_amount_only:
        # Parse as amount only
        _, parsed_amounts, _ = extract_payment_amount_currency(modify_text)
        if parsed_amounts:
            new_currency_amounts = parsed_amounts
            await update.message.reply_text(f"🔧 Changing amount to {modify_text}...")
            success, message, modified_transaction = modify_transaction_at_row(
                row_number, None, new_currency_amounts, None, new_notes
            )
            await send_modification_result(update, original_transaction, modified_transaction, success, message)
            return
    
    # Full parsing for other cases
    parsed_subcategory, parsed_amounts, parsed_payment = extract_payment_amount_currency(modify_text)
    
    if parsed_subcategory and parsed_subcategory.lower() not in ['same', 'keep', 'unchanged', '']:
        new_subcategory = parsed_subcategory
    
    if parsed_amounts:
        new_currency_amounts = parsed_amounts
    
    if parsed_payment and parsed_payment != "Cash":
        new_payment_type = parsed_payment
    
    # Special handling for "same" keyword
    if new_subcategory and new_subcategory.lower() in ['same', 'keep', 'unchanged']:
        new_subcategory = None
    
    await update.message.reply_text(f"🔧 Modifying transaction at row {row_number}...")
    
    success, message, modified_transaction = modify_transaction_at_row(
        row_number, new_subcategory, new_currency_amounts, new_payment_type, new_notes
    )
    
    await send_modification_result(update, original_transaction, modified_transaction, success, message)

async def send_modification_result(update, original_transaction, modified_transaction, success, message):
    """Send the modification result to user"""
    if success and modified_transaction:
        # Format the response
        response = format_transaction_response(modified_transaction, "Modified")
        
        # Show what changed
        changes = []
        
        if original_transaction.get('subcategory') != modified_transaction.get('subcategory'):
            old_item = original_transaction.get('subcategory') or original_transaction.get('category', 'Unknown')
            new_item = modified_transaction.get('subcategory') or modified_transaction.get('category', 'Unknown')
            if old_item != new_item:
                changes.append(f"Item: {old_item} → {new_item}")
        
        # Compare amounts
        for currency in ['USD', 'LBP', 'EURO']:
            old_val = original_transaction.get(currency.lower())
            new_val = modified_transaction.get(currency.lower())
            if old_val != new_val:
                old_str = format_currency_amount(old_val, currency) if old_val is not None else "None"
                new_str = format_currency_amount(new_val, currency) if new_val is not None else "None"
                changes.append(f"{currency}: {old_str} → {new_str}")
        
        if original_transaction.get('payment') != modified_transaction.get('payment'):
            changes.append(f"Payment: {original_transaction.get('payment', 'Cash')} → {modified_transaction.get('payment', 'Cash')}")
        
        if original_transaction.get('notes') != modified_transaction.get('notes'):
            old_notes = original_transaction.get('notes', 'None') or 'None'
            new_notes = modified_transaction.get('notes', 'None') or 'None'
            old_preview = str(old_notes)[:50] + "..." if len(str(old_notes)) > 50 else str(old_notes)
            new_preview = str(new_notes)[:50] + "..." if len(str(new_notes)) > 50 else str(new_notes)
            changes.append(f"Notes: {old_preview} → {new_preview}")
        
        if changes:
            changes_text = "\n".join([f"• {change}" for change in changes])
            full_response = f"{response}\n\n<b>Changes:</b>\n{changes_text}\n\n{message}"
        else:
            full_response = f"{response}\n\n{message}"
        
        await update.message.reply_text(full_response, parse_mode='HTML')
    else:
        await update.message.reply_text(message, parse_mode='HTML')

async def help_command(update, context):
    """Extended help command - UPDATED with new delete feature"""
    logger.info(f"Received /help from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    await update.message.reply_text(
        f"💰 <b>Smart Budget Tracker Bot - Help Guide</b>\n\n"
        
        f"<b>🎯 Smart Matching System:</b>\n"
        f"• Minimum confidence required: {MINIMUM_CONFIDENCE*100:.0f}%\n"
        f"• Rejects poor matches with suggestions\n"
        f"• Uses fuzzy matching for close matches\n\n"
        
        "<b>🗑️ NEW: Enhanced Delete Feature:</b>\n"
        "Now you can delete any of the last 10 transactions!\n\n"
        "<b>Delete Methods:</b>\n"
        "1. <code>/delete</code> - See recent transactions list\n"
        "2. <code>/delete 3</code> - Select transaction #3 for deletion\n"
        "3. <code>/delete 3 confirm</code> - Delete transaction #3 immediately\n"
        "4. <code>/delete confirm</code> - Delete last transaction (old method)\n\n"
        
        "<b>🔄 Negative Amount Support:</b>\n"
        "Use negative amounts for:\n"
        "• Refunds\n"
        "• Corrections\n"
        "• Negative adjustments\n"
        "• Transaction type remains 'Expenses'\n\n"
        
        "<b>💱 LBP Amount Format:</b>\n"
        "• <code>200000 LBP</code> - No commas needed\n"
        "• <code>200,000 LBP</code> - Commas also work\n"
        "• <code>-150000 lbp</code> - Negative also works\n\n"
        
        "<b>🌐 Multi-Currency Support:</b>\n"
        "• <code>KSC 10$, -150000 LBP</code> - Multiple currencies in one transaction\n"
        "• <code>Grocery 20€ 150000 LBP</code> - EURO + LBP\n"
        "• <code>Refund -5$, -75000 LBP</code> - Multiple refunds\n\n"
        
        "<b>📝 Basic Usage:</b>\n"
        "Simply send a message in this format:\n"
        "<code>ITEM [AMOUNT] [CURRENCY] [PAYMENT]</code>\n\n"
        
        "<b>✅ Will Accept (Examples):</b>\n"
        "• <code>Chamsin 10</code> - Expense\n"
        "• <code>Chamsin -10$</code> - Refund (still 'Expenses' type)\n"
        "• <code>Fuel Mazda 200000 LBP</code> - No commas needed\n"
        "• <code>Refund -500$ card</code> - Negative amount\n"
        "• <code>KSC 15.50 Card</code> - Expense\n"
        "• <code>Correction -200€</code> - Negative adjustment\n"
        "• <code>KSC 10$, -150000 LBP</code> - Multi-currency\n\n"
        
        "<b>❌ Will Reject:</b>\n"
        "• <code>OKAY 10$</code> - No matching category\n"
        "• <code>Test -5</code> - Common word\n"
        "• <code>Hello 20</code> - Greeting\n\n"
        
        "<b>📥 DOWNLOAD COMMANDS:</b>\n"
        "• <code>/download csv day</code> - Today's transactions\n"
        "• <code>/download csv week</code> - This week's transactions\n"
        "• <code>/download csv month</code> - This month's transactions\n"
        "• <code>/download csv year</code> - This year's transactions\n"
        "• <code>/download csv all</code> - All transactions\n"
        "• <code>/download summary</code> - Summary statistics (JSON)\n"
        "• <code>/download backup</code> - Excel backup file\n\n"
        
        "<b>⚙️ Available Commands:</b>\n"
        "/start - Welcome message\n"
        "/help - This help message\n"
        "/testparse [text] - Test parsing & matching\n"
        "/recent - Show last 10 transactions\n"
        "/stats - Show statistics\n"
        "/save - Force save to OneDrive\n"
        "/clearcache - Clear matching cache\n"
        "/download - Download data files\n\n"
        "/unlock - Emergency unlock Excel file\n\n"

        "<b>✏️ Edit Commands:</b>\n"
        "/delete - Delete any of last 10 transactions\n"
        "/modify - Modify any of last 10 transactions\n"

        "<b>💡 Tips:</b>\n"
        "• LBP amounts work with or without commas\n"
        "• Negative amounts keep 'Expenses' type\n"
        "• Be specific with item names\n"
        "• Check spelling if not matching\n"
        "• Use /testparse to test before adding\n"
        "• Download data regularly for backup",
        parse_mode='HTML'
    )

async def testparse_command(update, context):
    """Test input parsing and matching - UPDATED for multi-currency"""
    logger.info(f"Received /testparse from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        return
    
    if not context.args:
        await update.message.reply_text(
            "Usage: /testparse [text]\n\n"
            "Examples (Multi-Currency):\n"
            "• /testparse KSC 10$, -150000 LBP\n"
            "• /testparse Grocery 20€ 150000 LBP\n"
            "• /testparse Refund -5$, -75000 LBP\n\n"
            "Examples (Single Currency):\n"
            "• /testparse Chamsin 10$\n"
            "• /testparse Chamsin -10$\n"
            "• /testparse Fuel Mazda 200000 LBP"
        )
        return
    
    test_input = ' '.join(context.args)
    
    # Parse the input with multi-currency support
    subcategory, currency_amounts, payment_type = extract_payment_amount_currency(test_input)
    
    # Load tables and find match
    tables_dict = load_all_tables_with_details()
    matched, category, match_type, confidence = find_best_match_for_input(subcategory, tables_dict)
    
    # FIXED: Convert confidence to percentage
    confidence_percentage = confidence * 100
    
    # Determine if it would be accepted
    would_accept = confidence >= MINIMUM_CONFIDENCE
    accept_status = "✅ WOULD BE ACCEPTED" if would_accept else "❌ WOULD BE REJECTED"
    
    # Format currency amounts
    amounts_str = ""
    if currency_amounts:
        amount_parts = []
        for currency, amount in currency_amounts.items():
            if amount is not None:
                amount_parts.append(f"{format_currency_amount(amount, currency)}")
        amounts_str = ", ".join(amount_parts)
    else:
        amounts_str = "None"
    
    # Format the results
    result = [
        f"🔍 <b>Parsing & Matching Test (Multi-Currency)</b>",
        f"",
        f"<b>📝 Input:</b> <code>{test_input}</code>",
        f"",
        f"<b>📊 Parsing Results:</b>",
        f"• Subcategory: <code>{subcategory}</code>",
        f"• Amount{'s' if len(currency_amounts) > 1 else ''}: {amounts_str}",
        f"• Currencies Detected: {len(currency_amounts)}",
        f"• Payment Type: {payment_type}",
        f"",
        f"<b>🎯 Matching Results:</b>",
        f"• Matched to: <code>{matched if matched else 'No match'}</code>",
        f"• Category: {category}",
        f"• Match Type: {match_type}",
        f"• Confidence: {confidence_percentage:.1f}%",
        f"• Minimum Required: {MINIMUM_CONFIDENCE*100:.0f}%",
        f"",
        f"<b>📋 Final Result:</b> {accept_status}"
    ]
    
    await update.message.reply_text("\n".join(result), parse_mode='HTML')

async def recent_command(update, context):
    """Show recent transactions - UPDATED to show last 10 transactions"""
    logger.info(f"Received /recent from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        return
    
    if not EXCEL_SUPPORT:
        await update.message.reply_text("❌ Excel libraries not installed. Cannot read Excel file.")
        return
    
    try:
        await update.message.reply_text("📋 Loading last 10 transactions...")
        
        success, msg = copy_excel_from_onedrive()
        if not success:
            await update.message.reply_text(msg)
            return
        
        with excel_operation() as wb:
            sheet = wb[TRACKING_SHEET_NAME]
            
            recent = []
            # Find the last row with data
            last_row = 12
            while sheet.cell(row=last_row, column=3).value not in [None, ""]:
                last_row += 1
            
            # Get last 10 transactions (UPDATED from 5 to 10)
            start_row = max(12, last_row - 10)
            for row in range(start_row, last_row):
                date_val = sheet.cell(row=row, column=3).value
                if date_val:
                    payment = sheet.cell(row=row, column=4).value or "Cash"
                    transaction_type = sheet.cell(row=row, column=5).value or "Expenses"
                    category = sheet.cell(row=row, column=6).value or ""
                    subcategory = sheet.cell(row=row, column=7).value or ""
                    
                    usd = sheet.cell(row=row, column=8).value
                    lbp = sheet.cell(row=row, column=9).value
                    euro = sheet.cell(row=row, column=10).value
                    
                    # Determine which amount to show
                    amounts = []
                    if usd is not None:
                        prefix = "🔄 " if usd < 0 else ""
                        amounts.append(f"{prefix}${abs(usd):.2f}")
                    if lbp is not None:
                        prefix = "🔄 " if lbp < 0 else ""
                        amounts.append(f"{prefix}{abs(lbp):,.0f} LBP")
                    if euro is not None:
                        prefix = "🔄 " if euro < 0 else ""
                        amounts.append(f"{prefix}€{abs(euro):.2f}")
                    
                    amount_str = " + ".join(amounts) if amounts else "No amount"
                    
                    if subcategory or category:
                        date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
                        payment_emoji = "💳" if str(payment).lower() == "card" else "💵"
                        item = subcategory if subcategory else category
                        import html
                        safe_item = html.escape(item)
                        
                        # Add transaction type indicator
                        type_indicator = "🔄" if (usd is not None and usd < 0) or (lbp is not None and lbp < 0) or (euro is not None and euro < 0) else "📤"
                        
                        recent.append(f"{type_indicator} {payment_emoji} {date_str}: {safe_item} - {amount_str}")
        
        if recent:
            # Show count in the header
            response = [f"📋 <b>Recent Transactions (Last {len(recent)}):</b>", ""] + recent
            await update.message.reply_text("\n".join(response), parse_mode='HTML')
        else:
            await update.message.reply_text("No recent transactions found.")
        
    except Exception as e:
        logger.error(f"Error in recent_command: {str(e)}")
        await update.message.reply_text(f"❌ Error loading transactions: {str(e)[:200]}")

async def handle_message(update, context):
    """Handle regular messages - UPDATED for multi-currency & optional notes support"""
    user_id = update.effective_user.id
    text = update.message.text.strip()
    
    logger.info(f"Received message from user {user_id}: {text}")
    
    if not is_authorized(user_id):
        logger.warning(f"User {user_id} not authorized")
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    if text.startswith('/'):
        return
    
    if not text:
        await update.message.reply_text(
            "💰 <b>How to Add a Transaction (Multi-Currency):</b>\n\n"
            "Send: <code>ITEM [AMOUNTS WITH CURRENCIES] [PAYMENT]</code>\n\n"
            "<b>📝 Examples (Single Currency):</b>\n"
            "• <code>Chamsin 10</code> (Expense)\n"
            "• <code>Chamsin -10$</code> (Refund)\n"
            "• <code>Fuel 200000 lbp</code> (No commas needed)\n\n"
            "<b>🌐 Examples (Multi-Currency):</b>\n"
            "• <code>KSC 10$, -150000 LBP</code>\n"
            "• <code>Grocery 20€ 150000 LBP</code>\n"
            "• <code>Refund -5$, -75000 LBP</code>\n\n"
            "<b>📝 Optional Notes:</b>\n"
            "Add a second line starting with 'DETAILS : ' for notes\n"
            "• <code>ITEM AMOUNTS\nDETAILS : your notes</code>\n\n"
            f"<b>🎯 Minimum Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}%\n"
            "Poor matches will be rejected with suggestions.\n\n"
            "Need help? Use /help for detailed instructions.",
            parse_mode='HTML'
        )
        return
    
    # Check if there are optional notes (separated by newline with 'DETAILS : ')
    transaction_text = text
    optional_notes = ""
    
    # Look for "DETAILS : " on a new line (case insensitive)
    details_pattern = r'\n\s*DETAILS\s*:\s*(.+)'
    details_match = re.search(details_pattern, text, re.IGNORECASE | re.DOTALL)
    
    if details_match:
        # Extract transaction part (everything before DETAILS)
        transaction_text = text[:details_match.start()].strip()
        optional_notes = details_match.group(1).strip()
        logger.info(f"Found optional notes: '{optional_notes}'")
    
    # Parse the transaction part with multi-currency support
    subcategory, currency_amounts, payment_type = extract_payment_amount_currency(transaction_text)
    
    # Check if we found any amounts
    if not currency_amounts:
        await update.message.reply_text(
            f"⚠️ <b>Amount Not Detected</b>\n\n"
            f"Could not detect any amounts with currencies.\n\n"
            f"<b>For multi-currency, specify amounts like:</b>\n"
            f"• <code>KSC 10$, -150000 LBP</code>\n"
            f"• <code>Grocery 20€ 150000 LBP</code>\n"
            f"• <code>Refund -5$, -75000 LBP</code>\n\n"
            f"<b>For optional notes:</b>\n"
            f"<code>ITEM AMOUNTS\n"
            f"DETAILS : your notes here</code>\n\n"
            f"<i>Tip: Attach currency to each amount ($, LBP, €)</i>",
            parse_mode='HTML'
        )
        return
    
    if not subcategory:
        await update.message.reply_text(
            "❌ <b>Parsing Error</b>\n\n"
            "Could not extract item from message.\n"
            "Try: <code>Item [Amounts with Currencies] [Payment]</code>\n\n"
            "<b>Examples:</b>\n"
            "• <code>KSC 10$, -150000 LBP card</code>\n"
            "• <code>Grocery 20€ 150000 LBP</code>\n"
            "• <code>Refund -5$, -75000 LBP cash</code>\n\n"
            "<b>With optional notes:</b>\n"
            "<code>ITEM AMOUNTS\n"
            f"DETAILS : your notes</code>\n\n"
            "Use /help for more examples.",
            parse_mode='HTML'
        )
        return
    
    # Show what we're doing
    amount_display_parts = []
    has_negative = False
    
    for currency, amount in currency_amounts.items():
        if amount is not None:
            amount_str = format_currency_amount(amount, currency)
            if amount < 0:
                amount_display_parts.append(f"🔄 {amount_str}")
                has_negative = True
            else:
                amount_display_parts.append(amount_str)
    
    if amount_display_parts:
        amount_display = " + ".join(amount_display_parts)
        transaction_desc = "Expense (with refunds/corrections)" if has_negative else "Expense"
    else:
        amount_display = "No amount specified"
        transaction_desc = "Transaction"
    
    payment_emoji = PAYMENT_TYPES.get(payment_type, "💵")
    
    # Get currency emojis
    currency_emojis = []
    for currency in currency_amounts.keys():
        if currency == "USD":
            currency_emojis.append("💵")
        elif currency == "LBP":
            currency_emojis.append("🇱🇧")
        elif currency == "EURO":
            currency_emojis.append("💶")
    
    currency_emoji_str = " ".join(currency_emojis)
    
    import html
    safe_subcategory = html.escape(subcategory)
    
    # Show notes preview if provided
    notes_preview = f"\n📝 <b>Notes:</b> {html.escape(optional_notes[:50])}{'...' if len(optional_notes) > 50 else ''}" if optional_notes else ""
    
    # Add multi-currency indicator
    multi_currency_indicator = "🌐 " if len(currency_amounts) > 1 else ""
    
    await update.message.reply_text(
        f"{multi_currency_indicator}💰 <b>Processing {transaction_desc}:</b>\n"
        f"Item: <code>{safe_subcategory}</code>\n"
        f"Amount{'s' if len(currency_amounts) > 1 else ''}: {amount_display}\n"
        f"Payment: {payment_type}"
        f"{notes_preview}\n\n"
        f"⏳ Matching and adding...",
        parse_mode='HTML'
    )
    
    # Add transaction with optional notes support
    success, message = add_transaction_smart(subcategory, currency_amounts, payment_type, optional_notes)
    await update.message.reply_text(message, parse_mode='HTML')

async def stats_command(update, context):
    """Show statistics"""
    logger.info(f"Received /stats from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    try:
        success, msg = copy_excel_from_onedrive()
        if not success:
            await update.message.reply_text(msg)
            return
        
        total_transactions = 0
        if EXCEL_SUPPORT:
            wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
            
            if TRACKING_SHEET_NAME in wb.sheetnames:
                sheet = wb[TRACKING_SHEET_NAME]
                
                # Count transactions
                row = 12
                while sheet.cell(row=row, column=3).value not in [None, ""]:
                    total_transactions += 1
                    row += 1
                    if row > 1000:
                        break
            
            wb.close()
        else:
            total_transactions = "N/A (Excel libraries not installed)"
        
        # Get table count from cache
        table_count = len(_table_cache) if _table_cache else 0
        total_items = sum(data['count'] for data in _table_cache.values()) if _table_cache else 0
        
        # Get last backup time
        def get_last_backup_time() -> str:
            try:
                backups = list(BACKUP_DIR.glob("backup_*"))
                if not backups:
                    return "No backups yet"
                
                latest = max(backups, key=lambda x: x.stat().st_mtime)
                mod_time = datetime.fromtimestamp(latest.stat().st_mtime)
                return mod_time.strftime("%Y-%m-%d %H:%M")
            except:
                return "Error checking backups"
        
        # Check export directory
        export_files = list(EXPORT_DIR.glob("*")) if EXPORT_DIR.exists() else []
        
        await update.message.reply_text(
            f"📊 <b>Budget Tracker Statistics</b>\n\n"
            f"• <b>Excel Integration:</b> {'✅ Installed' if EXCEL_SUPPORT else '❌ Not installed'}\n"
            f"• <b>Total Transactions:</b> {total_transactions}\n"
            f"• <b>Categories Loaded:</b> {table_count}\n"
            f"• <b>Items in Database:</b> {total_items}\n"
            f"• <b>Cache Status:</b> {'✅ Loaded' if _table_cache else '⚠️ Empty'}\n"
            f"• <b>Last Backup:</b> {get_last_backup_time()}\n"
            f"• <b>Exports Available:</b> {len(export_files)} files\n"
            f"• <b>Min Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}%\n"
            f"• <b>Negative Amounts:</b> ✅ Supported\n"
            f"• <b>LBP without commas:</b> ✅ Supported\n"
            f"• <b>Bot Status:</b> ✅ Running",
            parse_mode='HTML'
        )
        
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def save_command(update, context):
    """Force save to OneDrive"""
    logger.info(f"Received /save from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        return
    
    await update.message.reply_text("💾 Saving to OneDrive...")
    success, message = save_excel_to_onedrive()
    await update.message.reply_text(message)

async def clear_cache_command(update, context):
    """Clear the table cache"""
    logger.info(f"Received /clearcache from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    global _table_cache, _table_cache_timestamp
    _table_cache = {}
    _table_cache_timestamp = None
    
    await update.message.reply_text("✅ Cache cleared! Will reload on next transaction.")

async def show_transaction_selection(update, context):
    """Show recent transactions for selection menu - UPDATED for last 10 transactions"""
    await update.message.reply_text("📋 Loading last 10 transactions...")
    
    # Get recent transactions - UPDATED to last 10
    recent_transactions = get_recent_transactions(10)  # Get last 10
    
    if not recent_transactions:
        await update.message.reply_text("❌ No transactions found.")
        return
    
    response_lines = ["📋 <b>Recent Transactions (Select one to modify):</b>\n"]
    
    for i, transaction in enumerate(recent_transactions, 1):
        date_val = transaction['date']
        date_str = date_val.strftime("%d/%m") if hasattr(date_val, 'strftime') else str(date_val)
        
        # Format amounts
        amounts = []
        if transaction['usd'] is not None:
            prefix = "🔄 " if transaction['usd'] < 0 else ""
            amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
        if transaction['lbp'] is not None:
            prefix = "🔄 " if transaction['lbp'] < 0 else ""
            amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
        if transaction['euro'] is not None:
            prefix = "🔄 " if transaction['euro'] < 0 else ""
            amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
        
        amount_str = " + ".join(amounts) if amounts else "No amount"
        
        payment_emoji = "💳" if str(transaction['payment']).lower() == "card" else "💵"
        item = transaction['subcategory'] if transaction['subcategory'] else transaction['category']
        
        import html
        safe_item = html.escape(item) if item else "Unknown"
        
        response_lines.append(
            f"\n<b>{i}.</b> {payment_emoji} {date_str}: <code>{safe_item}</code>\n"
            f"     Amount: {amount_str}\n"
            f"     Category: {html.escape(transaction['category'] or 'Unknown')}"
        )
    
    response_lines.append(
        f"\n\n<b>To modify, reply with:</b>\n"
        f"<code>/modify NUMBER</code>\n\n"
        f"<b>Example:</b> <code>/modify 1</code> to modify the first transaction\n"
        f"<i>Showing last {len(recent_transactions)} transactions</i>"
    )
    
    await update.message.reply_text("\n".join(response_lines), parse_mode='HTML')

async def show_selected_transaction_flexible(update, transaction):
    """Show selected transaction with flexible modification options"""
    date_val = transaction['date']
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
    
    # Format amounts
    amounts = []
    if transaction['usd'] is not None:
        prefix = "🔄 " if transaction['usd'] < 0 else ""
        amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
    if transaction['lbp'] is not None:
        prefix = "🔄 " if transaction['lbp'] < 0 else ""
        amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
    if transaction['euro'] is not None:
        prefix = "🔄 " if transaction['euro'] < 0 else ""
        amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
    
    amount_str = " + ".join(amounts) if amounts else "No amount"
    
    payment_emoji = "💳" if str(transaction['payment']).lower() == "card" else "💵"
    item = transaction['subcategory'] if transaction['subcategory'] else transaction['category']
    
    import html
    safe_item = html.escape(item) if item else "Unknown"
    
    response = (
        f"🔧 <b>Selected Transaction:</b>\n\n"
        f"{payment_emoji} {date_str}: <code>{safe_item}</code>\n"
        f"Amount: {amount_str}\n"
        f"Category: {html.escape(transaction['category'] or 'Unknown')}\n"
        f"Payment: {html.escape(transaction['payment'])}\n"
    )
    
    if transaction.get('notes'):
        response += f"Notes: {html.escape(transaction['notes'][:100])}{'...' if len(transaction['notes']) > 100 else ''}\n"
    
    response += (
        f"\n<b>Now modify what you want:</b>\n\n"
        f"<b>Change everything:</b>\n"
        f"<code>/modify NewItem 15$ card</code>\n\n"
        f"<b>Change amount only:</b>\n"
        f"<code>/modify 20$</code> or <code>/modify 150000 LBP</code>\n\n"
        f"<b>Change payment only:</b>\n"
        f"<code>/modify to card</code> or <code>/modify payment: cash</code>\n\n"
        f"<b>Change item only:</b>\n"
        f"<code>/modify Chamsin</code>\n\n"
        f"<b>Add notes:</b>\n"
        f"<code>/modify NOTES : Your notes here</code>\n\n"
        f"<b>Keep original:</b> Use <code>same</code> or <code>keep</code>\n"
        f"Example: <code>/modify same 20$</code> (keeps item, changes amount)"
    )
    
    await update.message.reply_text(response, parse_mode='HTML')

# ========== NEW: DOWNLOAD COMMAND ==========

async def download_command(update, context):
    """Handle download requests"""
    logger.info(f"Received /download from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "📥 <b>Download Options:</b>\n\n"
            "<b>CSV Exports:</b>\n"
            "• <code>/download csv day</code> - Today's transactions\n"
            "• <code>/download csv week</code> - This week's transactions\n"
            "• <code>/download csv month</code> - This month's transactions\n"
            "• <code>/download csv year</code> - This year's transactions\n"
            "• <code>/download csv all</code> - All transactions\n\n"
            "<b>Other Downloads:</b>\n"
            "• <code>/download summary</code> - Summary statistics (JSON)\n"
            "• <code>/download backup</code> - Excel backup file\n\n"
            "<b>Examples:</b>\n"
            "• <code>/download csv month</code>\n"
            "• <code>/download summary</code>\n"
            "• <code>/download backup</code>",
            parse_mode='HTML'
        )
        return
    
    command = context.args[0].lower()
    
    if command == "csv":
        if len(context.args) < 2:
            await update.message.reply_text(
                "Usage: /download csv [day/week/month/year/all]\n\n"
                "Examples:\n"
                "• /download csv day\n"
                "• /download csv month\n"
                "• /download csv all"
            )
            return
        
        time_range = context.args[1].lower()
        if time_range not in ["day", "week", "month", "year", "all"]:
            await update.message.reply_text(
                "❌ Invalid time range. Use: day, week, month, year, or all"
            )
            return
        
        await update.message.reply_text(f"📊 Exporting {time_range} transactions to CSV...")
        
        # Export to CSV
        success, result = export_to_csv(time_range)
        
        if success:
            try:
                # Send the file
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"transactions_{time_range}_{datetime.now().strftime('%Y%m%d')}.csv",
                        caption=f"📊 Transactions ({time_range}) - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Clean up the file after sending
                os.remove(result)
            except Exception as e:
                logger.error(f"Error sending CSV file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    elif command == "summary":
        await update.message.reply_text("📈 Exporting summary statistics...")
        
        # Export summary
        success, result = export_summary()
        
        if success:
            try:
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"budget_summary_{datetime.now().strftime('%Y%m%d')}.json",
                        caption=f"📈 Budget Summary - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Clean up
                os.remove(result)
            except Exception as e:
                logger.error(f"Error sending summary file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    elif command == "backup":
        await update.message.reply_text("💾 Creating Excel backup...")
        
        # Create backup copy
        success, result = create_backup_copy()
        
        if success:
            try:
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"budget_backup_{datetime.now().strftime('%Y%m%d')}.xlsm",
                        caption=f"💾 Excel Backup - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Note: Don't delete backup files from backup directory
            except Exception as e:
                logger.error(f"Error sending backup file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    else:
        await update.message.reply_text(
            "❌ Unknown download type.\n"
            "Use: csv, summary, or backup\n\n"
            "Example: /download csv month"
        )

async def error_handler(update, context):
    """Handle errors"""
    logger.error(f"Update {update} caused error {context.error}", exc_info=True)
    
    try:
        import html
        error_msg = html.escape(str(context.error)[:200])
        await update.message.reply_text(
            f"❌ <b>Bot Error</b>\n\n"
            f"{error_msg}\n\n"
            f"Please try again or use /help for assistance.",
            parse_mode='HTML'
        )
    except:
        pass

# ========== MAIN FUNCTION ==========

# ========== MAIN FUNCTION ==========

def main():
    """Start the bot"""
    print("\n" + "="*70)
    print("💰 SMART BUDGET TRACKER BOT - WITH DOWNLOAD FEATURE")
    print("="*70)
    
    # Check BOT_TOKEN
    if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        print("❌ ERROR: BOT_TOKEN not set in .env file!")
        input("\nPress Enter to exit...")
        return
    
    # Check ALLOWED_USER_IDS
    if not ALLOWED_USER_IDS:
        print("❌ ERROR: No ALLOWED_USER_IDS or ALLOWED_USER_ID set in .env file!")
        print("Add your user ID to .env file:")
        print("Example: ALLOWED_USER_ID=1663164223")
        input("\nPress Enter to exit...")
        return
    
    print(f"✅ BOT_TOKEN: {BOT_TOKEN[:10]}...")
    print(f"✅ ALLOWED_USER_IDS: {ALLOWED_USER_IDS}")
    print(f"✅ Excel support: {'✅ INSTALLED' if EXCEL_SUPPORT else '❌ NOT INSTALLED'}")
    print(f"✅ Minimum Confidence: {MINIMUM_CONFIDENCE*100:.0f}%")
    print(f"✅ Negative Amounts: ✅ SUPPORTED")
    print(f"✅ LBP without commas: ✅ SUPPORTED")
    print(f"✅ Download Feature: ✅ ADDED")
    
    # Check file paths
    if not ONEDRIVE_PATH.exists():
        print(f"❌ OneDrive file not found: {ONEDRIVE_PATH}")
        input("\nPress Enter to exit...")
        return
    
    print(f"✅ OneDrive file exists: {ONEDRIVE_PATH}")
    
    # Setup directories
    BACKUP_DIR.mkdir(exist_ok=True)
    EXPORT_DIR.mkdir(exist_ok=True)
    
    # Remove potentially locked file (without try-except indentation error)
    if LOCAL_COPY_PATH.exists():
        try:
            os.remove(LOCAL_COPY_PATH)
            print("✅ Removed potentially locked file")
        except Exception as e:
            print(f"⚠️ Could not remove file (may be in use): {e}")
    
    # Check file accessibility
    if LOCAL_COPY_PATH.exists():
        try:
            # Try to open the file to check if it's locked
            with open(LOCAL_COPY_PATH, 'r') as f:
                pass
            print("✅ No file locks detected")
        except PermissionError:
            print("⚠️  File appears to be locked")
            # You need to define unlock_excel_file() function or remove this part
            print("Trying to unlock...")
            # Since unlock_excel_file is not defined, let's skip or define it
            try:
                # Simple unlock attempt - just remove and recreate
                if LOCAL_COPY_PATH.exists():
                    os.remove(LOCAL_COPY_PATH)
                    print("✅ File removed and will be recreated")
            except:
                print("❌ Could not unlock file")
                print("Please close Excel and any other programs using the file")
                input("\nPress Enter to exit...")
                return

    print(f"✅ Backup directory: {BACKUP_DIR}")
    print(f"✅ Export directory: {EXPORT_DIR}")
    
    # Test file copy
    print("\n📂 Testing file operations...")
    success, msg = copy_excel_from_onedrive()
    print(f"File copy: {msg}")
    
    if not success:
        input("\nPress Enter to exit...")
        return
    
    # Load tables (check if EXCEL_SUPPORT is available)
    print("\n🔧 Loading data from Excel...")
    if EXCEL_SUPPORT:
        tables_dict = load_all_tables_with_details()
        if tables_dict:
            total_items = sum(data['count'] for data in tables_dict.values())
            print(f"✅ Loaded {len(tables_dict)} categories with {total_items} items")
        else:
            print("❌ Could not load data from Excel")
            print("⚠️  Make sure the Excel file has the correct sheet names:")
            print(f"    Tracking Sheet: '{TRACKING_SHEET_NAME}'")
            print(f"    Dropdown Sheet: '{DROPDOWN_SHEET_NAME}'")
    else:
        print("⚠️  Excel support not available. Please install openpyxl:")
        print("    pip install openpyxl")
    
    print("\n" + "="*70)
    print("🤖 BOT IS STARTING...")
    print("="*70)
    
    # Check if telegram is available
    if not TELEGRAM_AVAILABLE:
        print("❌ Telegram bot libraries not installed")
        print("Try: pip install python-telegram-bot")
        input("\nPress Enter to exit...")
        return
    
    # Import telegram here (already imported at top, but just to be safe)
    from telegram.ext import Application, CommandHandler, MessageHandler, filters
    
    # Create bot application
    try:
        app = Application.builder().token(BOT_TOKEN).build()
        logger.info("Bot application created successfully")
    except Exception as e:
        print(f"❌ Failed to create bot: {e}")
        input("\nPress Enter to exit...")
        return
    
    # Add command handlers
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("testparse", testparse_command))
    app.add_handler(CommandHandler("recent", recent_command))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CommandHandler("save", save_command))
    app.add_handler(CommandHandler("clearcache", clear_cache_command))
    app.add_handler(CommandHandler("download", download_command))
    app.add_handler(CommandHandler("delete", delete_command))
    app.add_handler(CommandHandler("modify", modify_command))
    
    # Add message handler
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Add error handler
    app.add_error_handler(error_handler)
    
    print("\n✅ Bot is ready and running!")
    print("📱 Open Telegram and send /start to your bot")
    print("📥 New: Use /download to export data")
    print("⏳ Waiting for messages...")
    print("="*70)
    
    try:
        # Start polling
        app.run_polling(
            poll_interval=1.0,
            timeout=10,
            drop_pending_updates=True
        )
    except KeyboardInterrupt:
        print("\n🛑 Bot stopped by user")
    except Exception as e:
        print(f"\n❌ Bot crashed: {e}")
        logger.error(f"Bot crashed: {e}", exc_info=True)
        input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()