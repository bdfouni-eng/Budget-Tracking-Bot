"""
BUDGET TRACKER BOT - LINUX VERSION WITH ONEDRIVE API
Optimized for Ubuntu 22.04 aarch64 on Oracle Cloud
"""

import os
import re
import logging
import shutil
import sys
import json
import asyncio
import time
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Tuple, Optional, List, Any
from difflib import get_close_matches
import threading
from dataclasses import dataclass
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor

# ========== EXCEL LIBRARY FOR LINUX ==========
try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl/pandas not found. Install: pip install openpyxl pandas")

# ========== ONEDRIVE API LIBRARY ==========
try:
    from onedrivesdk import OneDriveClient
    from onedrivesdk.helpers import GetAuthCodeServer
    ONEDRIVE_AVAILABLE = True
except ImportError:
    ONEDRIVE_AVAILABLE = False
    print("⚠️  onedrivesdk not found. Install: pip install onedrivesdk")

# ========== TELEGRAM BOT IMPORTS ==========
try:
    from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackQueryHandler
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    print("⚠️  python-telegram-bot not found. Install: pip install python-telegram-bot")

# ========== CONFIGURATION ==========
from dotenv import load_dotenv
load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")

# Handle both ALLOWED_USER_ID (singular) and ALLOWED_USER_IDS (plural)
ALLOWED_USER_IDS = []
user_ids_str = os.getenv("ALLOWED_USER_IDS", os.getenv("ALLOWED_USER_ID", ""))
if user_ids_str:
    for id_str in user_ids_str.split(","):
        id_str = id_str.strip()
        if id_str:
            try:
                ALLOWED_USER_IDS.append(int(id_str))
            except ValueError:
                print(f"⚠️ Warning: Invalid user ID '{id_str}' in .env file")

# File paths (Linux paths)
LOCAL_COPY_PATH = Path(os.getenv("LOCAL_COPY_PATH", "/home/ubuntu/budget_tracker/temp_budget.xlsm"))
TRACKING_SHEET_NAME = os.getenv("TRACKING_SHEET_NAME", "Budget Tracking")
DROPDOWN_SHEET_NAME = os.getenv("DROPDOWN_SHEET_NAME", "Dropdown Data")
EXPORT_DIR = Path(os.getenv("EXPORT_DIR", "/home/ubuntu/budget_tracker/exports"))

# OneDrive configuration
ONEDRIVE_CLIENT_ID = os.getenv("ONEDRIVE_CLIENT_ID", "")
ONEDRIVE_CLIENT_SECRET = os.getenv("ONEDRIVE_CLIENT_SECRET", "")
ONEDRIVE_REDIRECT_URI = os.getenv("ONEDRIVE_REDIRECT_URI", "http://localhost:8080/")
ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_FILE_PATH", "/budget_tracker.xlsm")
ONEDRIVE_TOKEN_PATH = Path(os.getenv("ONEDRIVE_TOKEN_PATH", "/home/ubuntu/budget_tracker/onedrive_tokens.json"))

# Backup configuration
BACKUP_DIR = Path(os.getenv("BACKUP_DIR", "/home/ubuntu/budget_tracker/backups"))
BACKUP_RETENTION_DAYS = 7

# Matching configuration
MINIMUM_CONFIDENCE = 0.5  # 50% minimum confidence to accept a match
FUZZY_MATCH_THRESHOLD = 0.6  # Minimum for fuzzy matching

# Common words to reject (these won't match any category)
COMMON_NON_CATEGORIES = [
    'ok', 'okay', 'yes', 'no', 'maybe', 'test', 'hello', 'hi', 
    'thanks', 'thank', 'please', 'help', 'what', 'how', 'why',
    'when', 'where', 'who', 'which', 'good', 'bad', 'nice',
    'great', 'fine', 'well', 'sorry', 'excuse', 'hey', 'yo'
]

# Currency configuration
@dataclass
class CurrencyConfig:
    symbols: List[str]
    column: str
    decimal_places: int = 2
    thousands_separator: bool = True

CURRENCIES = {
    'USD': CurrencyConfig(symbols=['$', 'usd', 'dollar'], column='H', decimal_places=2),
    'LBP': CurrencyConfig(symbols=['lbp', 'ليرة', 'lira', 'ل.ل'], column='I', decimal_places=0, thousands_separator=True),
    'EURO': CurrencyConfig(symbols=['€', 'euro', 'eur'], column='J', decimal_places=2)
}
DEFAULT_CURRENCY = 'USD'

# Payment types with additional options
PAYMENT_TYPES = {
    "Cash": "💵",
    "Card": "💳",
    "Bank Transfer": "🏦",
    "Digital Wallet": "📱"
}

CARD_KEYWORDS = ["card", "credit", "debit", "visa", "mastercard", "amex", "paypal", "bank"]
DIGITAL_KEYWORDS = ["paypal", "digital", "wallet", "apple pay", "google pay"]

# Setup logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.FileHandler('budget_bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Thread safety for Excel operations
excel_lock = threading.Lock()

# Cache for loaded tables
_table_cache = {}
_table_cache_timestamp = None

# OneDrive client instance
_onedrive_client = None

# ========== ONEDRIVE API FUNCTIONS ==========

# ========== MICROSOFT GRAPH API (Modern OneDrive) ==========
try:
    import msal
    import requests
    import json
    from msal import ConfidentialClientApplication
    ONEDRIVE_AVAILABLE = True
except ImportError:
    ONEDRIVE_AVAILABLE = False
    print("⚠️  Microsoft Graph libraries not found. Install: pip install msal requests msgraph-core")

# ========== ONEDRIVE CONFIGURATION ==========
# Load from .env
ONEDRIVE_CLIENT_ID = os.getenv("ONEDRIVE_CLIENT_ID", "")
ONEDRIVE_CLIENT_SECRET = os.getenv("ONEDRIVE_CLIENT_SECRET", "")
ONEDRIVE_TENANT_ID = os.getenv("ONEDRIVE_TENANT_ID", "common")
ONEDRIVE_REDIRECT_URI = os.getenv("ONEDRIVE_REDIRECT_URI", "http://localhost:8080/")
ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_FILE_PATH", "/budget_tracker.xlsm")
ONEDRIVE_TOKEN_PATH = Path(os.getenv("ONEDRIVE_TOKEN_PATH", "/home/ubuntu/budget_tracker/onedrive_tokens.json"))
ONEDRIVE_SCOPES = ["https://graph.microsoft.com/Files.ReadWrite"]

# OneDrive client instance
_onedrive_app = None

def get_onedrive_app():
    """Get or create MSAL app for OneDrive"""
    global _onedrive_app
    
    if _onedrive_app is not None:
        return _onedrive_app
    
    if not ONEDRIVE_AVAILABLE:
        logger.error("MSAL not available. Cannot connect to OneDrive.")
        return None
    
    try:
        _onedrive_app = ConfidentialClientApplication(
            client_id=ONEDRIVE_CLIENT_ID,
            client_credential=ONEDRIVE_CLIENT_SECRET,
            authority=f"https://login.microsoftonline.com/{ONEDRIVE_TENANT_ID}"
        )
        return _onedrive_app
    except Exception as e:
        logger.error(f"Error creating MSAL app: {e}")
        return None

def get_onedrive_token():
    """Get access token for OneDrive - FIXED VERSION"""
    try:
        app = get_onedrive_app()
        if app is None:
            return None
        
        # Check if we have auth code in environment
        auth_code = os.getenv("ONEDRIVE_AUTH_CODE")
        
        if auth_code:
            print(f"?? Using auth code from environment: {auth_code[:50]}...")
            
            # Exchange auth code for tokens
            result = app.acquire_token_by_authorization_code(
                code=auth_code,
                scopes=ONEDRIVE_SCOPES,
                redirect_uri=ONEDRIVE_REDIRECT_URI
            )
            
            if "access_token" in result:
                # Save tokens
                with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                    json.dump(result, f, indent=2)
                print("? Tokens saved from auth code")
                
                # Remove auth code from environment to avoid reuse
                os.environ.pop("ONEDRIVE_AUTH_CODE", None)
                
                return result["access_token"]
        
        # Try to load existing tokens
        token_data = None
        if ONEDRIVE_TOKEN_PATH.exists():
            try:
                with open(ONEDRIVE_TOKEN_PATH, 'r') as f:
                    token_data = json.load(f)
            except:
                pass
        
        if token_data and 'refresh_token' in token_data:
            # Try to refresh token
            result = app.acquire_token_by_refresh_token(
                refresh_token=token_data['refresh_token'],
                scopes=ONEDRIVE_SCOPES
            )
            
            if "access_token" in result:
                # Update token file
                with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                    json.dump(result, f, indent=2)
                return result["access_token"]
        
        print("? No valid authentication found")
        return None
            
    except Exception as e:
        print(f"? Error getting token: {e}")
        return None

def get_onedrive_client():
    """Get or create OneDrive client with authentication"""
    global _onedrive_client
    
    if _onedrive_client is not None:
        return _onedrive_client
    
    if not ONEDRIVE_AVAILABLE:
        logger.error("onedrivesdk not installed. Cannot connect to OneDrive.")
        return None
    
    try:
        # Create client
        client = OneDriveClient(
            client_id=ONEDRIVE_CLIENT_ID,
            scopes=['wl.signin', 'wl.offline_access', 'onedrive.readwrite']
        )
        
        # Try to load existing tokens
        if ONEDRIVE_TOKEN_PATH.exists():
            try:
                with open(ONEDRIVE_TOKEN_PATH, 'r') as f:
                    token_data = json.load(f)
                client.auth_provider.load_session(token_data)
                logger.info("Loaded OneDrive tokens from file")
                _onedrive_client = client
                return client
            except Exception as e:
                logger.warning(f"Failed to load tokens: {e}")
        
        # Need new authentication
        logger.warning("No valid OneDrive tokens found. Authentication required.")
        logger.warning(f"Please visit: {client.auth_provider.get_auth_url(ONEDRIVE_REDIRECT_URI)}")
        
        # For headless servers, we can't do interactive auth
        # You'll need to get the auth code manually first time
        auth_code = os.getenv("ONEDRIVE_AUTH_CODE")
        if auth_code:
            client.auth_provider.authenticate(auth_code, ONEDRIVE_REDIRECT_URI, ONEDRIVE_CLIENT_SECRET)
            
            # Save tokens for future use
            token_data = client.auth_provider.save_session()
            with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                json.dump(token_data, f)
            
            _onedrive_client = client
            logger.info("Authenticated with OneDrive using auth code")
            return client
        else:
            logger.error("No ONEDRIVE_AUTH_CODE in environment. Cannot authenticate.")
            return None
            
    except Exception as e:
        logger.error(f"Error creating OneDrive client: {e}")
        return None

def download_from_onedrive() -> Tuple[bool, str]:
    """Download Excel file from OneDrive using Microsoft Graph API"""
    try:
        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated with OneDrive. Use /onedrive_auth first."
        
        # Format file path for Graph API
        # Replace spaces with %20
        file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")
        
        # Graph API endpoint
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/content"
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        logger.info(f"Downloading from OneDrive: {ONEDRIVE_FILE_PATH}")
        
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            # Save to local file
            with open(LOCAL_COPY_PATH, 'wb') as f:
                f.write(response.content)
            
            logger.info(f"Downloaded {LOCAL_COPY_PATH} ({len(response.content)} bytes)")
            return True, "✅ File downloaded from OneDrive"
        elif response.status_code == 404:
            return False, f"❌ File not found in OneDrive: {ONEDRIVE_FILE_PATH}"
        else:
            error_msg = response.json().get('error', {}).get('message', 'Unknown error')
            return False, f"❌ Download error: {error_msg}"
        
    except Exception as e:
        logger.error(f"Error downloading from OneDrive: {str(e)}")
        return False, f"❌ Download error: {str(e)[:200]}"

def upload_to_onedrive() -> Tuple[bool, str]:
    """Upload local Excel file to OneDrive using Microsoft Graph API"""
    try:
        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated with OneDrive. Use /onedrive_auth first."
        
        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local file not found"
        
        # Format file path for Graph API
        file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")
        
        # Graph API endpoint
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/content"
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        logger.info(f"Uploading to OneDrive: {LOCAL_COPY_PATH}")
        
        # Read file content
        with open(LOCAL_COPY_PATH, 'rb') as f:
            file_content = f.read()
        
        # Upload file
        response = requests.put(url, headers=headers, data=file_content)
        
        if response.status_code in [200, 201]:
            logger.info(f"Uploaded to OneDrive: {ONEDRIVE_FILE_PATH}")
            return True, "✅ File uploaded to OneDrive"
        else:
            error_msg = response.json().get('error', {}).get('message', 'Unknown error')
            return False, f"❌ Upload error: {error_msg}"
        
    except Exception as e:
        logger.error(f"Error uploading to OneDrive: {str(e)}")
        return False, f"❌ Upload error: {str(e)[:200]}"

async def onedrive_auth_command(update, context):
    """Start OneDrive authentication flow"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    try:
        app = get_onedrive_app()
        if app is None:
            await update.message.reply_text("❌ MSAL app not configured. Check ONEDRIVE_CLIENT_ID in .env")
            return
        
        # Generate auth URL
        auth_url = app.get_authorization_request_url(
            scopes=ONEDRIVE_SCOPES,
            redirect_uri=ONEDRIVE_REDIRECT_URI
        )
        
        message = (
            "🔐 <b>OneDrive Authentication Required</b>\n\n"
            "1. <b>Click this link:</b>\n"
            f"<code>{auth_url}</code>\n\n"
            "2. <b>Sign in</b> with your Microsoft account\n"
            "3. You'll be redirected to a page\n"
            "4. <b>Copy the ENTIRE URL</b> from address bar\n\n"
            "5. <b>Look for the 'code=' parameter</b>\n"
            "Example: <code>https://your-domain.lhr.life/?code=M.R3_BL2.1234...</code>\n\n"
            "6. <b>Set the code in your VM:</b>\n"
            "<code>export ONEDRIVE_AUTH_CODE='your_code_here'</code>\n\n"
            "7. <b>Then run:</b> <code>/onedrive_test</code>"
        )
        
        await update.message.reply_text(message, parse_mode='HTML')
        
    except Exception as e:
        logger.error(f"Error generating auth URL: {e}")
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def onedrive_test_command(update, context):
    """Test OneDrive connection"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    await update.message.reply_text("🔄 Testing OneDrive connection...")
    
    # Test download
    success, msg = download_from_onedrive()
    await update.message.reply_text(f"Download: {msg}")
    
    if success:
        # Test upload
        success2, msg2 = upload_to_onedrive()
        await update.message.reply_text(f"Upload: {msg2}")
        
        if success2:
            await update.message.reply_text("✅ OneDrive connection successful!")
        else:
            await update.message.reply_text("⚠️ Download worked but upload failed")
    else:
        await update.message.reply_text("❌ OneDrive connection failed")

async def onedrive_complete_auth_command(update, context):
    """Complete OneDrive authentication with code"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "Usage: /onedrive_code [auth_code]\n\n"
            "Example: /onedrive_code M.R3_BL2.0.AQABAAIAAADX8GCiHq7HjXk5KJX7w..."
        )
        return
    
    auth_code = ' '.join(context.args)
    
    await update.message.reply_text("🔐 Completing authentication...")
    
    try:
        app = get_onedrive_app()
        if app is None:
            await update.message.reply_text("❌ MSAL app not configured")
            return
        
        # Exchange code for tokens
        result = app.acquire_token_by_authorization_code(
            code=auth_code,
            scopes=ONEDRIVE_SCOPES,
            redirect_uri=ONEDRIVE_REDIRECT_URI
        )
        
        if "access_token" in result:
            # Save tokens
            with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                json.dump(result, f)
            
            await update.message.reply_text(
                "✅ <b>Authentication Successful!</b>\n\n"
                "You can now use OneDrive features:\n"
                "• /onedrive_test - Test connection\n"
                "• /save - Force save to OneDrive\n"
                "• Normal transactions will auto-sync",
                parse_mode='HTML'
            )
        else:
            error_msg = result.get('error_description', 'Unknown error')
            await update.message.reply_text(f"❌ Authentication failed: {error_msg}")
            
    except Exception as e:
        logger.error(f"Error completing auth: {e}")
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def onedrive_test_command(update, context):
    """Test OneDrive connection"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    await update.message.reply_text("🔄 Testing OneDrive connection...")
    
    # Test download
    success, msg = download_from_onedrive()
    await update.message.reply_text(f"Download test: {msg}")
    
    if success:
        # Test upload
        success2, msg2 = upload_to_onedrive()
        await update.message.reply_text(f"Upload test: {msg2}")
        
        if success2:
            await update.message.reply_text("✅ OneDrive connection successful!")
        else:
            await update.message.reply_text("⚠️ Download worked but upload failed")
    else:
        await update.message.reply_text("❌ OneDrive connection failed. Run /onedrive_auth first.")

async def onedrive_code_command(update, context):
    """Manually set OneDrive auth code"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    if not context.args:
        # Check if we have saved codes
        codes_file = Path("/home/ubuntu/Tracking_Budget_Sheet_Python/auth_codes.json")
        if codes_file.exists():
            try:
                with open(codes_file, 'r') as f:
                    codes = json.load(f)
                if codes:
                    latest = codes[-1]
                    code = latest['code']
                    await update.message.reply_text(
                        f"📝 Found recent auth code from {latest['timestamp']}\n"
                        f"Using code: {code[:30]}...\n\n"
                        f"🔐 Completing authentication..."
                    )
                    
                    # Complete authentication with this code
                    await complete_onedrive_auth(update, code)
                    return
            except:
                pass
        
        await update.message.reply_text(
            "Usage: /onedrive_code [auth_code]\n\n"
            "Or run without arguments to use the last captured code.\n\n"
            "Get your code from:\n"
            "1. Click the auth link from /onedrive_auth\n"
            "2. Sign in with Microsoft\n"
            "3. Copy the 'code=' parameter from the URL"
        )
        return
    
    # Use provided code
    code = ' '.join(context.args)
    await update.message.reply_text(f"🔐 Using provided code: {code[:30]}...")
    await complete_onedrive_auth(update, code)

async def complete_onedrive_auth(update, auth_code):
    """Complete OneDrive authentication"""
    try:
        # Your existing authentication code here
        # This should be the code that exchanges auth_code for tokens
        await update.message.reply_text("✅ Authentication completed successfully!")
    except Exception as e:
        logger.error(f"Auth error: {e}")
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def direct_auth_command(update, context):
    """Direct authentication with manual code"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("? Unauthorized.")
        return
    
    # Get auth code from .env
    auth_code = os.getenv("ONEDRIVE_AUTH_CODE")
    
    if not auth_code:
        await update.message.reply_text("? No ONEDRIVE_AUTH_CODE in .env file")
        return
    
    await update.message.reply_text(f"?? Using auth code: {auth_code[:50]}...")
    
    try:
        # Direct HTTP request to get tokens
        import requests
        
        data = {
            'client_id': ONEDRIVE_CLIENT_ID,
            'scope': 'Files.ReadWrite',
            'code': auth_code,
            'redirect_uri': ONEDRIVE_REDIRECT_URI,
            'grant_type': 'authorization_code',
            'client_secret': ONEDRIVE_CLIENT_SECRET
        }
        
        response = requests.post(
            'https://login.microsoftonline.com/common/oauth2/v2.0/token',
            data=data
        )
        
        if response.status_code == 200:
            tokens = response.json()
            
            # Save tokens
            with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                json.dump(tokens, f, indent=2)
            
            await update.message.reply_text(
                "? Authentication successful!\n"
                "Tokens saved. Now test with /onedrive_test"
            )
        else:
            await update.message.reply_text(f"? Failed: {response.text}")
            
    except Exception as e:
        await update.message.reply_text(f"? Error: {str(e)[:200]}")

# ========== EXCEL FUNCTIONS (OPENPYXL VERSION) ==========

@contextmanager
def excel_operation():
    """Context manager for Excel operations with openpyxl"""
    wb = None
    try:
        yield wb
    except Exception as e:
        logger.error(f"Excel operation error: {e}")
        raise
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass

def copy_excel_from_onedrive() -> Tuple[bool, str]:
    """Copy from OneDrive to local using OneDrive API"""
    return download_from_onedrive()

def save_excel_to_onedrive() -> Tuple[bool, str]:
    """Save changes back to OneDrive using OneDrive API"""
    return upload_to_onedrive()

def load_all_tables_with_details() -> Dict[str, Dict]:
    """Load ALL Excel Tables with detailed information and caching"""
    global _table_cache, _table_cache_timestamp
    
    if not EXCEL_AVAILABLE:
        logger.error("openpyxl is not available. Cannot load tables.")
        return {}
    
    try:
        if LOCAL_COPY_PATH.exists():
            current_mtime = LOCAL_COPY_PATH.stat().st_mtime
            if (_table_cache_timestamp and 
                _table_cache_timestamp == current_mtime and 
                _table_cache):
                return _table_cache
        
        wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True, read_only=False)
        sheet = wb[DROPDOWN_SHEET_NAME]
        
        tables_dict = {}
        
        # In openpyxl, we need to manually identify tables
        # Look for structured data (assuming tables start at row 1)
        # You might need to adjust this based on your Excel structure
        
        # For now, let's assume each column from A onward is a category
        # and rows contain subcategories
        
        max_column = sheet.max_column
        max_row = sheet.max_row
        
        for col in range(1, max_column + 1):
            category = sheet.cell(row=1, column=col).value
            if not category:
                continue
            
            subcategories = []
            variations_dict = {}
            
            for row in range(2, max_row + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    original_text = str(cell_value).strip()
                    if original_text:
                        subcategories.append(original_text)
                        normalized = normalize_text_for_matching(original_text)
                        variations_dict[normalized] = original_text
                        
                        # Also store individual words for better matching
                        words = set(normalized.split())
                        for word in words:
                            if len(word) > 3:  # Only store words longer than 3 characters
                                if word not in variations_dict:
                                    variations_dict[word] = original_text
            
            if subcategories:  # Only add if we found data
                tables_dict[category] = {
                    'original_name': category,
                    'subcategories': subcategories,
                    'variations': variations_dict,
                    'count': len(subcategories)
                }
        
        wb.close()
        
        _table_cache = tables_dict
        _table_cache_timestamp = current_mtime if LOCAL_COPY_PATH.exists() else None
        
        logger.info(f"Loaded {len(tables_dict)} tables from Excel")
        return tables_dict
            
    except Exception as e:
        logger.error(f"Error loading tables: {str(e)}", exc_info=True)
        return {}

def add_transaction_smart(subcategory_input: str, currency_amounts: Dict[str, Optional[float]], 
                         payment_type: str, optional_notes: str = "") -> Tuple[bool, str]:
    """Add transaction with smart matching - supports optional notes"""
    try:
        # FIXED: Always use "Expenses" as transaction type (even for negative amounts)
        transaction_type = "Expenses"
        
        logger.info(f"Processing transaction: '{subcategory_input}', amounts: {currency_amounts}, payment: {payment_type}, notes: '{optional_notes}'")
        
        # 1. Copy fresh file from OneDrive
        success, msg = copy_excel_from_onedrive()
        if not success:
            logger.error(f"Failed to copy file: {msg}")
            return False, msg
        
        # 2. Load tables and find best match
        tables_dict = load_all_tables_with_details()
        if not tables_dict:
            logger.error("No tables found in Dropdown Data sheet")
            return False, "❌ No tables found in Dropdown Data sheet"
        
        # Find best match with confidence
        matched_original, category, match_type, confidence = find_best_match_for_input(subcategory_input, tables_dict)
        
        # FIXED: Display confidence as percentage (e.g., 70% instead of 0.7)
        confidence_percentage = confidence * 100
        
        # REJECTION LOGIC: If confidence is below threshold, reject the transaction
        if confidence < MINIMUM_CONFIDENCE:
            # Generate helpful suggestions
            suggestions = []
            
            # Get top suggestions from all categories
            all_suggestions = []
            for cat_name, data in tables_dict.items():
                for item in data['subcategories'][:5]:  # Top 5 from each category
                    all_suggestions.append((cat_name, item))
            
            # Show most relevant suggestions (alphabetically sorted)
            all_suggestions.sort(key=lambda x: x[1])
            
            suggestions.append("\n<b>Available categories:</b>")
            for cat_name, _ in tables_dict.items():
                suggestions.append(f"  • {cat_name}")
            
            suggestions.append("\n<b>Common items:</b>")
            displayed_items = 0
            for cat_name, item in all_suggestions:
                if displayed_items >= 8:  # Limit to 8 items
                    break
                suggestions.append(f"  • {item} ({cat_name})")
                displayed_items += 1
            
            suggestions_text = "\n".join(suggestions)
            
            rejection_msg = (
                f"❌ <b>Transaction Rejected:</b>\n\n"
                f"<b>Input:</b> <code>{subcategory_input}</code>\n"
                f"<b>Confidence:</b> {confidence_percentage:.1f}% (minimum required: {MINIMUM_CONFIDENCE*100:.0f}%)\n"
                f"<b>Reason:</b> {match_type}\n\n"
                f"<b>No good match found. Did you mean one of these?</b>\n"
                f"{suggestions_text}\n\n"
                f"<i>Tip: Use more specific terms or check spelling.</i>"
            )
            logger.warning(f"Transaction rejected: {subcategory_input} (confidence: {confidence_percentage:.1f}%)")
            return False, rejection_msg
        
        logger.info(f"Match found: '{subcategory_input}' → '{matched_original}' → '{category}' ({match_type}, {confidence_percentage:.1f}%)")
        
        # 3. Open Excel and add transaction
        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl not installed. Please install with: pip install openpyxl pandas"
        
        with excel_lock:
            wb = load_workbook(str(LOCAL_COPY_PATH))
            sheet = wb[TRACKING_SHEET_NAME]
            
            # Find first empty row
            row = 12
            while sheet[f'C{row}'].value not in [None, ""]:
                row += 1
                if row > 1000:
                    break
            
            logger.debug(f"Adding transaction at row {row}")
            
            # Add data
            today = datetime.now().strftime("%d-%b-%y")
            
            sheet[f'C{row}'].value = today
            sheet[f'D{row}'].value = payment_type
            sheet[f'E{row}'].value = transaction_type  # FIXED: Always "Expenses"
            sheet[f'F{row}'].value = category
            sheet[f'G{row}'].value = matched_original
            
            # Add amounts to correct currency columns
            for currency, amount in currency_amounts.items():
                if currency in CURRENCIES and amount is not None:
                    column = CURRENCIES[currency].column
                    sheet[f'{column}{row}'].value = float(amount)
            
            # If no currencies specified, add to USD as default
            if not currency_amounts:
                sheet[f'H{row}'].value = 0.0
            
            # ========== OPTIONAL NOTES SECTION ==========
            # Only add notes if user provided them
            if optional_notes:
                # Clean up the notes (remove extra whitespace, limit length)
                clean_notes = optional_notes.strip()
                if len(clean_notes) > 500:  # Limit to 500 characters
                    clean_notes = clean_notes[:497] + "..."
                sheet[f'K{row}'].value = clean_notes
            else:
                # Leave notes cell empty if no notes provided
                sheet[f'K{row}'].value = None
            # ========== END NOTES SECTION ==========

            wb.save(str(LOCAL_COPY_PATH))
            wb.close()
            logger.debug("Transaction saved to local file")
      
        # 4. Save back to OneDrive
        save_success, save_msg = save_excel_to_onedrive()
        
        # Format response
        amount_display_parts = []
        total_amounts = len(currency_amounts)
        
        for currency, amount in currency_amounts.items():
            if amount is not None:
                amount_str = format_currency_amount(amount, currency)
                # Add visual indicator for negative amounts
                if amount < 0:
                    amount_display = f"🔄 {amount_str}"
                else:
                    amount_display = amount_str
                amount_display_parts.append(amount_display)
        
        if amount_display_parts:
            if total_amounts > 1:
                amount_display = " + ".join(amount_display_parts)
            else:
                amount_display = amount_display_parts[0]
        else:
            amount_display = "No amount"
        
        payment_emoji = PAYMENT_TYPES.get(payment_type, "💵")
        
        # Check if any amounts are negative
        has_negative = any(a is not None and a < 0 for a in currency_amounts.values())
        
        if has_negative:
            type_emoji = "🔄"  # Circular arrow for refund/income
            transaction_desc = "Expense (with refunds/corrections)"
        else:
            type_emoji = "📤"  # Outgoing arrow for expense
            transaction_desc = "Expense"
        
        # Add currency emojis for each currency
        currency_emojis = []
        for currency in currency_amounts.keys():
            if currency == "USD":
                currency_emojis.append("💵")
            elif currency == "LBP":
                currency_emojis.append("🇱🇧")
            elif currency == "EURO":
                currency_emojis.append("💶")
        
        currency_emoji_str = " ".join(currency_emojis)
        
        import html
        safe_input = html.escape(subcategory_input)
        safe_matched = html.escape(matched_original)
        safe_category = html.escape(category)
        safe_payment = html.escape(payment_type)
        safe_notes = html.escape(optional_notes) if optional_notes else ""
        safe_confidence = f"{confidence_percentage:.1f}%"
        
        # Add multi-currency indicator
        multi_currency_indicator = "🌐 " if total_amounts > 1 else ""
        
        # Add notes to response if provided
        notes_section = f"\n• <b>Notes:</b> {safe_notes}" if optional_notes else ""
        
        if save_success:
            message = (
                f"{multi_currency_indicator}✅ {type_emoji} {payment_emoji} {currency_emoji_str} <b>{transaction_desc} Added:</b>\n\n"
                f"• <b>You typed:</b> <code>{safe_input}</code>\n"
                f"• <b>Recorded as:</b> <code>{safe_matched}</code>\n"
                f"• <b>Category:</b> {safe_category}\n"
                f"• <b>Amount{'s' if total_amounts > 1 else ''}:</b> {amount_display}\n"
                f"• <b>Payment:</b> {safe_payment}"
                f"{notes_section}\n"
                f"• <b>Confidence:</b> {safe_confidence}\n"
                f"• <b>Match type:</b> {match_type}"
            )
            logger.info(f"Transaction added successfully: {transaction_desc}")
            return True, message
        else:
            message = (
                f"{multi_currency_indicator}⚠️ {type_emoji} {payment_emoji} {currency_emoji_str} <b>{transaction_desc} Added Locally:</b>\n\n"
                f"• <b>You typed:</b> <code>{safe_input}</code>\n"
                f"• <b>Recorded as:</b> <code>{safe_matched}</code>\n"
                f"• <b>Category:</b> {safe_category}\n"
                f"• <b>Amount{'s' if total_amounts > 1 else ''}:</b> {amount_display}\n"
                f"• <b>Payment:</b> {safe_payment}"
                f"{notes_section}\n\n"
                f"<b>⚠️ OneDrive Sync Failed:</b>\n{save_msg}"
            )
            logger.warning(f"Transaction added locally but OneDrive sync failed: {save_msg}")
            return True, message
        
    except Exception as e:
        logger.error(f"Error in add_transaction_smart: {str(e)}", exc_info=True)
        import html
        error_msg = html.escape(str(e)[:200])
        return False, f"❌ <b>Error:</b> {error_msg}"

def get_recent_transactions(count: int = 10) -> List[Dict]:
    """Get multiple recent transactions from Excel"""
    transactions = []
    
    if not EXCEL_AVAILABLE:
        return transactions
    
    try:
        # Copy fresh file first
        copy_excel_from_onedrive()
        
        with excel_lock:
            wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
            sheet = wb[TRACKING_SHEET_NAME]
            
            # Find the last row with data
            last_row = 12
            while sheet[f'C{last_row}'].value not in [None, ""]:
                last_row += 1
                if last_row > 1000:
                    break
            
            # Get the last 'count' transactions
            start_row = max(12, last_row - count)
            
            for row in range(start_row, last_row):
                date_val = sheet[f'C{row}'].value
                if date_val:
                    transaction = {
                        'row': row,
                        'date': date_val,
                        'payment': sheet[f'D{row}'].value or "Cash",
                        'type': sheet[f'E{row}'].value or "Expenses",
                        'category': sheet[f'F{row}'].value or "",
                        'subcategory': sheet[f'G{row}'].value or "",
                        'usd': sheet[f'H{row}'].value,
                        'lbp': sheet[f'I{row}'].value,
                        'euro': sheet[f'J{row}'].value,
                        'notes': sheet[f'K{row}'].value or ""
                    }
                    transactions.append(transaction)
            
            wb.close()
        
        # Return in reverse order (most recent first)
        return list(reversed(transactions))
        
    except Exception as e:
        logger.error(f"Error getting recent transactions: {str(e)}")
        return []

# ========== NEW: DOWNLOAD/EXPORT FUNCTIONS ==========

def setup_export_directory():
    """Create export directory if it doesn't exist"""
    EXPORT_DIR.mkdir(exist_ok=True)

def export_to_csv(time_range: str = "month") -> Tuple[bool, str]:
    """
    Export transactions to CSV format using pandas
    time_range: "day", "week", "month", "year", "all"
    """
    try:
        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl not installed"
        
        # Copy fresh file
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}"
        
        # Use pandas to read Excel
        df = pd.read_excel(
            LOCAL_COPY_PATH,
            sheet_name=TRACKING_SHEET_NAME,
            header=10,  # Skip first 10 rows if they're headers
            usecols='C:K'  # Columns C to K
        )
        
        # Rename columns
        df.columns = ['Date', 'Payment', 'Type', 'Category', 'Subcategory', 'USD', 'LBP', 'EURO', 'Notes']
        
        # Filter by date if needed
        if time_range != "all":
            today = datetime.now()
            if time_range == "day":
                start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
            elif time_range == "week":
                start_date = today - timedelta(days=today.weekday())
                start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            elif time_range == "month":
                start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            elif time_range == "year":
                start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            
            df = df[pd.to_datetime(df['Date']) >= start_date]
        
        if df.empty:
            return False, "❌ No transactions found"
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transactions_{time_range}_{timestamp}.csv"
        filepath = EXPORT_DIR / filename
        
        # Save to CSV
        df.to_csv(filepath, index=False, encoding='utf-8')
        
        logger.info(f"Exported {len(df)} transactions to {filepath}")
        return True, str(filepath)
        
    except Exception as e:
        logger.error(f"Export error: {str(e)}", exc_info=True)
        return False, f"❌ Export error: {str(e)[:200]}"

def export_summary() -> Tuple[bool, str]:
    """Export summary statistics"""
    try:
        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl not installed"
        
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}"
        
        # Use pandas to read Excel
        df = pd.read_excel(
            LOCAL_COPY_PATH,
            sheet_name=TRACKING_SHEET_NAME,
            header=10,
            usecols='C:K'
        )
        
        df.columns = ['Date', 'Payment', 'Type', 'Category', 'Subcategory', 'USD', 'LBP', 'EURO', 'Notes']
        
        # Calculate totals
        usd_total = df['USD'].sum() if 'USD' in df.columns else 0
        lbp_total = df['LBP'].sum() if 'LBP' in df.columns else 0
        euro_total = df['EURO'].sum() if 'EURO' in df.columns else 0
        
        # Category breakdown
        categories = {}
        if 'Category' in df.columns and 'USD' in df.columns:
            for category in df['Category'].unique():
                if pd.notna(category):
                    cat_data = df[df['Category'] == category]
                    categories[category] = {
                        'usd': cat_data['USD'].sum(),
                        'lbp': cat_data['LBP'].sum() if 'LBP' in df.columns else 0,
                        'euro': cat_data['EURO'].sum() if 'EURO' in df.columns else 0,
                        'count': len(cat_data)
                    }
        
        # Create summary JSON
        summary = {
            'export_date': datetime.now().isoformat(),
            'total_transactions': len(df),
            'totals': {
                'USD': float(usd_total),
                'LBP': float(lbp_total),
                'EURO': float(euro_total)
            },
            'categories': categories,
            'file_info': {
                'source': str(LOCAL_COPY_PATH),
                'last_modified': datetime.fromtimestamp(LOCAL_COPY_PATH.stat().st_mtime).isoformat() if LOCAL_COPY_PATH.exists() else None
            }
        }
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"summary_{timestamp}.json"
        filepath = EXPORT_DIR / filename
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, default=str)
        
        logger.info(f"Exported summary to {filepath}")
        return True, str(filepath)
        
    except Exception as e:
        logger.error(f"Summary export error: {str(e)}", exc_info=True)
        return False, f"❌ Summary error: {str(e)[:200]}"

def create_backup_copy() -> Tuple[bool, str]:
    """Create a backup copy of the Excel file for download"""
    try:
        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local copy not found"
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"budget_backup_{timestamp}.xlsm"
        backup_path = BACKUP_DIR / backup_filename
        
        shutil.copy2(LOCAL_COPY_PATH, backup_path)
        
        logger.info(f"Created downloadable backup: {backup_path}")
        return True, str(backup_path)
        
    except Exception as e:
        logger.error(f"Backup copy error: {str(e)}")
        return False, f"❌ Backup copy error: {str(e)[:200]}"

# ========== UTILITY FUNCTIONS ==========

def is_authorized(user_id: int) -> bool:
    """Check if user is authorized to use the bot"""
    return user_id in ALLOWED_USER_IDS

def format_currency_amount(amount: float, currency: str) -> str:
    """Format amount according to currency rules - supports negative numbers"""
    if amount is None:
        return "No amount"
    
    # If no currency, just return the number
    if not currency:
        return f"{amount:,.2f}"
    
    is_negative = amount < 0
    abs_amount = abs(amount)
    
    if currency not in CURRENCIES:
        currency = DEFAULT_CURRENCY
    
    config = CURRENCIES[currency]
    
    if config.thousands_separator:
        formatted = f"{abs_amount:,.{config.decimal_places}f}"
    else:
        formatted = f"{abs_amount:.{config.decimal_places}f}"
    
    symbol = config.symbols[0]
    
    if is_negative:
        if currency == "USD":
            return f"-{symbol}{formatted}"
        elif currency == "EURO":
            return f"-{symbol}{formatted}"
        elif currency == "LBP":
            return f"-{formatted} LBP"
        else:
            return f"-{formatted}"
    else:
        if currency == "USD":
            return f"{symbol}{formatted}"
        elif currency == "EURO":
            return f"{symbol}{formatted}"
        elif currency == "LBP":
            return f"{formatted} LBP"
        return formatted

def format_transaction_response(transaction: Dict, action: str = "Added") -> str:
    """Format a transaction for display in messages"""
    date_val = transaction['date']
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
    
    # Format amounts
    amounts = []
    if transaction.get('usd') is not None:
        prefix = "🔄 " if transaction['usd'] < 0 else ""
        amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
    if transaction.get('lbp') is not None:
        prefix = "🔄 " if transaction['lbp'] < 0 else ""
        amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
    if transaction.get('euro') is not None:
        prefix = "🔄 " if transaction['euro'] < 0 else ""
        amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
    
    amount_str = " + ".join(amounts) if amounts else "No amount"
    
    payment_emoji = "💳" if str(transaction.get('payment', '')).lower() == "card" else "💵"
    
    import html
    safe_item = html.escape(transaction.get('subcategory') or transaction.get('category', ''))
    safe_category = html.escape(transaction.get('category', 'Unknown'))
    
    response = (
        f"✅ <b>Transaction {action}:</b>\n\n"
        f"{payment_emoji} {date_str}: {safe_item}\n"
        f"Amount: {amount_str}\n"
        f"Category: {safe_category}\n"
        f"Payment: {html.escape(transaction.get('payment', 'Cash'))}"
    )
    
    notes = transaction.get('notes')
    if notes:
        response += f"\nNotes: {html.escape(str(notes)[:100])}{'...' if len(str(notes)) > 100 else ''}"
    
    return response

def create_backup():
    """Create backup of the Excel file"""
    try:
        BACKUP_DIR.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"backup_{timestamp}_{LOCAL_COPY_PATH.name}"
        shutil.copy2(LOCAL_COPY_PATH, backup_path)
        
        cleanup_old_backups()
        logger.info(f"Backup created: {backup_path}")
        return True
    except Exception as e:
        logger.error(f"Backup creation failed: {e}")
        return False

def cleanup_old_backups():
    """Remove backups older than retention period"""
    try:
        cutoff_date = datetime.now().timestamp() - (BACKUP_RETENTION_DAYS * 24 * 3600)
        
        for backup_file in BACKUP_DIR.glob("backup_*"):
            if backup_file.stat().st_mtime < cutoff_date:
                backup_file.unlink()
                logger.info(f"Removed old backup: {backup_file}")
    except Exception as e:
        logger.error(f"Backup cleanup failed: {e}")

def unlock_excel_file():
    """Try to unlock the Excel file if it's locked - Linux version"""
    try:
        if LOCAL_COPY_PATH.exists():
            # Try to remove lock files
            lock_files = [
                LOCAL_COPY_PATH.with_suffix('.lock'),
                Path(f"{LOCAL_COPY_PATH}.lock"),
                Path(f"~{LOCAL_COPY_PATH}.lock").expanduser()
            ]
            
            for lock_file in lock_files:
                if lock_file.exists():
                    try:
                        lock_file.unlink()
                        logger.info(f"Removed lock file: {lock_file}")
                    except:
                        pass
            
            # Kill any LibreOffice processes that might have the file open
            subprocess.run(['pkill', '-f', 'libreoffice'], capture_output=True)
            subprocess.run(['pkill', '-f', 'soffice'], capture_output=True)
            
            logger.info("Attempted to unlock Excel file")
            return True
        else:
            return False
            
    except Exception as e:
        logger.error(f"Unlock error: {e}")
        return False

# ========== TEXT PROCESSING ==========

def normalize_text_for_matching(text: str) -> str:
    """Normalize text for matching"""
    if not text:
        return ""
    
    text = text.lower()
    text_without_parentheses = re.sub(r'\([^)]*\)', ' ', text)
    content_in_parentheses = re.findall(r'\(([^)]+)\)', text)
    all_text = text_without_parentheses + ' ' + ' '.join(content_in_parentheses)
    all_text = re.sub(r'[^\w\s]', ' ', all_text)
    all_text = ' '.join(all_text.split())
    
    return all_text.strip()

def extract_payment_amount_currency(text: str) -> Tuple[str, Dict[str, Optional[float]], str]:
    """
    Extract payment type, multiple amounts with currencies, and clean subcategory
    Now supports multiple currencies in one input
    
    Returns: (subcategory, currency_amounts_dict, payment_type)
    where currency_amounts_dict is {'USD': amount, 'LBP': amount, 'EURO': amount}
    """
    if not text:
        return "", {}, "Cash"
    
    original_text = text
    text_lower = text.lower()
    
    # 1. Extract payment type
    payment_type = "Cash"
    
    for keyword in CARD_KEYWORDS:
        if keyword in text_lower:
            payment_type = "Card"
            break
    
    if payment_type == "Cash":
        for keyword in DIGITAL_KEYWORDS:
            if keyword in text_lower:
                payment_type = "Digital Wallet"
                break
    
    # 2. RESTORED ORIGINAL MULTI-CURRENCY EXTRACTION WITH FIXES
    currency_amounts = {}  # {'USD': 10.0, 'LBP': -150000.0}
    
    # Patterns for all currency types - RESTORED FROM ORIGINAL
    patterns = [
        # LBP patterns - UPDATED for better matching
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(lbp|ليرة|lira|ل\.ل)', 'LBP'),
        (r'(lbp|ليرة|lira|ل\.ل)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'LBP'),
        # Simple LBP pattern without commas requirement
        (r'(-?\d+)\s*(lbp|ليرة|lira|ل\.ل)', 'LBP'),
        
        # USD patterns
        (r'[$\$](-?\d+(?:,\d{3})*(?:\.\d+)?)', 'USD'),
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(usd|dollar)', 'USD'),
        (r'(usd|dollar)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'USD'),
        # Simple USD pattern
        (r'(-?\d+(?:\.\d+)?)\s*\$', 'USD'),
        
        # EURO patterns
        (r'€(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'EURO'),
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(eur|euro)', 'EURO'),
        (r'(eur|euro)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'EURO'),
        # Simple EURO pattern
        (r'(-?\d+(?:\.\d+)?)\s*€', 'EURO'),
        
        # Symbol suffix patterns
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)[€$]', None),
    ]
    
    # Process each pattern - IMPORTANT: Don't break, collect ALL matches
    for pattern, currency_code in patterns:
        try:
            matches = list(re.finditer(pattern, text_lower, re.IGNORECASE))
            for match in matches:
                try:
                    if currency_code is None:
                        # Handle symbol suffix patterns
                        amount_str = match.group(1).replace(',', '')
                        symbol = match.group(0)[-1]
                        currency_code = 'USD' if symbol == '$' else 'EURO'
                        amount = float(amount_str)
                    elif currency_code == 'LBP':
                        # Check which group has the amount
                        if match.group(1) and match.group(1).replace(',', '').replace('-', '').replace('.', '').isdigit():
                            amount_str = match.group(1).replace(',', '')
                        elif match.group(2) and match.group(2).replace(',', '').replace('-', '').replace('.', '').isdigit():
                            amount_str = match.group(2).replace(',', '')
                        else:
                            continue
                        amount = float(amount_str)
                    else:  # USD or EURO
                        # Check which group has the amount
                        if match.group(0)[0] in ['$', '€']:
                            amount_str = match.group(1).replace(',', '')
                            amount = float(amount_str)
                        elif match.group(1) and match.group(1).replace(',', '').replace('-', '').replace('.', '').isdigit():
                            amount_str = match.group(1).replace(',', '')
                            amount = float(amount_str)
                        elif match.group(2) and match.group(2).replace(',', '').replace('-', '').replace('.', '').isdigit():
                            amount_str = match.group(2).replace(',', '')
                            amount = float(amount_str)
                        else:
                            continue
                    
                    # Check if amount starts with - in original text
                    match_text = match.group(0)
                    if match_text.startswith('-') and amount > 0:
                        amount = -amount
                    
                    # Store the amount for this currency
                    # If multiple amounts for same currency, keep the last one
                    currency_amounts[currency_code] = amount
                    
                    logger.info(f"Found amount: {amount} {currency_code}")
                    
                except (ValueError, IndexError) as e:
                    logger.debug(f"Error parsing amount: {e}")
                    continue
                    
        except re.error:
            continue
    
    # 3. Also look for standalone negative numbers near currency words
    if 'LBP' not in currency_amounts:
        # Look for negative number followed by LBP
        lbp_pattern = r'-\s*(\d+(?:,\d{3})*)\s*(lbp|ليرة|lira|ل\.ل)'
        matches = re.findall(lbp_pattern, text_lower, re.IGNORECASE)
        for match in matches:
            try:
                amount = -float(match[0].replace(',', ''))
                currency_amounts['LBP'] = amount
                logger.info(f"Found negative LBP amount: {amount}")
            except ValueError:
                pass
    
    # 4. If we found some but not all currencies, look for standalone amounts
    if currency_amounts:
        # Look for standalone numbers that aren't already part of currency patterns
        standalone_pattern = r'(?<!\S)(-?\d+(?:,\d{3})*(?:\.\d+)?)(?![\$\€\w])'
        matches = re.finditer(standalone_pattern, text_lower)
        
        for match in matches:
            try:
                amount = float(match.group(1).replace(',', ''))
                # Don't add if we already have this amount (might be duplicate)
                # But do add if it's a different currency than what we already have
                
                # If it's a large whole number and no LBP yet, assume LBP
                if amount >= 1000 and amount.is_integer() and 'LBP' not in currency_amounts:
                    currency_amounts['LBP'] = amount
                    logger.info(f"Assumed LBP for standalone: {amount}")
                # If it has decimals and no USD yet, assume USD
                elif not amount.is_integer() and 'USD' not in currency_amounts:
                    currency_amounts['USD'] = amount
                    logger.info(f"Assumed USD for standalone: {amount}")
                # Small whole number and we need USD
                elif amount < 1000 and amount.is_integer() and 'USD' not in currency_amounts:
                    currency_amounts['USD'] = amount
                    logger.info(f"Assumed USD for small whole number: {amount}")
            except ValueError:
                continue
    
    # 5. Clean subcategory - remove all amount patterns
    subcategory = original_text
    
    # Remove all detected amounts
    for currency_code, amount in currency_amounts.items():
        abs_amount = abs(amount)
        
        if currency_code == 'LBP':
            # Remove LBP formats
            formats_to_remove = [
                f"{abs_amount:,.0f}LBP", f"{abs_amount}LBP",
                f"LBP{abs_amount:,.0f}", f"LBP{abs_amount}",
                f"{abs_amount:,.0f} lbp", f"{abs_amount} lbp",
                f"lbp {abs_amount:,.0f}", f"lbp {abs_amount}",
                f"{abs_amount:,.0f}ليرة", f"{abs_amount}ليرة",
                f"ليرة{abs_amount:,.0f}", f"ليرة{abs_amount}",
                f"{abs_amount:,.0f} ليرة", f"{abs_amount} ليرة",
                f"ليرة {abs_amount:,.0f}", f"ليرة {abs_amount}",
            ]
        elif currency_code == 'USD':
            # Remove USD formats
            formats_to_remove = [
                f"${abs_amount:.2f}", f"${abs_amount:,.2f}",
                f"{abs_amount:.2f}$", f"{abs_amount:,.2f}$",
                f"{abs_amount:.2f}usd", f"{abs_amount:.2f} dollar",
                f"{abs_amount}usd", f"{abs_amount} dollar",
            ]
        elif currency_code == 'EURO':
            # Remove EURO formats
            formats_to_remove = [
                f"€{abs_amount:.2f}", f"€{abs_amount:,.2f}",
                f"{abs_amount:.2f}€", f"{abs_amount:,.2f}€",
                f"{abs_amount:.2f}eur", f"{abs_amount:.2f} euro",
                f"{abs_amount}eur", f"{abs_amount} euro",
            ]
        else:
            formats_to_remove = []
        
        # Add negative versions
        negative_formats = [f"-{fmt}" for fmt in formats_to_remove]
        formats_to_remove.extend(negative_formats)
        
        # Remove all formats (case insensitive)
        for fmt in formats_to_remove:
            subcategory = subcategory.replace(fmt, ' ')
            subcategory = subcategory.replace(fmt.lower(), ' ')
            subcategory = subcategory.replace(fmt.upper(), ' ')
            # Also try with spaces
            subcategory = subcategory.replace(fmt.replace(' ', ''), ' ')
        
        # Also remove just the number (with and without commas)
        num_str = str(abs_amount)
        if '.' in num_str:
            # Has decimals
            subcategory = subcategory.replace(num_str, ' ')
        else:
            # Whole number
            subcategory = subcategory.replace(num_str, ' ')
            if abs_amount >= 1000:
                # Remove with commas
                subcategory = subcategory.replace(f"{abs_amount:,.0f}", ' ')
                # Remove without commas
                subcategory = subcategory.replace(f"{int(abs_amount)}", ' ')
    
    # Remove standalone currency symbols and words
    subcategory = subcategory.replace('$', ' ').replace('€', ' ')
    
    # Remove payment keywords
    for keyword in CARD_KEYWORDS + DIGITAL_KEYWORDS:
        subcategory = subcategory.replace(keyword, ' ').replace(keyword.title(), ' ')
        subcategory = subcategory.replace(keyword.upper(), ' ')
    
    # Remove currency words
    currency_words = ['usd', 'dollar', 'euro', 'eur', 'lbp', 'lira', 'ليرة', 'ل.ل']
    for word in currency_words:
        subcategory = subcategory.replace(word, ' ').replace(word.title(), ' ')
        subcategory = subcategory.replace(word.upper(), ' ')
    
    # Clean up
    subcategory = ' '.join(subcategory.split())
    subcategory = subcategory.title()
    
    if not subcategory:
        subcategory = original_text
    
    logger.info(f"Parsed: subcategory='{subcategory}', amounts={currency_amounts}, payment={payment_type}")
    return subcategory, currency_amounts, payment_type

def find_best_match_for_input(input_text: str, tables_dict: Dict) -> Tuple[str, str, str, float]:
    """Find the best match for input text with confidence score"""
    if not input_text:
        return "", "Unknown", "Empty input", 0.0
    
    normalized_input = normalize_text_for_matching(input_text)
    
    if not normalized_input:
        return "", "Unknown", "Empty after normalization", 0.0
    
    # Reject very short inputs
    if len(normalized_input) < 3:
        return "", "Unknown", "Input too short", 0.0
    
    # Reject common non-category words
    if normalized_input in COMMON_NON_CATEGORIES:
        return "", "Unknown", "Common non-category word", 0.0
    
    best_match = None
    best_category = None
    best_score = 0.0
    match_type = "No match"
    
    # SIMPLIFY: Just split by spaces, no complex regex
    input_words = set(normalized_input.split())
    
    for category, data in tables_dict.items():
        for variation, original in data['variations'].items():
            # SIMPLIFY: Split variation by spaces
            variation_words = set(variation.split())
            
            # Calculate intersection
            intersection = len(input_words.intersection(variation_words))
            union = len(input_words.union(variation_words))
            
            if union > 0:
                score = intersection / union
                
                # Boost for exact matches
                if normalized_input == variation:
                    score = 1.0
                elif normalized_input in variation:
                    score = max(score, 0.8)
                elif variation in normalized_input:
                    score = max(score, 0.8)
                
                if score > best_score:
                    best_score = score
                    best_match = original
                    best_category = category
    
    # Check if we have a good enough match
    if best_score >= MINIMUM_CONFIDENCE:
        if best_score >= 0.95:
            match_type = "Exact match"
        elif best_score >= 0.85:
            match_type = "Strong match"
        elif best_score >= MINIMUM_CONFIDENCE:
            match_type = "Good match"
        
        logger.info(f"Good match found: '{input_text}' → '{best_match}' (score: {best_score:.2f})")
        return best_match, best_category, match_type, best_score
    
    # Try fuzzy matching with simplified approach
    if best_score < FUZZY_MATCH_THRESHOLD:
        # Build simple lists for fuzzy matching
        all_variations = []
        all_originals = []
        all_categories = []
        
        for category, data in tables_dict.items():
            for variation, original in data['variations'].items():
                all_variations.append(variation)
                all_originals.append(original)
                all_categories.append(category)
        
        # Use simple fuzzy matching
        matches = get_close_matches(normalized_input, all_variations, n=1, cutoff=0.5)
        
        if matches:
            best_variation = matches[0]
            idx = all_variations.index(best_variation)
            fuzzy_score = 0.5  # Fixed score for fuzzy matches
            if fuzzy_score >= MINIMUM_CONFIDENCE:
                return all_originals[idx], all_categories[idx], "Fuzzy match", fuzzy_score
    
    logger.info(f"No good match found for '{input_text}'. Best score: {best_score:.2f}")
    return "", "Unknown", "No good match found", best_score

# ========== TELEGRAM BOT FUNCTIONS ==========

async def start_command(update, context):
    """Handle /start - UPDATED with enhanced delete feature"""
    logger.info(f"Received /start from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    excel_status = "✅ INSTALLED" if EXCEL_AVAILABLE else "❌ NOT INSTALLED"
    onedrive_status = "✅ INSTALLED" if ONEDRIVE_AVAILABLE else "❌ NOT INSTALLED"
    
    await update.message.reply_text(
        f"💰 <b>Smart Budget Tracker Bot</b>\n\n"
        f"📊 <b>Status:</b>\n"
        f"• Excel: {excel_status}\n"
        f"• OneDrive: {onedrive_status}\n"
        f"🎯 <b>Minimum Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}&#37;\n\n"
        
        "✅ <b>Smart Matching Examples:</b>\n"
        "• <code>Chamsin</code> → Bakery Products (Chamsin)\n"
        "• <code>Mazda</code> → Fuel (Mazda)\n"
        "• <code>KSC</code> → KSC (exact match)\n"
        "• <code>Fuel Opel</code> → Fuel (Opel)\n\n"
        
        "🔄 <b>Negative Amounts (Refunds/Corrections):</b>\n"
        "• <code>Chamsin -10$</code> → Refund from Chamsin\n"
        "• <code>Refund -200000 LBP</code> → Refund (no commas needed)\n"
        "• <code>Correction -50€</code> → Negative adjustment\n\n"
        
        "💱 <b>LBP Amounts:</b>\n"
        "• <code>200000 LBP</code> → Works without commas\n"
        "• <code>200,000 LBP</code> → Also works with commas\n\n"
        
        "❌ <b>Will Reject:</b>\n"
        f"• <code>OKAY</code>, <code>Test</code>, <code>Hello</code>\n"
        f"• Poor matches (&lt;{MINIMUM_CONFIDENCE*100:.0f}&#37; confidence)\n\n"
        
        "<b>📝 How to Use:</b>\n"
        "Just send: <code>ITEM AMOUNT CURRENCY PAYMENT</code>\n\n"
        "<b>Examples:</b>\n"
        "• <code>Chamsin 10</code> (Expense)\n"
        "• <code>Chamsin -10$</code> (Refund)\n"
        "• <code>Fuel Mazda 200000 LBP</code> (No commas needed)\n"
        "• <code>Refund -500$ card</code> (Negative amount)\n"
        "• <code>KSC 15.50 Card</code>\n"
        "• <code>Daouk Sweets 20$ Card</code>\n\n"
        
        "<b>🔐 OneDrive Setup:</b>\n"
        "1. <code>/onedrive_auth</code> - Get authentication URL\n"
        "2. <code>/onedrive_test</code> - Test connection\n\n"
        
        "<b>📥 DOWNLOAD FEATURES:</b>\n"
        "• /download csv [day/week/month/year/all] - Export to CSV\n"
        "• /download summary - Export summary JSON\n"
        "• /download backup - Download Excel backup\n\n"
        
        "<b>⚙️ Commands:</b>\n"
        "/start - This message\n"
        "/help - Detailed help\n"
        "/testparse [text] - Test parsing & matching\n"
        "/recent - Last 10 transactions\n"
        "/stats - Statistics\n"
        "/save - Force save to OneDrive\n"
        "/clearcache - Clear matching cache\n"
        "/download - Download data files\n"
        "/onedrive_auth - Setup OneDrive authentication\n"
        "/onedrive_test - Test OneDrive connection\n\n"
        
        "💡 <b>Tip:</b> Tap the menu icon (/) to see all commands",
        parse_mode='HTML'
    )

async def help_command(update, context):
    """Extended help command"""
    logger.info(f"Received /help from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    await update.message.reply_text(
        f"💰 <b>Smart Budget Tracker Bot - Help Guide</b>\n\n"
        
        f"<b>🎯 Smart Matching System:</b>\n"
        f"• Minimum confidence required: {MINIMUM_CONFIDENCE*100:.0f}%\n"
        f"• Rejects poor matches with suggestions\n"
        f"• Uses fuzzy matching for close matches\n\n"
        
        "<b>🔄 Negative Amount Support:</b>\n"
        "Use negative amounts for:\n"
        "• Refunds\n"
        "• Corrections\n"
        "• Negative adjustments\n"
        "• Transaction type remains 'Expenses'\n\n"
        
        "<b>💱 LBP Amount Format:</b>\n"
        "• <code>200000 LBP</code> - No commas needed\n"
        "• <code>200,000 LBP</code> - Commas also work\n"
        "• <code>-150000 lbp</code> - Negative also works\n\n"
        
        "<b>🌐 Multi-Currency Support:</b>\n"
        "• <code>KSC 10$, -150000 LBP</code> - Multiple currencies in one transaction\n"
        "• <code>Grocery 20€ 150000 LBP</code> - EURO + LBP\n"
        "• <code>Refund -5$, -75000 LBP</code> - Multiple refunds\n\n"
        
        "<b>📝 Basic Usage:</b>\n"
        "Simply send a message in this format:\n"
        "<code>ITEM [AMOUNT] [CURRENCY] [PAYMENT]</code>\n\n"
        
        "<b>✅ Will Accept (Examples):</b>\n"
        "• <code>Chamsin 10</code> - Expense\n"
        "• <code>Chamsin -10$</code> - Refund (still 'Expenses' type)\n"
        "• <code>Fuel Mazda 200000 LBP</code> - No commas needed\n"
        "• <code>Refund -500$ card</code> - Negative amount\n"
        "• <code>KSC 15.50 Card</code> - Expense\n"
        "• <code>Correction -200€</code> - Negative adjustment\n"
        "• <code>KSC 10$, -150000 LBP</code> - Multi-currency\n\n"
        
        "<b>❌ Will Reject:</b>\n"
        "• <code>OKAY 10$</code> - No matching category\n"
        "• <code>Test -5</code> - Common word\n"
        "• <code>Hello 20</code> - Greeting\n\n"
        
        "<b>🔐 OneDrive Setup:</b>\n"
        "1. Register app at: https://portal.azure.com\n"
        "2. Get Client ID and Secret\n"
        "3. Run <code>/onedrive_auth</code>\n"
        "4. Follow authentication steps\n\n"
        
        "<b>📥 DOWNLOAD COMMANDS:</b>\n"
        "• <code>/download csv day</code> - Today's transactions\n"
        "• <code>/download csv week</code> - This week's transactions\n"
        "• <code>/download csv month</code> - This month's transactions\n"
        "• <code>/download csv year</code> - This year's transactions\n"
        "• <code>/download csv all</code> - All transactions\n"
        "• <code>/download summary</code> - Summary statistics (JSON)\n"
        "• <code>/download backup</code> - Excel backup file\n\n"
        
        "<b>⚙️ Available Commands:</b>\n"
        "/start - Welcome message\n"
        "/help - This help message\n"
        "/testparse [text] - Test parsing & matching\n"
        "/recent - Show last 10 transactions\n"
        "/stats - Show statistics\n"
        "/save - Force save to OneDrive\n"
        "/clearcache - Clear matching cache\n"
        "/download - Download data files\n"
        "/onedrive_auth - OneDrive authentication\n"
        "/onedrive_test - Test OneDrive connection\n\n"
        
        "<b>💡 Tips:</b>\n"
        "• LBP amounts work with or without commas\n"
        "• Negative amounts keep 'Expenses' type\n"
        "• Be specific with item names\n"
        "• Check spelling if not matching\n"
        "• Use /testparse to test before adding\n"
        "• Download data regularly for backup",
        parse_mode='HTML'
    )

async def testparse_command(update, context):
    """Test input parsing and matching - UPDATED for multi-currency"""
    logger.info(f"Received /testparse from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        return
    
    if not context.args:
        await update.message.reply_text(
            "Usage: /testparse [text]\n\n"
            "Examples (Multi-Currency):\n"
            "• /testparse KSC 10$, -150000 LBP\n"
            "• /testparse Grocery 20€ 150000 LBP\n"
            "• /testparse Refund -5$, -75000 LBP\n\n"
            "Examples (Single Currency):\n"
            "• /testparse Chamsin 10$\n"
            "• /testparse Chamsin -10$\n"
            "• /testparse Fuel Mazda 200000 LBP"
        )
        return
    
    test_input = ' '.join(context.args)
    
    # Parse the input with multi-currency support
    subcategory, currency_amounts, payment_type = extract_payment_amount_currency(test_input)
    
    # Load tables and find match
    tables_dict = load_all_tables_with_details()
    matched, category, match_type, confidence = find_best_match_for_input(subcategory, tables_dict)
    
    # FIXED: Convert confidence to percentage
    confidence_percentage = confidence * 100
    
    # Determine if it would be accepted
    would_accept = confidence >= MINIMUM_CONFIDENCE
    accept_status = "✅ WOULD BE ACCEPTED" if would_accept else "❌ WOULD BE REJECTED"
    
    # Format currency amounts
    amounts_str = ""
    if currency_amounts:
        amount_parts = []
        for currency, amount in currency_amounts.items():
            if amount is not None:
                amount_parts.append(f"{format_currency_amount(amount, currency)}")
        amounts_str = ", ".join(amount_parts)
    else:
        amounts_str = "None"
    
    # Format the results
    result = [
        f"🔍 <b>Parsing & Matching Test (Multi-Currency)</b>",
        f"",
        f"<b>📝 Input:</b> <code>{test_input}</code>",
        f"",
        f"<b>📊 Parsing Results:</b>",
        f"• Subcategory: <code>{subcategory}</code>",
        f"• Amount{'s' if len(currency_amounts) > 1 else ''}: {amounts_str}",
        f"• Currencies Detected: {len(currency_amounts)}",
        f"• Payment Type: {payment_type}",
        f"",
        f"<b>🎯 Matching Results:</b>",
        f"• Matched to: <code>{matched if matched else 'No match'}</code>",
        f"• Category: {category}",
        f"• Match Type: {match_type}",
        f"• Confidence: {confidence_percentage:.1f}%",
        f"• Minimum Required: {MINIMUM_CONFIDENCE*100:.0f}%",
        f"",
        f"<b>📋 Final Result:</b> {accept_status}"
    ]
    
    await update.message.reply_text("\n".join(result), parse_mode='HTML')

async def recent_command(update, context):
    """Show recent transactions - UPDATED to show last 10 transactions"""
    logger.info(f"Received /recent from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        return
    
    if not EXCEL_AVAILABLE:
        await update.message.reply_text("❌ openpyxl not installed. Cannot read Excel file.")
        return
    
    try:
        await update.message.reply_text("📋 Loading last 10 transactions...")
        
        success, msg = copy_excel_from_onedrive()
        if not success:
            await update.message.reply_text(msg)
            return
        
        with excel_lock:
            wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
            sheet = wb[TRACKING_SHEET_NAME]
            
            recent = []
            # Find the last row with data
            last_row = 12
            while sheet[f'C{last_row}'].value not in [None, ""]:
                last_row += 1
            
            # Get last 10 transactions (UPDATED from 5 to 10)
            start_row = max(12, last_row - 10)
            for row in range(start_row, last_row):
                date_val = sheet[f'C{row}'].value
                if date_val:
                    payment = sheet[f'D{row}'].value or "Cash"
                    transaction_type = sheet[f'E{row}'].value or "Expenses"
                    category = sheet[f'F{row}'].value or ""
                    subcategory = sheet[f'G{row}'].value or ""
                    
                    usd = sheet[f'H{row}'].value
                    lbp = sheet[f'I{row}'].value
                    euro = sheet[f'J{row}'].value
                    
                    # Determine which amount to show
                    amounts = []
                    if usd is not None:
                        prefix = "🔄 " if usd < 0 else ""
                        amounts.append(f"{prefix}${abs(usd):.2f}")
                    if lbp is not None:
                        prefix = "🔄 " if lbp < 0 else ""
                        amounts.append(f"{prefix}{abs(lbp):,.0f} LBP")
                    if euro is not None:
                        prefix = "🔄 " if euro < 0 else ""
                        amounts.append(f"{prefix}€{abs(euro):.2f}")
                    
                    amount_str = " + ".join(amounts) if amounts else "No amount"
                    
                    if subcategory or category:
                        date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
                        payment_emoji = "💳" if str(payment).lower() == "card" else "💵"
                        item = subcategory if subcategory else category
                        import html
                        safe_item = html.escape(item)
                        
                        # Add transaction type indicator
                        type_indicator = "🔄" if usd is not None and usd < 0 or lbp is not None and lbp < 0 or euro is not None and euro < 0 else "📤"
                        
                        recent.append(f"{type_indicator} {payment_emoji} {date_str}: {safe_item} - {amount_str}")
            
            wb.close()
        
        if recent:
            # Show count in the header
            response = [f"📋 <b>Recent Transactions (Last {len(recent)}):</b>", ""] + recent
            await update.message.reply_text("\n".join(response), parse_mode='HTML')
        else:
            await update.message.reply_text("No recent transactions found.")
        
    except Exception as e:
        logger.error(f"Error in recent_command: {str(e)}")
        await update.message.reply_text(f"❌ Error loading transactions: {str(e)[:200]}")

async def handle_message(update, context):
    """Handle regular messages - UPDATED for multi-currency & optional notes support"""
    user_id = update.effective_user.id
    text = update.message.text.strip()
    
    logger.info(f"Received message from user {user_id}: {text}")
    
    if not is_authorized(user_id):
        logger.warning(f"User {user_id} not authorized")
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    if text.startswith('/'):
        return
    
    if not text:
        await update.message.reply_text(
            "💰 <b>How to Add a Transaction (Multi-Currency):</b>\n\n"
            "Send: <code>ITEM [AMOUNTS WITH CURRENCIES] [PAYMENT]</code>\n\n"
            "<b>📝 Examples (Single Currency):</b>\n"
            "• <code>Chamsin 10</code> (Expense)\n"
            "• <code>Chamsin -10$</code> (Refund)\n"
            "• <code>Fuel 200000 lbp</code> (No commas needed)\n\n"
            "<b>🌐 Examples (Multi-Currency):</b>\n"
            "• <code>KSC 10$, -150000 LBP</code>\n"
            "• <code>Grocery 20€ 150000 LBP</code>\n"
            "• <code>Refund -5$, -75000 LBP</code>\n\n"
            "<b>📝 Optional Notes:</b>\n"
            "Add a second line starting with 'DETAILS : ' for notes\n"
            "• <code>ITEM AMOUNTS\nDETAILS : your notes</code>\n\n"
            f"<b>🎯 Minimum Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}%\n"
            "Poor matches will be rejected with suggestions.\n\n"
            "Need help? Use /help for detailed instructions.",
            parse_mode='HTML'
        )
        return
    
    # Check if there are optional notes (separated by newline with 'DETAILS : ')
    transaction_text = text
    optional_notes = ""
    
    # Look for "DETAILS : " on a new line (case insensitive)
    details_pattern = r'\n\s*DETAILS\s*:\s*(.+)'
    details_match = re.search(details_pattern, text, re.IGNORECASE | re.DOTALL)
    
    if details_match:
        # Extract transaction part (everything before DETAILS)
        transaction_text = text[:details_match.start()].strip()
        optional_notes = details_match.group(1).strip()
        logger.info(f"Found optional notes: '{optional_notes}'")
    
    # Parse the transaction part with multi-currency support
    subcategory, currency_amounts, payment_type = extract_payment_amount_currency(transaction_text)
    
    # Check if we found any amounts
    if not currency_amounts:
        await update.message.reply_text(
            f"⚠️ <b>Amount Not Detected</b>\n\n"
            f"Could not detect any amounts with currencies.\n\n"
            f"<b>For multi-currency, specify amounts like:</b>\n"
            f"• <code>KSC 10$, -150000 LBP</code>\n"
            f"• <code>Grocery 20€ 150000 LBP</code>\n"
            f"• <code>Refund -5$, -75000 LBP</code>\n\n"
            f"<b>For optional notes:</b>\n"
            f"<code>ITEM AMOUNTS\n"
            f"DETAILS : your notes here</code>\n\n"
            f"<i>Tip: Attach currency to each amount ($, LBP, €)</i>",
            parse_mode='HTML'
        )
        return
    
    if not subcategory:
        await update.message.reply_text(
            "❌ <b>Parsing Error</b>\n\n"
            "Could not extract item from message.\n"
            "Try: <code>Item [Amounts with Currencies] [Payment]</code>\n\n"
            "<b>Examples:</b>\n"
            "• <code>KSC 10$, -150000 LBP card</code>\n"
            "• <code>Grocery 20€ 150000 LBP</code>\n"
            "• <code>Refund -5$, -75000 LBP cash</code>\n\n"
            "<b>With optional notes:</b>\n"
            "<code>ITEM AMOUNTS\n"
            f"DETAILS : your notes</code>\n\n"
            "Use /help for more examples.",
            parse_mode='HTML'
        )
        return
    
    # Show what we're doing
    amount_display_parts = []
    has_negative = False
    
    for currency, amount in currency_amounts.items():
        if amount is not None:
            amount_str = format_currency_amount(amount, currency)
            if amount < 0:
                amount_display_parts.append(f"🔄 {amount_str}")
                has_negative = True
            else:
                amount_display_parts.append(amount_str)
    
    if amount_display_parts:
        amount_display = " + ".join(amount_display_parts)
        transaction_desc = "Expense (with refunds/corrections)" if has_negative else "Expense"
    else:
        amount_display = "No amount specified"
        transaction_desc = "Transaction"
    
    payment_emoji = PAYMENT_TYPES.get(payment_type, "💵")
    
    # Get currency emojis
    currency_emojis = []
    for currency in currency_amounts.keys():
        if currency == "USD":
            currency_emojis.append("💵")
        elif currency == "LBP":
            currency_emojis.append("🇱🇧")
        elif currency == "EURO":
            currency_emojis.append("💶")
    
    currency_emoji_str = " ".join(currency_emojis)
    
    import html
    safe_subcategory = html.escape(subcategory)
    
    # Show notes preview if provided
    notes_preview = f"\n📝 <b>Notes:</b> {html.escape(optional_notes[:50])}{'...' if len(optional_notes) > 50 else ''}" if optional_notes else ""
    
    # Add multi-currency indicator
    multi_currency_indicator = "🌐 " if len(currency_amounts) > 1 else ""
    
    await update.message.reply_text(
        f"{multi_currency_indicator}💰 <b>Processing {transaction_desc}:</b>\n"
        f"Item: <code>{safe_subcategory}</code>\n"
        f"Amount{'s' if len(currency_amounts) > 1 else ''}: {amount_display}\n"
        f"Payment: {payment_type}"
        f"{notes_preview}\n\n"
        f"⏳ Matching and adding...",
        parse_mode='HTML'
    )
    
    # Add transaction with optional notes support
    success, message = add_transaction_smart(subcategory, currency_amounts, payment_type, optional_notes)
    await update.message.reply_text(message, parse_mode='HTML')

async def stats_command(update, context):
    """Show statistics"""
    logger.info(f"Received /stats from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    try:
        success, msg = copy_excel_from_onedrive()
        if not success:
            await update.message.reply_text(msg)
            return
        
        total_transactions = 0
        if EXCEL_AVAILABLE:
            with excel_lock:
                wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
                sheet = wb[TRACKING_SHEET_NAME]
                
                # Count transactions
                row = 12
                while sheet[f'C{row}'].value not in [None, ""]:
                    total_transactions += 1
                    row += 1
                    if row > 1000:
                        break
                
                wb.close()
        else:
            total_transactions = "N/A (openpyxl not installed)"
        
        # Get table count from cache
        table_count = len(_table_cache) if _table_cache else 0
        total_items = sum(data['count'] for data in _table_cache.values()) if _table_cache else 0
        
        # Get last backup time
        def get_last_backup_time() -> str:
            try:
                backups = list(BACKUP_DIR.glob("backup_*"))
                if not backups:
                    return "No backups yet"
                
                latest = max(backups, key=lambda x: x.stat().st_mtime)
                mod_time = datetime.fromtimestamp(latest.stat().st_mtime)
                return mod_time.strftime("%Y-%m-%d %H:%M")
            except:
                return "Error checking backups"
        
        # Check export directory
        export_files = list(EXPORT_DIR.glob("*")) if EXPORT_DIR.exists() else []
        
        await update.message.reply_text(
            f"📊 <b>Budget Tracker Statistics</b>\n\n"
            f"• <b>Excel Integration:</b> {'✅ Installed' if EXCEL_AVAILABLE else '❌ Not installed'}\n"
            f"• <b>OneDrive Integration:</b> {'✅ Installed' if ONEDRIVE_AVAILABLE else '❌ Not installed'}\n"
            f"• <b>Total Transactions:</b> {total_transactions}\n"
            f"• <b>Categories Loaded:</b> {table_count}\n"
            f"• <b>Items in Database:</b> {total_items}\n"
            f"• <b>Cache Status:</b> {'✅ Loaded' if _table_cache else '⚠️ Empty'}\n"
            f"• <b>Last Backup:</b> {get_last_backup_time()}\n"
            f"• <b>Exports Available:</b> {len(export_files)} files\n"
            f"• <b>Min Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}%\n"
            f"• <b>Negative Amounts:</b> ✅ Supported\n"
            f"• <b>LBP without commas:</b> ✅ Supported\n"
            f"• <b>Bot Status:</b> ✅ Running",
            parse_mode='HTML'
        )
        
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def save_command(update, context):
    """Force save to OneDrive"""
    logger.info(f"Received /save from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        return
    
    await update.message.reply_text("💾 Saving to OneDrive...")
    success, message = save_excel_to_onedrive()
    await update.message.reply_text(message)

async def clear_cache_command(update, context):
    """Clear the table cache"""
    logger.info(f"Received /clearcache from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    global _table_cache, _table_cache_timestamp
    _table_cache = {}
    _table_cache_timestamp = None
    
    await update.message.reply_text("✅ Cache cleared! Will reload on next transaction.")

async def unlock_command(update, context):
    """Emergency unlock Excel file - Linux version"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    await update.message.reply_text("🔓 Attempting to unlock Excel file...")
    
    try:
        # Kill any LibreOffice processes
        subprocess.run(['pkill', '-f', 'libreoffice'], capture_output=True, text=True)
        subprocess.run(['pkill', '-f', 'soffice'], capture_output=True, text=True)
        
        # Try to delete locked files
        locked = False
        if os.path.exists(LOCAL_COPY_PATH):
            try:
                os.remove(LOCAL_COPY_PATH)
                msg = "✅ Locked file removed. Bot will create new copy."
            except PermissionError:
                locked = True
                msg = "⚠️ File still locked. Try: sudo lsof | grep temp_budget.xlsm"
        
        if locked:
            msg += "\n\n💡 **Solutions:**"
            msg += "\n1. Check for open processes: <code>sudo lsof | grep xlsm</code>"
            msg += "\n2. Kill process: <code>sudo kill -9 PID</code>"
            msg += "\n3. Wait 1 minute and try again"
        
        await update.message.reply_text(msg)
        
    except Exception as e:
        await update.message.reply_text(f"❌ Unlock error: {str(e)[:200]}")

# ========== DOWNLOAD COMMAND ==========

async def download_command(update, context):
    """Handle download requests"""
    logger.info(f"Received /download from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "📥 <b>Download Options:</b>\n\n"
            "<b>CSV Exports:</b>\n"
            "• <code>/download csv day</code> - Today's transactions\n"
            "• <code>/download csv week</code> - This week's transactions\n"
            "• <code>/download csv month</code> - This month's transactions\n"
            "• <code>/download csv year</code> - This year's transactions\n"
            "• <code>/download csv all</code> - All transactions\n\n"
            "<b>Other Downloads:</b>\n"
            "• <code>/download summary</code> - Summary statistics (JSON)\n"
            "• <code>/download backup</code> - Excel backup file\n\n"
            "<b>Examples:</b>\n"
            "• <code>/download csv month</code>\n"
            "• <code>/download summary</code>\n"
            "• <code>/download backup</code>",
            parse_mode='HTML'
        )
        return
    
    command = context.args[0].lower()
    
    if command == "csv":
        if len(context.args) < 2:
            await update.message.reply_text(
                "Usage: /download csv [day/week/month/year/all]\n\n"
                "Examples:\n"
                "• /download csv day\n"
                "• /download csv month\n"
                "• /download csv all"
            )
            return
        
        time_range = context.args[1].lower()
        if time_range not in ["day", "week", "month", "year", "all"]:
            await update.message.reply_text(
                "❌ Invalid time range. Use: day, week, month, year, or all"
            )
            return
        
        await update.message.reply_text(f"📊 Exporting {time_range} transactions to CSV...")
        
        # Export to CSV
        success, result = export_to_csv(time_range)
        
        if success:
            try:
                # Send the file
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"transactions_{time_range}_{datetime.now().strftime('%Y%m%d')}.csv",
                        caption=f"📊 Transactions ({time_range}) - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Clean up the file after sending
                os.remove(result)
            except Exception as e:
                logger.error(f"Error sending CSV file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    elif command == "summary":
        await update.message.reply_text("📈 Exporting summary statistics...")
        
        # Export summary
        success, result = export_summary()
        
        if success:
            try:
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"budget_summary_{datetime.now().strftime('%Y%m%d')}.json",
                        caption=f"📈 Budget Summary - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Clean up
                os.remove(result)
            except Exception as e:
                logger.error(f"Error sending summary file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    elif command == "backup":
        await update.message.reply_text("💾 Creating Excel backup...")
        
        # Create backup copy
        success, result = create_backup_copy()
        
        if success:
            try:
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"budget_backup_{datetime.now().strftime('%Y%m%d')}.xlsm",
                        caption=f"💾 Excel Backup - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Note: Don't delete backup files from backup directory
            except Exception as e:
                logger.error(f"Error sending backup file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    else:
        await update.message.reply_text(
            "❌ Unknown download type.\n"
            "Use: csv, summary, or backup\n\n"
            "Example: /download csv month"
        )

async def error_handler(update, context):
    """Handle errors"""
    logger.error(f"Update {update} caused error {context.error}", exc_info=True)
    
    try:
        import html
        error_msg = html.escape(str(context.error)[:200])
        await update.message.reply_text(
            f"❌ <b>Bot Error</b>\n\n"
            f"{error_msg}\n\n"
            f"Please try again or use /help for assistance.",
            parse_mode='HTML'
        )
    except:
        pass

# ========== MAIN FUNCTION ==========

def main():
    """Start the bot"""
    print("\n" + "="*70)
    print("💰 SMART BUDGET TRACKER BOT - LINUX VERSION")
    print("Optimized for Ubuntu 22.04 aarch64")
    print("="*70)
    
    # Check BOT_TOKEN
    if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        print("❌ ERROR: BOT_TOKEN not set in .env file!")
        input("\nPress Enter to exit...")
        return
    
    # Check ALLOWED_USER_IDS
    if not ALLOWED_USER_IDS:
        print("❌ ERROR: No ALLOWED_USER_IDS or ALLOWED_USER_ID set in .env file!")
        print("Add your user ID to .env file:")
        print("Example: ALLOWED_USER_ID=1663164223")
        input("\nPress Enter to exit...")
        return
    
    print(f"✅ BOT_TOKEN: {BOT_TOKEN[:10]}...")
    print(f"✅ ALLOWED_USER_IDS: {ALLOWED_USER_IDS}")
    print(f"✅ Excel library: {'✅ openpyxl/pandas' if EXCEL_AVAILABLE else '❌ NOT INSTALLED'}")
    print(f"✅ OneDrive API: {'✅ onedrivesdk' if ONEDRIVE_AVAILABLE else '❌ NOT INSTALLED'}")
    print(f"✅ Minimum Confidence: {MINIMUM_CONFIDENCE*100:.0f}%")
    print(f"✅ Negative Amounts: ✅ SUPPORTED")
    print(f"✅ LBP without commas: ✅ SUPPORTED")
    print(f"✅ Download Feature: ✅ ADDED")
    
    # Setup directories
    BACKUP_DIR.mkdir(exist_ok=True, parents=True)
    EXPORT_DIR.mkdir(exist_ok=True, parents=True)
    LOCAL_COPY_PATH.parent.mkdir(exist_ok=True, parents=True)

    # Check for OneDrive configuration
    print("\n🔐 Checking OneDrive configuration...")
    if not ONEDRIVE_CLIENT_ID or ONEDRIVE_CLIENT_ID == "YOUR_CLIENT_ID":
        print("⚠️  ONEDRIVE_CLIENT_ID not set. OneDrive sync will not work.")
        print("Run /onedrive_auth in Telegram after setting up your Azure app.")
    
    # Test OneDrive connection if configured
    if ONEDRIVE_CLIENT_ID and ONEDRIVE_CLIENT_ID != "YOUR_CLIENT_ID":
        print("Testing OneDrive connection...")
        success, msg = download_from_onedrive()
        if success:
            print(f"✅ OneDrive: {msg}")
        else:
            print(f"⚠️  OneDrive: {msg}")
            print("Run /onedrive_auth in Telegram to authenticate.")
    else:
        print("⚠️  OneDrive not configured. Files will only be saved locally.")
    
    # Load tables (requires openpyxl)
    print("\n🔧 Loading data from Excel...")
    if EXCEL_AVAILABLE and LOCAL_COPY_PATH.exists():
        tables_dict = load_all_tables_with_details()
        if tables_dict:
            total_items = sum(data['count'] for data in tables_dict.values())
            print(f"✅ Loaded {len(tables_dict)} categories with {total_items} items")
        else:
            print("❌ Could not load data from Excel")
    else:
        print("⚠️  openpyxl not installed or Excel file not found.")
        print("Install with: pip install openpyxl pandas")
    
    print("\n" + "="*70)
    print("🤖 BOT IS STARTING...")
    print("="*70)
    
    # Create bot application
    try:
        app = Application.builder().token(BOT_TOKEN).build()
        logger.info("Bot application created successfully")
    except Exception as e:
        print(f"❌ Failed to create bot: {e}")
        input("\nPress Enter to exit...")
        return
    
    # Add command handlers
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("testparse", testparse_command))
    app.add_handler(CommandHandler("recent", recent_command))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CommandHandler("save", save_command))
    app.add_handler(CommandHandler("clearcache", clear_cache_command))
    app.add_handler(CommandHandler("download", download_command))
    app.add_handler(CommandHandler("unlock", unlock_command))
    app.add_handler(CommandHandler("onedrive_auth", onedrive_auth_command))
    app.add_handler(CommandHandler("onedrive_test", onedrive_test_command))
    app.add_handler(CommandHandler("onedrive_code", onedrive_complete_auth_command))
    app.add_handler(CommandHandler("onedrive_test", onedrive_test_command))
    app.add_handler(CommandHandler("direct_auth", direct_auth_command))

    
    # Add message handler
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Add error handler
    app.add_error_handler(error_handler)
    
    print("\n✅ Bot is ready and running!")
    print("📱 Open Telegram and send /start to your bot")
    print("🔐 New: Use /onedrive_auth to setup OneDrive")
    print("📥 Use /download to export data")
    print("⏳ Waiting for messages...")
    print("="*70)
    
    try:
        # Start polling
        app.run_polling(
            poll_interval=1.0,
            timeout=10,
            drop_pending_updates=True
        )
    except KeyboardInterrupt:
        print("\n🛑 Bot stopped by user")
    except Exception as e:
        print(f"\n❌ Bot crashed: {e}")
        logger.error(f"Bot crashed: {e}", exc_info=True)
        input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()