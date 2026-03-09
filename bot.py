"""
BUDGET TRACKER BOT - LINUX VERSION WITH ONEDRIVE API
Optimized for Ubuntu 22.04 aarch64 on Oracle Cloud
"""
import random
import functools
import os
import re
import logging
import shutil
import sys
import json
import asyncio
import time
import subprocess
from zipfile import ZipFile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Tuple, Optional, List, Any
from difflib import get_close_matches
import threading
from dataclasses import dataclass
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor, Future

# ========== EXCEL LIBRARY FOR LINUX ==========
try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import range_boundaries, get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    # Check if openpyxl supports keep_vba
    import openpyxl
    OPENPYXL_VERSION = tuple(map(int, openpyxl.__version__.split('.')))
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not found. Install: pip install openpyxl pandas")

# ========== TELEGRAM BOT IMPORTS ==========
try:
    from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackQueryHandler
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    print("⚠️  python-telegram-bot not found. Install: pip install python-telegram-bot")

# ========== SETUP LOGGING (MUST BE BEFORE ANY LOGGER USE) ==========
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.FileHandler('budget_bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ========== CONFIGURATION ==========
from dotenv import load_dotenv
load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")

# Handle both ALLOWED_USER_ID (singular) and ALLOWED_USER_IDS (plural)
ALLOWED_USER_IDS = []
user_ids_str = os.getenv("ALLOWED_USER_IDS", os.getenv("ALLOWED_USER_ID", ""))
if user_ids_str:
    for id_str in user_ids_str.split(","):
        id_str = id_str.strip()
        if id_str:
            try:
                ALLOWED_USER_IDS.append(int(id_str))
            except ValueError:
                print(f"⚠️ Warning: Invalid user ID '{id_str}' in .env file")

# File paths (Linux paths)
LOCAL_COPY_PATH = Path(os.getenv("LOCAL_COPY_PATH", "/home/ubuntu/budget_tracker/temp_budget.xlsm"))
TRACKING_SHEET_NAME = os.getenv("TRACKING_SHEET_NAME", "Budget Tracking")
DROPDOWN_SHEET_NAME = os.getenv("DROPDOWN_SHEET_NAME", "Dropdown Data")
EXPORT_DIR = Path(os.getenv("EXPORT_DIR", "/home/ubuntu/budget_tracker/exports"))

# Backup configuration
BACKUP_DIR = Path(os.getenv("BACKUP_DIR", "/home/ubuntu/budget_tracker/backups"))
BACKUP_RETENTION_DAYS = 7

# Matching configuration
MINIMUM_CONFIDENCE = 0.5  # 50% minimum confidence to accept a match
FUZZY_MATCH_THRESHOLD = 0.6  # Minimum for fuzzy matching

# Common words to reject (these won't match any category)
COMMON_NON_CATEGORIES = [
    'ok', 'okay', 'yes', 'no', 'maybe', 'test', 'hello', 'hi',
    'thanks', 'thank', 'please', 'help', 'what', 'how', 'why',
    'when', 'where', 'who', 'which', 'good', 'bad', 'nice',
    'great', 'fine', 'well', 'sorry', 'excuse', 'hey', 'yo'
]

# Currency configuration
@dataclass
class CurrencyConfig:
    symbols: List[str]
    column: str
    decimal_places: int = 2
    thousands_separator: bool = True

CURRENCIES = {
    'USD': CurrencyConfig(symbols=['$', 'usd', 'dollar'], column='H', decimal_places=2),
    'LBP': CurrencyConfig(symbols=['lbp', 'ليرة', 'lira', 'ل.ل'], column='I', decimal_places=0, thousands_separator=True),
    'EURO': CurrencyConfig(symbols=['€', 'euro', 'eur'], column='J', decimal_places=2)
}
DEFAULT_CURRENCY = 'USD'

# Payment types with additional options
PAYMENT_TYPES = {
    "Cash": "💵",
    "Card": "💳",
    "Bank Transfer": "🏦",
    "Digital Wallet": "📱"
}

CARD_KEYWORDS = ["card", "credit", "debit", "visa", "mastercard", "amex", "paypal", "bank"]
DIGITAL_KEYWORDS = ["paypal", "digital", "wallet", "apple pay", "google pay"]

# Thread safety for Excel operations
excel_lock = threading.Lock()

# Cache for loaded tables
_table_cache = {}
_table_cache_timestamp = None

# OneDrive client instance
_onedrive_client = None

# ========== ONEDRIVE API FUNCTIONS ==========
# ========== MICROSOFT GRAPH API (Modern OneDrive) ==========
try:
    import msal
    import requests
    from msal import ConfidentialClientApplication
    ONEDRIVE_AVAILABLE = True
except ImportError:
    ONEDRIVE_AVAILABLE = False
    print("⚠️  Microsoft Graph libraries not found. Install: pip install msal requests msgraph-core")

# ========== ONEDRIVE CONFIGURATION (single consolidated block) ==========
ONEDRIVE_CLIENT_ID = os.getenv("ONEDRIVE_CLIENT_ID", "")
ONEDRIVE_CLIENT_SECRET = os.getenv("ONEDRIVE_CLIENT_SECRET", "")
ONEDRIVE_TENANT_ID = os.getenv("ONEDRIVE_TENANT_ID", "common")
ONEDRIVE_REDIRECT_URI = os.getenv("ONEDRIVE_REDIRECT_URI", "http://localhost:8080/")
ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_FILE_PATH", "/budget_tracker.xlsm")
ONEDRIVE_TOKEN_PATH = Path(os.getenv("ONEDRIVE_TOKEN_PATH", "/home/ubuntu/budget_tracker/onedrive_tokens.json"))

# Scopes (from .env). Example: ONEDRIVE_SCOPES=Files.ReadWrite,offline_access
_scopes_raw = os.getenv("ONEDRIVE_SCOPES", "Files.ReadWrite,offline_access")
ONEDRIVE_SCOPES = [s.strip() for s in _scopes_raw.split(",") if s.strip()]

# MSAL requires scopes as a plain list[str] (not set/frozenset)
if not isinstance(ONEDRIVE_SCOPES, list):
    ONEDRIVE_SCOPES = list(ONEDRIVE_SCOPES)

logger.info(f"OneDrive scopes loaded: {ONEDRIVE_SCOPES} (type={type(ONEDRIVE_SCOPES)})")

TRACKING_TABLE_NAME = os.getenv("TRACKING_TABLE_NAME", "Tracking")

# OneDrive client instance
_onedrive_app = None

# Background uploader single-thread executor to serialize retries
_background_uploader = ThreadPoolExecutor(max_workers=1)

# Track pending/last background sync future so we don't queue duplicates
_bg_sync_future: Optional[Future] = None

# Serialize all OneDrive uploads across the process
_onedrive_upload_lock = threading.Lock()
_upload_in_progress = threading.Event()

def get_onedrive_app():
    """Get or create MSAL app for OneDrive"""
    global _onedrive_app

    if _onedrive_app is not None:
        return _onedrive_app

    if not ONEDRIVE_AVAILABLE:
        logger.error("MSAL not available. Cannot connect to OneDrive.")
        return None

    try:
        _onedrive_app = ConfidentialClientApplication(
            client_id=ONEDRIVE_CLIENT_ID,
            client_credential=ONEDRIVE_CLIENT_SECRET,
            authority=f"https://login.microsoftonline.com/{ONEDRIVE_TENANT_ID}"
        )
        return _onedrive_app
    except Exception as e:
        logger.error(f"Error creating MSAL app: {e}")
        return None

def get_onedrive_token():
    """Get access token for OneDrive - fixed version"""
    try:
        app = get_onedrive_app()
        if app is None:
            return None

        # Check if we have auth code in environment
        auth_code = os.getenv("ONEDRIVE_AUTH_CODE")

        if auth_code:
            logger.info("Using ONEDRIVE_AUTH_CODE from environment (truncated)...")
            try:
                result = app.acquire_token_by_authorization_code(
                    code=auth_code,
                    scopes=list(ONEDRIVE_SCOPES),
                    redirect_uri=ONEDRIVE_REDIRECT_URI
                )
            except Exception as ex:
                logger.error(f"Error exchanging auth code: {ex}")
                result = {}

            if result and "access_token" in result:
                # Save tokens
                with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                    json.dump(result, f, indent=2)
                logger.info("Tokens saved from auth code")
                os.environ.pop("ONEDRIVE_AUTH_CODE", None)
                return result["access_token"]

        # Try to load existing tokens
        token_data = None
        if ONEDRIVE_TOKEN_PATH.exists():
            try:
                with open(ONEDRIVE_TOKEN_PATH, 'r') as f:
                    token_data = json.load(f)
            except Exception:
                token_data = None

        if token_data and 'refresh_token' in token_data:
            try:
                # Use acquire_token_by_refresh_token if available; fallback to client credential flow for app-only
                if hasattr(app, 'acquire_token_by_refresh_token'):
                    result = app.acquire_token_by_refresh_token(
                        refresh_token=token_data['refresh_token'],
                        scopes=list(ONEDRIVE_SCOPES)
                    )
                else:
                    # Try client credential flow (app-only) as fallback
                    result = app.acquire_token_for_client(scopes=list(ONEDRIVE_SCOPES))
            except Exception as ex:
                logger.warning(f"Refresh token flow failed: {ex}")
                result = {}

            if result and "access_token" in result:
                with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                    json.dump(result, f, indent=2)
                return result["access_token"]

        logger.info("No valid OneDrive authentication found")
        return None

    except Exception as e:
        logger.error(f"Error getting token: {e}")
        return None

def _graph_request(
    method: str,
    url: str,
    token: str,
    *,
    json_body: Optional[dict] = None,
    timeout: int = 30
) -> Tuple[bool, str, Optional[dict]]:
    """Small helper for Microsoft Graph requests with consistent error handling."""
    try:
        headers = {"Authorization": f"Bearer {token}"}
        if json_body is not None:
            headers["Content-Type"] = "application/json"

        resp = requests.request(method, url, headers=headers, json=json_body, timeout=timeout)

        if resp.status_code in (200, 201, 202, 204):
            if resp.text and resp.text.strip():
                try:
                    return True, "OK", resp.json()
                except Exception:
                    return True, "OK", None
            return True, "OK", None

        # Error path
        try:
            err = resp.json()
            msg = err.get("error", {}).get("message", resp.text)
        except Exception:
            msg = resp.text

        return False, f"{resp.status_code}: {msg}", None

    except Exception as e:
        return False, str(e), None

def _graph_request_with_retry(
    method: str,
    url: str,
    token: str,
    *,
    json_body: Optional[dict] = None,
    timeout: int = 60,
    max_retries: int = 6,
) -> Tuple[bool, str, Optional[dict]]:
    """
    Retry wrapper for transient Graph workbook errors (503/504/502/429).
    """
    for attempt in range(1, max_retries + 1):
        ok, msg, data = _graph_request(method, url, token, json_body=json_body, timeout=timeout)
        if ok:
            return ok, msg, data

        # msg looks like: "503: ...." or "429: ...."
        status = None
        try:
            status = int(str(msg).split(":", 1)[0].strip())
        except Exception:
            status = None

        if status in (429, 502, 503, 504):
            # exponential backoff with jitter
            sleep_s = min(20.0, (2 ** (attempt - 1)) + random.random())
            logger.warning(f"Graph transient error {status}. Retry {attempt}/{max_retries} in {sleep_s:.1f}s. msg={msg}")
            time.sleep(sleep_s)
            continue

        # Non-retryable error
        return False, msg, None

    return False, f"Max retries exceeded. Last error: {msg}", None

def graph_get_table_row_count() -> Tuple[bool, str, Optional[int]]:
    """Return number of rows in the Excel Table via Graph."""
    token = get_onedrive_token()
    if not token:
        return False, "❌ Not authenticated with OneDrive.", None

    file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")
    url = (
        f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/workbook/"
        f"tables/{TRACKING_TABLE_NAME}/rows?$top=2000"
    )

    ok, msg, data = _graph_request_with_retry("GET", url, token, timeout=60)
    if not ok:
        return False, msg, None

    try:
        return True, "OK", len((data or {}).get("value", []))
    except Exception as e:
        return False, f"Bad response: {e}", None

def _worksheet_row_to_table_index(row: int) -> int:
    """
    Convert worksheet row number to 0-based Excel Table row index.
    Assuming:
    - Header row is at worksheet row 11 (table header)
    - First data row is worksheet row 12 (table index 0)
    """
    # Subtract 12 because:
    # Row 12 (first transaction) -> Index 0
    # Row 13 -> Index 1
    # etc.
    return row - 12

def graph_delete_transaction_at_row(row: int) -> Tuple[bool, str]:
    """
    Physically removes the row in the table via Microsoft Graph DELETE endpoint.
    This reduces the table's length and shifts up all rows below.
    """
    token = get_onedrive_token()
    if not token:
        return False, "❌ Not authenticated with OneDrive. Use /onedrive_auth first."

    if row < 12:
        return False, f"❌ Invalid row: {row}. First transaction row is 12."

    index = _worksheet_row_to_table_index(row)
    file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")

    url = (
        f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/workbook/"
        f"tables/{TRACKING_TABLE_NAME}/rows/itemAt(index={index})"
    )

    # Make the DELETE request
    ok, msg, _ = _graph_request("DELETE", url, token, timeout=30)
    if ok:
        return True, f"✅ Transaction deleted by removing table row {row} (table index {index}). Table shrank and rows shifted up."
    return False, f"❌ Graph DELETE failed: {msg}"

def graph_get_table_row_values(index: int) -> Tuple[bool, str, Optional[List[Any]]]:
     """Read one table row (all table columns) as a list via Graph."""
     token = get_onedrive_token()
     if not token:
         return False, "❌ Not authenticated with OneDrive.", None
 
     file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")
     url = (
         f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/workbook/"
         f"tables/{TRACKING_TABLE_NAME}/rows/itemAt(index={index})/range"
     )
 
     ok, msg, data = _graph_request_with_retry("GET", url, token, timeout=60)
     if not ok:
         return False, msg, None
 
     try:
         values = (data or {}).get("values", [])
         if not values or not values[0]:
             return True, "OK", []
         return True, "OK", values[0]
     except Exception as e:
         return False, f"Bad response: {e}", None
 
 
def graph_set_table_row_values(index: int, row_values: List[Any]) -> Tuple[bool, str]:
     """Write one table row (all table columns) via Graph."""
     token = get_onedrive_token()
     if not token:
         return False, "❌ Not authenticated with OneDrive."
 
     # ensure 13 columns
     row_values = (row_values + [None] * 13)[:13]
 
     file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")
     url = (
         f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/workbook/"
         f"tables/{TRACKING_TABLE_NAME}/rows/itemAt(index={index})/range"
     )
 
     ok, msg, _ = _graph_request_with_retry("PATCH", url, token, json_body={"values": [row_values]}, timeout=60)
     if ok:
         return True, "OK"
     return False, msg

def graph_update_transaction_at_row(
    row: int,
    *,
    date_value,
    payment: str,
    tx_type: str,
    category: str,
    subcategory: str,
    usd: Optional[float],
    lbp: Optional[float],
    euro: Optional[float],
    details: Optional[str],
) -> Tuple[bool, str]:
    """Update a transaction row in the Excel Table using Graph (preserves shapes/formulas)."""
    token = get_onedrive_token()
    if not token:
        return False, "❌ Not authenticated with OneDrive. Use /onedrive_auth first."

    if row < 12:
        return False, f"❌ Invalid row: {row}. First transaction row is 12."

    index = _worksheet_row_to_table_index(row)
    file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")

    # Update by PATCHing the range of that row
    url = (
        f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/workbook/"
        f"tables/{TRACKING_TABLE_NAME}/rows/itemAt(index={index})/range"
    )

    # Force date-only format: YYYY-MM-DD
    from datetime import datetime, date

    def _to_iso_date_only(value) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        s = str(value)
        # If already like 2026-01-22T00:00:00 -> keep only date part
        return s.split("T")[0]

    date_str = _to_iso_date_only(date_value)

    row_values = [
        date_str,         # C Date
        payment,          # D Payment Type
        tx_type,          # E Type
        category,         # F Category
        subcategory,      # G Sub-Category
        usd,              # H Amount $
        lbp,              # I Amount LBP
        euro,             # J Amount €
        details or None,  # K Details
        None,             # L Balance $ (formula)
        None,             # M Balance LBP (formula)
        None,             # N Balance € (formula)
        None,             # O Effective Date (formula)
    ]

    ok, msg, _ = _graph_request("PATCH", url, token, json_body={"values": [row_values]}, timeout=60)
    if ok:
        return True, f"✅ Transaction updated (Graph) at row {row} (table index {index})."
    return False, f"❌ Graph update failed: {msg}"

def append_transaction_to_tracking_table(
    *,
    date_value,
    payment: str,
    tx_type: str,
    category: str,
    subcategory: str,
    usd: Optional[float],
    lbp: Optional[float],
    euro: Optional[float],
    details: Optional[str],
) -> Tuple[bool, str]:
    """
    Append a transaction row to the Excel Table 'Tracking' using Microsoft Graph workbook API.
    This preserves shapes/text boxes and triggers Excel Online recalculation.
    """
    token = get_onedrive_token()
    if not token:
        return False, "❌ Not authenticated with OneDrive. Use /onedrive_auth first."

    # URL-encode spaces for Graph path usage
    file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")

    # Append row endpoint (table name in TRACKING_TABLE_NAME)
    url = f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/workbook/tables/{TRACKING_TABLE_NAME}/rows/add"

    # Dates: safest to send ISO string
    try:
        date_str = date_value.isoformat()
    except Exception:
        date_str = str(date_value)

    # Your table columns order (13 columns total):
    # Date | Payment | Type | Category | Subcategory | Amount $ | Amount LBP | Amount € | Details |
    # Balance $ | Balance LBP | Balance € | Effective Date
    #
    # Formula-driven columns -> send null so Excel calculates them.
    row = [
        date_str,
        payment,
        tx_type,
        category,
        subcategory,
        usd,
        lbp,
        euro,
        details or None,
        None,  # Balance $ (formula)
        None,  # Balance LBP (formula)
        None,  # Balance € (formula)
        None,  # Effective Date (formula)
    ]

    body = {"values": [row]}

    ok, msg, _ = _graph_request("POST", url, token, json_body=body, timeout=60)
    if ok:
        return True, "✅ Transaction appended to Excel table (Graph)."

    return False, f"❌ Graph append failed: {msg}"

def download_from_onedrive() -> Tuple[bool, str]:
    """Download Excel file from OneDrive using Microsoft Graph API"""
    try:
        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated with OneDrive. Use /onedrive_auth first."

        # Format file path for Graph API
        file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")

        # Graph API endpoint
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/content"

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }

        logger.info(f"Downloading from OneDrive: {ONEDRIVE_FILE_PATH}")

        response = requests.get(url, headers=headers, timeout=30)

        if response.status_code == 200:
            # Ensure parent exists
            LOCAL_COPY_PATH.parent.mkdir(parents=True, exist_ok=True)
            # Save to local file
            with open(LOCAL_COPY_PATH, 'wb') as f:
                f.write(response.content)

            logger.info(f"Downloaded {LOCAL_COPY_PATH} ({len(response.content)} bytes)")
            return True, "✅ File downloaded from OneDrive"
        elif response.status_code == 404:
            return False, f"❌ File not found in OneDrive: {ONEDRIVE_FILE_PATH}"
        else:
            try:
                error_msg = response.json().get('error', {}).get('message', response.text)
            except Exception:
                error_msg = response.text
            return False, f"❌ Download error: {error_msg}"

    except Exception as e:
        logger.error(f"Error downloading from OneDrive: {str(e)}")
        return False, f"❌ Download error: {str(e)[:200]}"

def copy_excel_from_onedrive() -> Tuple[bool, str]:
    """Download fresh copy from OneDrive before operations"""
    try:
        success, msg = download_from_onedrive()
        if success:
            logger.info("Successfully downloaded fresh copy from OneDrive")
        else:
            logger.error(f"Failed to download from OneDrive: {msg}")
        return success, msg
    except Exception as e:
        logger.error(f"Error in copy_excel_from_onedrive: {e}")
        return False, f"❌ Error downloading: {str(e)[:200]}"

def upload_to_onedrive() -> Tuple[bool, str]:
    """
    Upload local Excel file to OneDrive using a resumable upload session, with robust retries.
    Now serialized with a global lock to prevent parallel uploads to the same path.
    """
    try:
        # Make uploads single-flight
        with _onedrive_upload_lock:
            _upload_in_progress.set()
            try:
                token = get_onedrive_token()
                if not token:
                    return False, "❌ Not authenticated with OneDrive. Use /onedrive_auth first."
                if not LOCAL_COPY_PATH.exists():
                    return False, "❌ Local file not found"

                with open(LOCAL_COPY_PATH, "rb") as f:
                    data_bytes = f.read()
                file_size = len(data_bytes)

                session_url = f"https://graph.microsoft.com/v1.0/me/drive/root:{ONEDRIVE_FILE_PATH}:/createUploadSession"
                session_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
                session_body = {"item": {"@microsoft.graph.conflictBehavior": "replace",
                                         "name": Path(ONEDRIVE_FILE_PATH).name}}

                logger.info(f"Creating upload session for {ONEDRIVE_FILE_PATH}")
                resp = requests.post(session_url, headers=session_headers, json=session_body, timeout=30)
                if resp.status_code not in (200, 201):
                    try:
                        err_msg = resp.json().get("error", {}).get("message", resp.text)
                    except Exception:
                        err_msg = resp.text
                    return False, f"❌ Upload error: {err_msg}"

                upload_url = resp.json().get("uploadUrl")
                if not upload_url:
                    return False, "❌ Upload error: Missing uploadUrl from session"

                content_range = f"bytes 0-{file_size-1}/{file_size}"
                max_retries = 10
                base_sleep = 2.0
                last_err = None

                for attempt in range(1, max_retries + 1):
                    try:
                        put_headers = {
                            "Content-Length": str(file_size),
                            "Content-Range": content_range
                        }
                        logger.info(f"Upload attempt {attempt}/{max_retries} via session")
                        put_resp = requests.put(upload_url, headers=put_headers, data=data_bytes, timeout=60)

                        if put_resp.status_code in (200, 201):
                            logger.info(f"Uploaded to OneDrive: {ONEDRIVE_FILE_PATH}")
                            return True, "✅ File uploaded to OneDrive"

                        # Parse error
                        try:
                            err_json = put_resp.json()
                            err_msg = err_json.get("error", {}).get("message", put_resp.text)
                            err_code = err_json.get("error", {}).get("code", "")
                        except Exception:
                            err_msg = put_resp.text
                            err_code = ""

                        # Handle "same name is currently being uploaded"
                        msg_l = err_msg.lower()
                        same_name_busy = "same name is currently being uploaded" in msg_l

                        # Retry on lock, conflict, throttling, or concurrent upload
                        if put_resp.status_code in (423, 409, 429) or "locked" in msg_l \
                           or err_code in ("resourceLocked", "fileLocked") or same_name_busy:
                            retry_after = put_resp.headers.get("Retry-After")
                            if retry_after and retry_after.isdigit():
                                sleep_s = float(retry_after)
                            else:
                                # backoff with a bit of jitter
                                sleep_s = base_sleep * attempt + 0.3 * attempt
                                # If we hit the "same name uploading" case, wait a bit longer
                                if same_name_busy:
                                    sleep_s += 2.0
                            last_err = f"{put_resp.status_code} {err_code} {err_msg}"
                            logger.warning(f"Upload busy/locked. Retrying in {sleep_s:.1f}s. Error: {last_err}")
                            time.sleep(sleep_s)
                            continue

                        last_err = f"{put_resp.status_code} {err_code} {err_msg}"
                        logger.error(f"Upload session error (no retry): {last_err}")
                        return False, f"❌ Upload error: {err_msg}"

                    except Exception as inner:
                        last_err = str(inner)
                        logger.warning(f"Upload session attempt {attempt} failed: {inner}")
                        time.sleep(base_sleep * attempt)

                return False, f"❌ Upload error: {last_err or 'Unknown error'}"

            finally:
                _upload_in_progress.clear()

    except Exception as e:
        logger.error(f"Error uploading to OneDrive: {str(e)}")
        return False, f"❌ Upload error: {str(e)[:200]}"

async def onedrive_auth_command(update, context):
    """Start OneDrive authentication flow"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    try:
        app = get_onedrive_app()
        if app is None:
            await update.message.reply_text("❌ MSAL app not configured. Check ONEDRIVE_CLIENT_ID in .env")
            return

        logger.info(f"DEBUG scopes used for auth: {ONEDRIVE_SCOPES} type={type(ONEDRIVE_SCOPES)}")
        logger.info(f"DEBUG redirect used for auth: {ONEDRIVE_REDIRECT_URI}")

        auth_url = app.get_authorization_request_url(
            scopes=list(ONEDRIVE_SCOPES),
            redirect_uri=ONEDRIVE_REDIRECT_URI
        )

        logger.info(f"DEBUG auth_url generated: {auth_url}")

        message = (
            "🔐 <b>OneDrive Authentication Required</b>\n\n"
            "1. <b>Click this link:</b>\n"
            f"<code>{auth_url}</code>\n\n"
            "2. <b>Sign in</b> with your Microsoft account\n"
            "3. You'll be redirected to a page\n"
            "4. <b>Copy the ENTIRE URL</b> from address bar\n\n"
            "5. <b>Look for the 'code=' parameter</b>\n"
            "Example: <code>https://your-domain.lhr.life/?code=M.R3_BL2.1234...</code>\n\n"
            "6. <b>Set the code in your VM:</b>\n"
            "<code>export ONEDRIVE_AUTH_CODE='your_code_here'</code>\n\n"
            "7. <b>Then run:</b> <code>/onedrive_test</code>"
        )

        await update.message.reply_text(message, parse_mode="HTML")

    except Exception as e:
        logger.error(f"Error generating auth URL: {e}")
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def onedrive_complete_auth_command(update, context):
    """Complete OneDrive authentication with code"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    if not context.args:
        await update.message.reply_text(
            "Usage: /onedrive_code [auth_code]\n\n"
            "Example: /onedrive_code M.R3_BL2.0.AQABAAIAAADX8GCiHq7HjXk5KJX7w..."
        )
        return

    auth_code = ' '.join(context.args)

    await update.message.reply_text("🔐 Completing authentication...")

    try:
        app = get_onedrive_app()
        if app is None:
            await update.message.reply_text("❌ MSAL app not configured")
            return

        # Exchange code for tokens
        result = app.acquire_token_by_authorization_code(
            code=auth_code,
            scopes=list(ONEDRIVE_SCOPES),
            redirect_uri=ONEDRIVE_REDIRECT_URI
        )

        if "access_token" in result:
            # Save tokens
            with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                json.dump(result, f)

            await update.message.reply_text(
                "✅ <b>Authentication Successful!</b>\n\n"
                "You can now use OneDrive features:\n"
                "• /onedrive_test - Test connection\n"
                "• /save - Force save to OneDrive\n"
                "• Normal transactions will auto-sync",
                parse_mode='HTML'
            )
        else:
            error_msg = result.get('error_description', 'Unknown error')
            await update.message.reply_text(f"❌ Authentication failed: {error_msg}")

    except Exception as e:
        logger.error(f"Error completing auth: {e}")
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def onedrive_complete_auth_from_url_command(update, context):
    """Complete OneDrive auth by pasting the FULL redirect URL (safer than pasting code)."""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    if not context.args:
        await update.message.reply_text(
            "Usage: /onedrive_code_url <full_redirect_url>\n\n"
            "Example:\n"
            "/onedrive_code_url http://localhost:8080/?code=...&session_state=..."
        )
        return

    full_url = " ".join(context.args).strip()

    try:
        from urllib.parse import urlparse, parse_qs

        parsed = urlparse(full_url)
        qs = parse_qs(parsed.query)
        code = qs.get("code", [None])[0]

        if not code:
            await update.message.reply_text("❌ Could not find `code=` in the URL you pasted.")
            return

        await update.message.reply_text("🔐 Completing authentication...")

        app = get_onedrive_app()
        if app is None:
            await update.message.reply_text("❌ MSAL app not configured")
            return

        result = app.acquire_token_by_authorization_code(
            code=code,
            scopes=list(ONEDRIVE_SCOPES),
            redirect_uri=ONEDRIVE_REDIRECT_URI
        )

        if "access_token" in result:
            with open(ONEDRIVE_TOKEN_PATH, "w") as f:
                json.dump(result, f)

            await update.message.reply_text(
                "✅ <b>Authentication Successful!</b>\n\n"
                "You can now use OneDrive features:\n"
                "• /onedrive_test - Test connection\n"
                "• Normal transactions will write via Graph",
                parse_mode="HTML"
            )
        else:
            error_msg = result.get("error_description", str(result))
            await update.message.reply_text(f"❌ Authentication failed: {error_msg}")

    except Exception as e:
        logger.error(f"Error completing auth from URL: {e}", exc_info=True)
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")


async def onedrive_test_command(update, context):
    """Test OneDrive connection (non-blocking)"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    await update.message.reply_text("🔄 Testing OneDrive connection...")

    try:
        # Run download in a thread
        success, msg = await run_blocking(copy_excel_from_onedrive)
        await update.message.reply_text(f"Download test: {msg}")

        if success:
            # Run upload in a thread
            success2, msg2 = await run_blocking(upload_to_onedrive)
            await update.message.reply_text(f"Upload test: {msg2}")

            if success2:
                await update.message.reply_text("✅ OneDrive connection successful!")
            else:
                await update.message.reply_text("⚠️ Download worked but upload failed")
        else:
            await update.message.reply_text("❌ OneDrive connection failed. Run /onedrive_auth first.")
    except Exception as e:
        logger.exception("Error during OneDrive test")
        await update.message.reply_text(f"❌ Error running OneDrive test: {str(e)[:200]}")

async def onedrive_code_command(update, context):
    """Manually set OneDrive auth code"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    if not context.args:
        # Check if we have saved codes
        codes_file = Path("/home/ubuntu/Tracking_Budget_Sheet_Python/auth_codes.json")
        if codes_file.exists():
            try:
                with open(codes_file, 'r') as f:
                    codes = json.load(f)
                if codes:
                    latest = codes[-1]
                    code = latest['code']
                    await update.message.reply_text(
                        f"📝 Found recent auth code from {latest['timestamp']}\n"
                        f"Using code: {code[:30]}...\n\n"
                        f"🔐 Completing authentication..."
                    )

                    # Complete authentication with this code
                    await complete_onedrive_auth(update, code)
                    return
            except:
                pass

        await update.message.reply_text(
            "Usage: /onedrive_code [auth_code]\n\n"
            "Or run without arguments to use the last captured code.\n\n"
            "Get your code from:\n"
            "1. Click the auth link from /onedrive_auth\n"
            "2. Sign in with Microsoft\n"
            "3. Copy the 'code=' parameter from the URL"
        )
        return

    # Use provided code
    code = ' '.join(context.args)
    await update.message.reply_text(f"🔐 Using provided code: {code[:30]}...")
    await complete_onedrive_auth(update, code)

async def complete_onedrive_auth(update, auth_code):
    """Complete OneDrive authentication"""
    try:
        app = get_onedrive_app()
        if app is None:
            await update.message.reply_text("❌ MSAL app not configured")
            return

        result = app.acquire_token_by_authorization_code(
            code=auth_code,
            scopes=list(ONEDRIVE_SCOPES),
            redirect_uri=ONEDRIVE_REDIRECT_URI
        )

        if "access_token" in result:
            with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                json.dump(result, f)
            await update.message.reply_text("✅ Authentication completed successfully!")
        else:
            await update.message.reply_text(f"❌ Authentication failed: {result.get('error_description','Unknown')}")

    except Exception as e:
        logger.error(f"Auth error: {e}")
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def direct_auth_command(update, context):
    """Direct authentication with manual code"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("? Unauthorized.")
        return

    # Get auth code from .env
    auth_code = os.getenv("ONEDRIVE_AUTH_CODE")

    if not auth_code:
        await update.message.reply_text("? No ONEDRIVE_AUTH_CODE in .env file")
        return

    await update.message.reply_text(f"?? Using auth code: {auth_code[:50]}...")

    try:
        # Direct HTTP request to get tokens
        import requests

        data = {
            'client_id': ONEDRIVE_CLIENT_ID,
            'scope': 'Files.ReadWrite',
            'code': auth_code,
            'redirect_uri': ONEDRIVE_REDIRECT_URI,
            'grant_type': 'authorization_code',
            'client_secret': ONEDRIVE_CLIENT_SECRET
        }

        response = requests.post(
            'https://login.microsoftonline.com/common/oauth2/v2.0/token',
            data=data,
            timeout=30
        )

        if response.status_code == 200:
            tokens = response.json()

            # Save tokens
            with open(ONEDRIVE_TOKEN_PATH, 'w') as f:
                json.dump(tokens, f, indent=2)

            await update.message.reply_text(
                "? Authentication successful!\n"
                "Tokens saved. Now test with /onedrive_test"
            )
        else:
            await update.message.reply_text(f"? Failed: {response.text}")

    except Exception as e:
        await update.message.reply_text(f"? Error: {str(e)[:200]}")

# ========== VALIDATION & UTILITY FUNCTIONS ==========

def validate_configuration() -> bool:
    """Validate all required configuration"""
    errors = []

    if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        errors.append("BOT_TOKEN not configured in .env")

    if not ALLOWED_USER_IDS:
        errors.append("ALLOWED_USER_IDS not configured in .env")

    if not EXCEL_AVAILABLE:
        errors.append("Excel libraries (openpyxl, pandas) not installed")

    if ONEDRIVE_CLIENT_ID and ONEDRIVE_CLIENT_ID == "YOUR_CLIENT_ID":
        logger.warning("OneDrive not configured - local mode only")

    if errors:
        for error in errors:
            logger.error(error)
        return False
    return True

def retry_onedrive_operation(max_retries: int = 3):
    """Decorator for retrying OneDrive operations"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            last_error = None
            for attempt in range(1, max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_error = e
                    logger.warning(f"Attempt {attempt}/{max_retries} failed: {e}")
                    if attempt < max_retries:
                        time.sleep(2 ** attempt)  # Exponential backoff
            raise last_error
        return wrapper
    return decorator

def validate_transaction_data(
    subcategory: str,
    currency_amounts: Dict[str, Optional[float]],
    payment_type: str
) -> Tuple[bool, str]:
    """Validate transaction data before processing"""

    # Check subcategory
    if not subcategory or len(subcategory.strip()) < 2:
        return False, "❌ Item name too short or empty"

    # Check amounts
    if not currency_amounts:
        return False, "❌ No amounts specified"

    valid_currencies = set(CURRENCIES.keys())
    for currency in currency_amounts:
        if currency not in valid_currencies:
            return False, f"❌ Invalid currency: {currency}"

    # Check for unrealistic amounts
    for currency, amount in currency_amounts.items():
        if amount is not None:
            if abs(amount) > 1000000:  # 1 million threshold
                return False, f"❌ Amount too large: {format_currency_amount(amount, currency)}"
            if amount == 0:
                return False, "❌ Amount cannot be zero"

    # Check payment type
    valid_payments = list(PAYMENT_TYPES.keys())
    if payment_type not in valid_payments:
        return False, f"❌ Invalid payment type. Use: {', '.join(valid_payments)}"

    return True, "✅ Valid transaction data"

def ensure_excel_file_exists() -> Tuple[bool, str]:
    """Ensure Excel file exists locally"""
    try:
        if LOCAL_COPY_PATH.exists():
            # Validate file
            file_size = LOCAL_COPY_PATH.stat().st_size
            if file_size < 1000:
                logger.warning("Excel file too small, downloading fresh copy")
                return copy_excel_from_onedrive()

            # Try to open it
            try:
                wb = load_workbook(str(LOCAL_COPY_PATH), read_only=True)
                wb.close()
                return True, "✅ Excel file is valid"
            except Exception:
                logger.warning("Excel file corrupted, downloading fresh copy")
                return copy_excel_from_onedrive()
        else:
            logger.info("Excel file not found locally, downloading")
            return copy_excel_from_onedrive()

    except Exception as e:
        return False, f"❌ Error ensuring Excel file: {str(e)[:200]}"

def cleanup_temp_files(max_age_hours: int = 24):
    """Clean up temporary and old files"""
    try:
        cutoff_time = time.time() - (max_age_hours * 3600)

        # Clean backup directory
        for file_path in BACKUP_DIR.glob("*.xlsm"):
            if file_path.stat().st_mtime < cutoff_time:
                file_path.unlink()
                logger.info(f"Cleaned up old backup: {file_path.name}")

        # Clean export directory
        for file_path in EXPORT_DIR.glob("*.csv"):
            if file_path.stat().st_mtime < cutoff_time:
                file_path.unlink()
                logger.info(f"Cleaned up old export: {file_path.name}")

        # Clean temp files in main directory
        for file_path in LOCAL_COPY_PATH.parent.glob("*.temp.*"):
            if file_path.stat().st_mtime < cutoff_time:
                file_path.unlink()
                logger.info(f"Cleaned up temp file: {file_path.name}")

    except Exception as e:
        logger.error(f"Cleanup error: {e}")

def sanitize_user_input(text: str, max_length: int = 500) -> str:
    """Sanitize user input to prevent injection attacks"""
    if not text:
        return ""

    # Remove control characters
    sanitized = ''.join(char for char in text if ord(char) >= 32)

    # Limit length
    if len(sanitized) > max_length:
        sanitized = sanitized[:max_length]

    return sanitized.strip()

# ========== EXCEL FUNCTIONS (OPENPYXL VERSION) ==========

@contextmanager
def excel_operation(path: Optional[Path] = None, *, keep_vba: bool = True, read_only: bool = False, data_only: bool = False):
    """
    Context manager for Excel operations using openpyxl.

    Use this when you need a safe way to open a workbook and ensure it is closed
    reliably even on errors.

    Parameters:
    - path: Path to the workbook. If None, LOCAL_COPY_PATH is used.
    - keep_vba: pass keep_vba to openpyxl.load_workbook (needed for .xlsm).
    - read_only: open workbook read-only (faster and safer for read ops).
    - data_only: if True, formulas are evaluated and values returned (openpyxl reads stored values).

    Usage (synchronous / in a background thread):
        with excel_operation(LOCAL_COPY_PATH, keep_vba=True, read_only=False) as wb:
            sheet = wb[TRACKING_SHEET_NAME]
            # do read/write with wb / sheet

    Important:
    - This context manager performs blocking I/O (load_workbook / close). Do NOT call it directly
      inside an asyncio event loop handler. Instead run the code that uses it in a thread:
          await asyncio.to_thread(_blocking_work)
      where _blocking_work contains the `with excel_operation(...)` block.
    """
    wb = None
    path_to_open: Optional[Path] = None

    try:
        # Resolve path to open
        path_to_open = LOCAL_COPY_PATH if path is None else Path(path)

        # Ensure path exists before trying to open
        if not path_to_open.exists():
            raise FileNotFoundError(f"Excel file not found: {path_to_open}")

        # Open workbook with requested options
        wb = load_workbook(str(path_to_open), keep_vba=keep_vba, read_only=read_only, data_only=data_only)

        yield wb

    except Exception as e:
        logger.error(f"Excel operation error for '{path_to_open}': {e}", exc_info=True)
        raise

    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception as close_err:
                logger.debug(f"Error closing workbook '{path_to_open}': {close_err}")

def save_excel_to_onedrive() -> Tuple[bool, str]:
    """Upload with file validation"""
    try:
        # Verify file exists and is valid
        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local file not found"

        file_size = LOCAL_COPY_PATH.stat().st_size
        if file_size < 1000:
            return False, "❌ File is too small (likely corrupted)"

        # Try to validate it's a valid Excel file
        try:
            # Quick test - try to open with pandas
            import pandas as pd
            # Just read header to verify
            df = pd.read_excel(LOCAL_COPY_PATH, sheet_name=0, nrows=1)
            logger.info(f"File validated, size: {file_size} bytes")
        except Exception as e:
            logger.error(f"File validation failed: {e}")
            return False, "❌ File appears corrupted"

        # Use simple direct upload
        return upload_to_onedrive_direct()

    except Exception as e:
        logger.error(f"Upload error: {e}")
        return False, f"❌ Upload error: {str(e)[:200]}"


def upload_to_onedrive_direct() -> Tuple[bool, str]:
    """Simple direct upload without sessions. Picks Content-Type based on file extension."""
    try:
        if requests is None:
            return False, "❌ requests library is not installed. Install with: pip install requests"

        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated"

        # Format path
        file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")

        # Determine appropriate Content-Type based on file extension
        suffix = Path(LOCAL_COPY_PATH).suffix.lower()
        if suffix == ".xlsm":
            content_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
        elif suffix == ".xlsx":
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            # fallback generic binary
            content_type = "application/octet-stream"

        # Simple PUT request
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/content"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": content_type
        }

        logger.info(f"Direct upload to: {ONEDRIVE_FILE_PATH} (Content-Type: {content_type})")

        with open(LOCAL_COPY_PATH, 'rb') as f:
            response = requests.put(url, headers=headers, data=f, timeout=60)

        if response.status_code in (200, 201):
            return True, "✅ File uploaded successfully"
        else:
            try:
                error_json = response.json()
                error_msg = error_json.get('error', {}).get('message', response.text)
            except Exception:
                error_msg = response.text

            # Check if it's a busy/locked error
            emsg_l = error_msg.lower()
            if "busy" in emsg_l or "423" in emsg_l or "same name" in emsg_l or "locked" in emsg_l:
                # Queue background sync (guarded by the upload lock to avoid duplicate submissions)
                global _bg_sync_future
                try:
                    with _onedrive_upload_lock:
                        # If there is no pending background future or the previous one finished, submit a new task
                        if _bg_sync_future is None or getattr(_bg_sync_future, "done", lambda: True)():
                            try:
                                _bg_sync_future = _background_uploader.submit(
                                    _simple_background_sync,
                                    attempts=5,
                                    base_sleep=3.0
                                )
                                logger.info("Background sync queued")
                            except Exception as e:
                                logger.warning(f"Failed to queue background sync: {e}")
                except Exception as e:
                    logger.warning(f"Error while attempting to schedule background sync: {e}")

                return False, "⚠️ OneDrive is busy/locked. Background sync queued."

            return False, f"❌ Upload error: {error_msg}"

    except Exception as e:
        logger.error(f"Direct upload exception: {e}", exc_info=True)
        return False, f"❌ Error: {str(e)[:200]}"

def _simple_background_sync(attempts: int = 5, base_sleep: float = 3.0):
    """Simple background sync"""
    for attempt in range(1, attempts + 1):
        logger.info(f"Background sync attempt {attempt}/{attempts}")
        time.sleep(base_sleep * attempt)

        success, msg = upload_to_onedrive_direct()
        if success:
            logger.info("Background sync succeeded")
            return

    logger.warning("Background sync failed after all attempts")

def _robust_background_sync(attempts: int = 8, base_sleep: float = 5.0):
    """Background sync with multiple strategies"""
    logger.info("Starting robust background sync...")

    for attempt in range(1, attempts + 1):
        logger.info(f"Background sync attempt {attempt}/{attempts}")

        # Try different strategies in order
        strategies = [
            ("Simple upload", simple_direct_upload),
            ("Wait and retry", lambda: wait_and_retry_upload(10)),
            ("Temp name upload", upload_to_different_name_then_move),
            ("Unique name upload", force_upload_with_unique_name)
        ]

        for strategy_name, strategy_func in strategies:
            logger.info(f"Trying {strategy_name}...")
            success, msg = strategy_func()

            if success:
                logger.info(f"Background sync succeeded with {strategy_name}")
                return

            # Check if we should continue
            error_lower = msg.lower()
            if "not authenticated" in error_lower:
                logger.error("Authentication failed in background sync")
                return

        # Wait before next full attempt cycle
        if attempt < attempts:
            sleep_time = base_sleep * attempt
            logger.info(f"All strategies failed, waiting {sleep_time}s before retry...")
            time.sleep(sleep_time)

    logger.error("Background sync exhausted all attempts")

def _retry_sync_background(attempts: int = 12, base_sleep: float = 2.0):
    """
    Background retry loop for busy/locked files. Waits if a foreground upload is running.
    """
    for attempt in range(1, attempts + 1):
        # If a foreground upload is happening, wait a bit and retry
        if _upload_in_progress.is_set():
            time.sleep(1.5)
            continue

        ok, msg = upload_to_onedrive()
        if ok:
            logger.info("Background sync succeeded.")
            return

        msg_l = msg.lower()
        if "423" in msg or "locked" in msg_l or "429" in msg or "same name is currently being uploaded" in msg_l:
            sleep_s = base_sleep * attempt + 0.3 * attempt
            logger.warning(f"Background sync attempt {attempt}/{attempts} failed: {msg}. Retrying in {sleep_s:.1f}s")
            time.sleep(sleep_s)
            continue

        logger.error(f"Background sync aborted: {msg}")
        return

    logger.error("Background sync exhausted all retries and still failed due to lock or concurrent upload.")


def _upload_backup_temp_copy() -> Tuple[bool, str]:
    """
    Upload the local file to OneDrive under a temporary name so that the latest
    state exists in the cloud even if the original path is locked.
    Example name: budget_tracker.temp_20260117_1845.xlsm
    """
    try:
        token = get_onedrive_token()
        if not token:
            return False, "Not authenticated"

        if not LOCAL_COPY_PATH.exists():
            return False, "Local file not found"

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        orig = ONEDRIVE_FILE_PATH
        folder = orig.rsplit("/", 1)[0] if "/" in orig else ""
        name = (orig.rsplit("/", 1)[1] if "/" in orig else orig).replace(".xlsm", f".temp_{ts}.xlsm")
        temp_path = f"{folder}/{name}" if folder else name

        # Create an upload session to the temp path
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:{temp_path}:/createUploadSession"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        body = {"item": {"@microsoft.graph.conflictBehavior": "replace", "name": name}}
        r = requests.post(url, headers=headers, json=body, timeout=30)
        if r.status_code not in (200, 201):
            try:
                err_msg = r.json().get("error", {}).get("message", r.text)
            except Exception:
                err_msg = r.text
            return False, f"Temp session error: {err_msg}"

        upload_url = r.json().get("uploadUrl")
        if not upload_url:
            return False, "Temp session missing uploadUrl"

        with open(LOCAL_COPY_PATH, "rb") as f:
            data = f.read()
        size = len(data)

        put_headers = {"Content-Length": str(size), "Content-Range": f"bytes 0-{size-1}/{size}"}
        pr = requests.put(upload_url, headers=put_headers, data=data, timeout=60)
        if pr.status_code in (200, 201):
            return True, "Temp backup uploaded"
        try:
            err_msg = pr.json().get("error", {}).get("message", pr.text)
        except Exception:
            err_msg = pr.text
        return False, f"Temp upload error: {err_msg}"
    except Exception as e:
        return False, f"Temp backup exception: {str(e)[:200]}"


def load_all_tables_with_details() -> Dict[str, Dict]:
    """
    Load ALL Excel Tables from the Dropdown Data sheet with detailed information and caching.
    First tries real Excel tables (ListObjects). If none are found, falls back to the
    existing column-scan heuristic.
    """
    global _table_cache, _table_cache_timestamp

    if not EXCEL_AVAILABLE:
        logger.error("openpyxl is not available. Cannot load tables.")
        return {}

    def _fallback_column_scan(ws) -> Dict[str, Dict]:
        """Existing behavior: treat row 1 as categories and rows below as items."""
        tables_dict: Dict[str, Dict] = {}
        max_col = ws.max_column
        max_row = ws.max_row
        for col in range(1, max_col + 1):
            category = ws.cell(row=1, column=col).value
            if not category:
                continue
            subcategories: List[str] = []
            variations: Dict[str, str] = {}
            for row in range(2, max_row + 1):
                cell_val = ws.cell(row=row, column=col).value
                if not cell_val:
                    continue
                original_text = str(cell_val).strip()
                if not original_text:
                    continue
                subcategories.append(original_text)
                normalized = normalize_text_for_matching(original_text)
                variations[normalized] = original_text
                # add individual words to improve matching
                for word in set(normalized.split()):
                    if len(word) > 3 and word not in variations:
                        variations[word] = original_text

            if subcategories:
                tables_dict[str(category)] = {
                    "original_name": str(category),
                    "subcategories": subcategories,
                    "variations": variations,
                    "count": len(subcategories),
                }
        return tables_dict

    try:
        # cache check
        if LOCAL_COPY_PATH.exists():
            current_mtime = LOCAL_COPY_PATH.stat().st_mtime
            if _table_cache_timestamp and _table_cache_timestamp == current_mtime and _table_cache:
                return _table_cache

        # keep_vba=True preserves macros in .xlsm files
        wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True, read_only=False, keep_vba=True)
        ws = wb[DROPDOWN_SHEET_NAME]

        tables_dict: Dict[str, Dict] = {}
        total_items = 0

        # openpyxl exposes real Excel Tables under ws.tables
        # Each table has .name and .ref (range like 'A1:C10')
        if getattr(ws, "tables", None) and len(ws.tables) > 0:
            for tbl in ws.tables.values():
                try:
                    table_name = getattr(tbl, "displayName", None) or getattr(tbl, "name", None) or "UnnamedTable"
                    ref = tbl.ref  # e.g., "B2:F25"
                    min_col, min_row, max_col, max_row = range_boundaries(ref)

                    subcategories: List[str] = []
                    variations: Dict[str, str] = {}

                    # Skip header row inside the table: start at min_row + 1
                    for row_vals in ws.iter_rows(
                        min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
                    ):
                        for val in row_vals:
                            if val is None:
                                continue
                            original_text = str(val).strip()
                            if not original_text:
                                continue
                            subcategories.append(original_text)
                            normalized = normalize_text_for_matching(original_text)
                            variations[normalized] = original_text
                            # add individual words as variations
                            for word in set(normalized.split()):
                                if len(word) > 3 and word not in variations:
                                    variations[word] = original_text

                    if subcategories:
                        tables_dict[table_name] = {
                            "original_name": table_name,
                            "subcategories": subcategories,
                            "variations": variations,
                            "count": len(subcategories),
                        }
                        total_items += len(subcategories)
                except Exception as inner:
                    logger.warning(f"Error reading table: {inner}")

            wb.close()

            _table_cache = tables_dict
            _table_cache_timestamp = current_mtime if LOCAL_COPY_PATH.exists() else None

            logger.info(f"Loaded {len(tables_dict)} tables from Excel")
            print(f"✅ Loaded {len(tables_dict)} categories with {total_items} items")

            # If for some reason table read returned nothing, try fallback
            if not tables_dict:
                logger.info("No items found inside Excel tables. Falling back to column scan.")
                tables_dict = _fallback_column_scan(ws)

            return tables_dict

        # No Excel tables present on the sheet. Use fallback.
        logger.info("No Excel tables found on sheet. Falling back to column scan.")
        tables_dict = _fallback_column_scan(ws)
        wb.close()

        _table_cache = tables_dict
        _table_cache_timestamp = current_mtime if LOCAL_COPY_PATH.exists() else None
        logger.info(f"Loaded {len(tables_dict)} tables from Excel (fallback)")
        print(f"✅ Loaded {len(tables_dict)} categories with {sum(d['count'] for d in tables_dict.values())} items")
        return tables_dict

    except Exception as e:
        logger.error(f"Error loading tables: {str(e)}", exc_info=True)
        return {}

def add_transaction_smart(
    subcategory_input: str,
    currency_amounts: Dict[str, Optional[float]],
    payment_type: str,
    optional_notes: str = ""
) -> Tuple[bool, str]:
    """Add transaction with smart matching - supports optional notes"""
    try:
        # Always use "Expenses" as transaction type (even for negative amounts)
        transaction_type = "Expenses"

        logger.info(
            f"Processing transaction: '{subcategory_input}', amounts: {currency_amounts}, "
            f"payment: {payment_type}, notes: '{optional_notes}'"
        )

        # 1) Copy fresh file from OneDrive
        success, msg = copy_excel_from_onedrive()
        if not success:
            logger.error(f"Failed to copy file: {msg}")
            return False, msg

        # 2) Load tables and find best match
        tables_dict = load_all_tables_with_details()
        if not tables_dict:
            logger.error("No tables found in Dropdown Data sheet")
            return False, "❌ No tables found in Dropdown Data sheet"

        matched_original, category, match_type, confidence = find_best_match_for_input(
            subcategory_input, tables_dict
        )
        confidence_percentage = confidence * 100

        # Reject if below threshold, include helpful suggestions
        if confidence < MINIMUM_CONFIDENCE:
            suggestions = []

            # Collect common suggestions
            all_suggestions = []
            for cat_name, data in tables_dict.items():
                for item in data['subcategories'][:5]:
                    all_suggestions.append((cat_name, item))
            all_suggestions.sort(key=lambda x: x[1])

            suggestions.append("\n<b>Available categories:</b>")
            for cat_name, _ in tables_dict.items():
                suggestions.append(f" • {cat_name}")

            suggestions.append("\n<b>Common items:</b>")
            displayed_items = 0
            for cat_name, item in all_suggestions:
                if displayed_items >= 8:
                    break
                suggestions.append(f" • {item} ({cat_name})")
                displayed_items += 1

            suggestions_text = "\n".join(suggestions)

            rejection_msg = (
                f"❌ <b>Transaction Rejected:</b>\n\n"
                f"<b>Input:</b> <code>{subcategory_input}</code>\n"
                f"<b>Confidence:</b> {confidence_percentage:.1f}% (minimum required: {MINIMUM_CONFIDENCE*100:.0f}%)\n"
                f"<b>Reason:</b> {match_type}\n\n"
                f"<b>No good match found. Did you mean one of these?</b>\n"
                f"{suggestions_text}\n\n"
                f"<i>Tip: Use more specific terms or check spelling.</i>"
            )
            logger.warning(
                f"Transaction rejected: {subcategory_input} (confidence: {confidence_percentage:.1f}%)"
            )
            return False, rejection_msg

        logger.info(
            f"Match found: '{subcategory_input}' → '{matched_original}' → '{category}' "
            f"({match_type}, {confidence_percentage:.1f}%)"
        )

        # 3) Append transaction via Microsoft Graph to Excel Table (preserves shapes + recalculates formulas)
        now_date = datetime.now().date()

        usd = float(currency_amounts.get("USD")) if currency_amounts.get("USD") is not None else None
        lbp = float(currency_amounts.get("LBP")) if currency_amounts.get("LBP") is not None else None
        eur = float(currency_amounts.get("EURO")) if currency_amounts.get("EURO") is not None else None

        # Prepare details/notes safely (limit length like before)
        details = None
        if optional_notes:
            details = optional_notes.strip()
            if len(details) > 500:
                details = details[:497] + "..."

        graph_ok, graph_msg = append_transaction_to_tracking_table(
            date_value=now_date,
            payment=payment_type,
            tx_type=transaction_type,
            category=category,
            subcategory=matched_original,
            usd=usd,
            lbp=lbp,
            euro=eur,
            details=details,
        )

        if not graph_ok:
            logger.error(f"Graph append failed: {graph_msg}")
            return False, graph_msg

        # Graph succeeded; Excel Online will recalc formulas automatically
        save_success, save_msg = True, graph_msg
        
        # Build response
        amount_display_parts = []
        total_amounts = len(currency_amounts)

        for currency, amount in currency_amounts.items():
            if amount is not None:
                amt_str = format_currency_amount(amount, currency)
                amount_display_parts.append(f"🔄 {amt_str}" if amount < 0 else amt_str)

        if amount_display_parts:
            amount_display = " + ".join(amount_display_parts) if total_amounts > 1 else amount_display_parts[0]
        else:
            amount_display = "No amount"

        payment_emoji = PAYMENT_TYPES.get(payment_type, "💵")
        has_negative = any(a is not None and a < 0 for a in currency_amounts.values())
        if has_negative:
            type_emoji = "🔄"
            transaction_desc = "Expense (with refunds/corrections)"
        else:
            type_emoji = "📤"
            transaction_desc = "Expense"

        currency_emojis = []
        for cur in currency_amounts.keys():
            if cur == "USD":
                currency_emojis.append("💵")
            elif cur == "LBP":
                currency_emojis.append("🇱🇧")
            elif cur == "EURO":
                currency_emojis.append("💶")
        currency_emoji_str = " ".join(currency_emojis)

        import html
        safe_input = html.escape(subcategory_input)
        safe_matched = html.escape(matched_original)
        safe_category = html.escape(category)
        safe_payment = html.escape(payment_type)
        safe_notes = html.escape(optional_notes) if optional_notes else ""
        safe_confidence = f"{confidence_percentage:.1f}%"
        multi_currency_indicator = "🌐 " if total_amounts > 1 else ""
        notes_section = f"\n• <b>Notes:</b> {safe_notes}" if optional_notes else ""

        if save_success:
            message = (
                f"{multi_currency_indicator}✅ {type_emoji} {payment_emoji} {currency_emoji_str} "
                f"<b>{transaction_desc} Added:</b>\n\n"
                f"• <b>You typed:</b> <code>{safe_input}</code>\n"
                f"• <b>Recorded as:</b> <code>{safe_matched}</code>\n"
                f"• <b>Category:</b> {safe_category}\n"
                f"• <b>Amount{'s' if total_amounts > 1 else ''}:</b> {amount_display}\n"
                f"• <b>Payment:</b> {safe_payment}"
                f"{notes_section}\n"
                f"• <b>Confidence:</b> {safe_confidence}\n"
                f"• <b>Match type:</b> {match_type}"
            )
            logger.info(f"Transaction added successfully: {transaction_desc}")
            return True, message
        else:
            message = (
                f"{multi_currency_indicator}⚠️ {type_emoji} {payment_emoji} {currency_emoji_str} "
                f"<b>{transaction_desc} Added Locally:</b>\n\n"
                f"• <b>You typed:</b> <code>{safe_input}</code>\n"
                f"• <b>Recorded as:</b> <code>{safe_matched}</code>\n"
                f"• <b>Category:</b> {safe_category}\n"
                f"• <b>Amount{'s' if total_amounts > 1 else ''}:</b> {amount_display}\n"
                f"• <b>Payment:</b> {safe_payment}"
                f"{notes_section}\n\n"
                f"<b>⚠️ OneDrive Sync Failed:</b>\n{save_msg}"
            )
            logger.warning(f"Transaction added locally but OneDrive sync failed: {save_msg}")
            return True, message

    except Exception as e:
        logger.error(f"Error in add_transaction_smart: {str(e)}", exc_info=True)
        import html
        error_msg = html.escape(str(e)[:200])
        return False, f"❌ <b>Error:</b> {error_msg}"

def _append_row_to_table(ws, tbl, row_values: Dict[str, Any]) -> int:
    """
    Append a row to an openpyxl table `tbl` on worksheet `ws`.
    row_values is a dict with keys:
      'date', 'payment', 'type', 'category', 'subcategory', 'USD', 'LBP', 'EURO', 'details'
    Returns the row number written, or 0 on failure.
    """
    try:
        ref = tbl.ref  # e.g. "C11:K133"
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        header_row = min_row
        new_row = max_row + 1

        # Build new table ref that extends table down by 1 row
        new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_row}"
        tbl.ref = new_ref  # update table ref in memory

        # Map headers -> target cell column
        headers = {}
        for col in range(min_col, max_col + 1):
            hdr_val = ws.cell(row=header_row, column=col).value
            # Coerce to string safely before strip/lower
            if hdr_val is None:
                hdr_text = ""
            else:
                hdr_text = str(hdr_val).strip().lower()
            headers[col] = hdr_text

        # Helper: try to coerce a string date into a datetime (returns datetime or None)
        def _coerce_to_datetime(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val
            if isinstance(val, str):
                # try common formats (include the pattern you observed: '19-Jan-26')
                fmts = ['%d-%b-%y', '%d-%b-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']
                for fmt in fmts:
                    try:
                        return datetime.strptime(val.strip(), fmt)
                    except Exception:
                        continue
            return None

        # Helper: reduce a sample Excel number_format to a date-only format by stripping time tokens.
        def _date_only_format(fmt: str) -> str:
            if not fmt:
                return ''
            # Split on whitespace and keep only parts that look like date tokens (d/m/y/Y)
            parts = re.split(r'\s+', str(fmt))
            date_parts = []
            for p in parts:
                # skip parts that clearly contain time tokens
                if re.search(r'[hHsS]|AM/PM|am/pm', p):
                    continue
                # keep parts that contain date letters
                if re.search(r'[dmyDMY]', p):
                    date_parts.append(p)
            if date_parts:
                return ' '.join(date_parts).strip()
            # fallback: take substring before any 'h' or 'H'
            m = re.split(r'(?=[hH])', fmt, maxsplit=1)
            return m[0].strip()

        # For each column in the table, determine which value to write
        for col in range(min_col, max_col + 1):
            hdr = headers[col]
            val = None

            # Basic heuristics to map header text to our values
            if "date" in hdr:
                # prefer a datetime/date object without time component
                raw = row_values.get('date')
                coerced = _coerce_to_datetime(raw)
                if isinstance(coerced, datetime):
                    # convert to date (drop time)
                    val = coerced.date()
                elif isinstance(raw, datetime):
                    val = raw.date()
                else:
                    # raw might already be a date object, or None
                    val = raw
            elif "payment" in hdr:
                val = row_values.get('payment')
            elif ("type" in hdr) and ("payment" not in hdr):
                val = row_values.get('type')
            elif ("sub" in hdr and "cat" in hdr) or "sub-category" in hdr or "sub category" in hdr:
                val = row_values.get('subcategory')
            elif "category" in hdr and "sub" not in hdr:
                val = row_values.get('category')
            elif "lbp" in hdr or "amount lbp" in hdr:
                val = row_values.get('LBP')
            elif "€" in hdr or "euro" in hdr or "amount €" in hdr or "amount eur" in hdr:
                val = row_values.get('EURO')
            elif "$" in hdr or "amount $" in hdr or ("amount" in hdr and "lbp" not in hdr and "euro" not in hdr):
                # Prefer USD for generic "Amount" column
                val = row_values.get('USD')
            elif "detail" in hdr or "note" in hdr:
                val = row_values.get('details')
            else:
                # fallback: attempt matches by keywords
                if "usd" in hdr or "amount $" in hdr or "$" in hdr:
                    val = row_values.get('USD')
                elif "lbp" in hdr:
                    val = row_values.get('LBP')
                elif "euro" in hdr:
                    val = row_values.get('EURO')

            # write the value into the new row cell
            cell = ws.cell(row=new_row, column=col)

            # If this is the date column and we have a date-like value, set a date-only object and preserve table format
            if "date" in hdr and val is not None:
                # If val is datetime.datetime, convert to date; if it's date, keep
                write_val = val
                try:
                    # If it's a datetime-like string, try coercion again
                    if isinstance(write_val, datetime):
                        write_val = write_val.date()
                except Exception:
                    pass

                # Try to detect the table's existing number format for this column by scanning the current table rows
                sample_format = None
                try:
                    for r in range(min_row + 1, max_row + 1):
                        sample_cell = ws.cell(row=r, column=col)
                        if sample_cell.value is not None:
                            sample_format = sample_cell.number_format
                            break
                except Exception:
                    sample_format = None

                cell.value = write_val

                # Apply the sample format if present and non-empty; otherwise use US short date
                if sample_format and str(sample_format).strip() != '':
                    # reduce to date-only if sample includes time
                    cleaned = _date_only_format(sample_format)
                    if cleaned:
                        cell.number_format = cleaned
                    else:
                        cell.number_format = 'm/d/yyyy'
                else:
                    cell.number_format = 'm/d/yyyy'
            else:
                # non-date columns: write value directly
                cell.value = val

        # openpyxl keeps the ws.tables dict updated by ref mutation
        logger.debug(f"Appended row {new_row} to table {tbl.name} (new ref {tbl.ref})")
        return new_row

    except Exception as e:
        logger.error(f"Error appending row to table: {e}", exc_info=True)
        return 0

def get_recent_transactions(count: int = 10) -> List[Dict]:
    """Get multiple recent transactions from Excel"""
    transactions = []

    if not EXCEL_AVAILABLE:
        return transactions

    try:
        # Copy fresh file first
        copy_excel_from_onedrive()

        with excel_lock:
            wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
            sheet = wb[TRACKING_SHEET_NAME]

            # Find the last row with data
            last_row = 12
            while sheet[f'C{last_row}'].value not in [None, ""]:
                last_row += 1
                if last_row > 1000:
                    break

            # Get the last 'count' transactions
            start_row = max(12, last_row - count)

            for row in range(start_row, last_row):
                date_val = sheet[f'C{row}'].value
                if date_val:
                    transaction = {
                        'row': row,
                        'date': date_val,
                        'payment': sheet[f'D{row}'].value or "Cash",
                        'type': sheet[f'E{row}'].value or "Expenses",
                        'category': sheet[f'F{row}'].value or "",
                        'subcategory': sheet[f'G{row}'].value or "",
                        'usd': sheet[f'H{row}'].value,
                        'lbp': sheet[f'I{row}'].value,
                        'euro': sheet[f'J{row}'].value,
                        'notes': sheet[f'K{row}'].value or ""
                    }
                    transactions.append(transaction)

            wb.close()

        # Return in reverse order (most recent first)
        return list(reversed(transactions))

    except Exception as e:
        logger.error(f"Error getting recent transactions: {str(e)}")
        return []


# ========== DELETE / MODIFY FUNCTIONS (openpyxl) ==========

def delete_transaction_at_row(row: int) -> Tuple[bool, str, Optional[Dict]]:
    """Delete a transaction using Graph table row delete (preserves workbook/shapes)."""
    try:
        logger.info(f"Deleting transaction at row {row}")

        # 1) Download fresh copy (only to read what we're deleting)
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}", None

        # 2) Read transaction data before deleting (for confirmation message)
        deleted_transaction = None
        try:
            wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True, read_only=True, keep_vba=True)
            sheet = wb[TRACKING_SHEET_NAME]

            if row < 12 or row > 1000:
                wb.close()
                return False, f"❌ Invalid row: {row}", None

            cell_value = sheet[f"C{row}"].value
            if cell_value in [None, ""]:
                wb.close()
                return False, f"❌ No transaction at row {row}", None

            deleted_transaction = {
                "row": row,
                "date": cell_value,
                "payment": sheet[f"D{row}"].value or "Cash",
                "type": sheet[f"E{row}"].value or "Expenses",
                "category": sheet[f"F{row}"].value or "",
                "subcategory": sheet[f"G{row}"].value or "",
                "usd": sheet[f"H{row}"].value,
                "lbp": sheet[f"I{row}"].value,
                "euro": sheet[f"J{row}"].value,
                "notes": sheet[f"K{row}"].value or "",
            }
            wb.close()
        except Exception as e:
            logger.error(f"Error reading transaction: {e}")
            return False, "❌ Error reading Excel file", None

        # 3) Delete via Graph (table API)
        del_ok, del_msg = graph_delete_transaction_at_row(row)
        if not del_ok:
            return False, del_msg, None

        # 4) Build response
        date_val = deleted_transaction["date"]
        date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, "strftime") else str(date_val)
        item = deleted_transaction["subcategory"] or deleted_transaction["category"]

        response = (
            f"✅ <b>Transaction Deleted Successfully!</b>\n\n"
            f"🗑️ <b>Deleted:</b> {item}\n"
            f"📅 <b>Date:</b> {date_str}\n"
            f"{del_msg}"
        )
        return True, response, deleted_transaction

    except Exception as e:
        logger.error("Delete error", exc_info=True)
        return False, f"❌ Error: {str(e)[:200]}", None

async def debug_onedrive_command(update, context):
    """Debug OneDrive issues"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    await update.message.reply_text("🔍 Running OneDrive diagnostics...")

    # Check local file
    local_exists = LOCAL_COPY_PATH.exists()
    local_size = LOCAL_COPY_PATH.stat().st_size if local_exists else 0

    # Check token
    token = get_onedrive_token()
    token_status = "✅ Valid" if token else "❌ Invalid/Missing"

    # Try a test operation
    test_success, test_msg = simple_direct_upload()

    response = (
        f"<b>OneDrive Diagnostics:</b>\n\n"
        f"📁 <b>Local File:</b> {'✅ Exists' if local_exists else '❌ Missing'}\n"
        f"📏 <b>Size:</b> {local_size:,} bytes\n"
        f"🔐 <b>Token:</b> {token_status}\n"
        f"🔄 <b>Test Upload:</b> {'✅ Success' if test_success else '❌ Failed'}\n\n"
    )

    if not test_success:
        response += f"<b>Error:</b> {test_msg}\n\n"

    # Check background sync status
    if _bg_sync_future and not _bg_sync_future.done():
        response += "🔄 <b>Background Sync:</b> Running\n"
    elif _upload_in_progress.is_set():
        response += "🔄 <b>Upload:</b> In Progress\n"
    else:
        response += "🔄 <b>Background Sync:</b> Idle\n"

    response += (
        f"\n<b>Quick Fixes:</b>\n"
        f"1. Wait 1-2 minutes for any ongoing sync\n"
        f"2. Use /sync to force a new attempt\n"
        f"3. Check OneDrive website for file locks\n"
    )

    await update.message.reply_text(response, parse_mode='HTML')

def delete_last_transaction() -> Tuple[bool, str, Optional[Dict]]:
    """Delete the last transaction with proper error handling"""
    try:
        logger.info("Starting delete last transaction operation")

        # Step 1: Download fresh copy
        success, msg = copy_excel_from_onedrive()
        if not success:
            logger.error(f"Failed to download: {msg}")
            return False, f"❌ {msg}", None

        # Step 2: Find last transaction
        with excel_lock:
            if not EXCEL_AVAILABLE:
                return False, "❌ openpyxl not installed", None

            wb = load_workbook(str(LOCAL_COPY_PATH))
            sheet = wb[TRACKING_SHEET_NAME]

            # Find last row
            last_row = 12
            while sheet[f"C{last_row}"].value not in [None, ""]:
                last_row += 1
                if last_row > 1000:
                    break

            if last_row == 12:
                wb.close()
                return False, "❌ No transactions found to delete", None

            row_to_delete = last_row - 1
            wb.close()

        # Step 3: Delete using the robust function
        return delete_transaction_at_row(row_to_delete)

    except Exception as e:
        logger.error(f"Error in delete_last_transaction: {e}", exc_info=True)
        return False, f"❌ Error: {str(e)[:200]}", None

def modify_transaction_at_row(
    row: int,
    new_subcategory: Optional[str] = None,
    new_currency_amounts: Optional[Dict[str, Optional[float]]] = None,
    new_payment_type: Optional[str] = None,
    new_notes: Optional[str] = None,
) -> Tuple[bool, str, Optional[Dict]]:
    """Modify selected fields at a specific row using Microsoft Graph (preserves shapes/formulas)."""
    try:
        if row < 12 or row > 1000:
            return False, f"❌ Invalid row number: {row}", None

        # Download fresh copy only for reading original + matching tables
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}", None

        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl not installed. Install: pip install openpyxl pandas", None

        with excel_lock:
            wb = load_workbook(str(LOCAL_COPY_PATH), keep_vba=True, read_only=True, data_only=True)
            sheet = wb[TRACKING_SHEET_NAME]

            if sheet[f"C{row}"].value in [None, ""]:
                wb.close()
                return False, f"❌ No transaction found at row {row}", None

            original = {
                "row": row,
                "date": sheet[f"C{row}"].value,
                "payment": sheet[f"D{row}"].value or "Cash",
                "type": sheet[f"E{row}"].value or "Expenses",
                "category": sheet[f"F{row}"].value or "",
                "subcategory": sheet[f"G{row}"].value or "",
                "usd": sheet[f"H{row}"].value,
                "lbp": sheet[f"I{row}"].value,
                "euro": sheet[f"J{row}"].value,
                "notes": sheet[f"K{row}"].value or "",
            }
            wb.close()

        modified = original.copy()
        changes: List[str] = []

        # 1) Subcategory change with smart matching
        if new_subcategory:
            tables = load_all_tables_with_details()
            if not tables:
                return False, "❌ No tables found in Dropdown Data sheet", None

            matched, category, match_type, confidence = find_best_match_for_input(new_subcategory, tables)
            if confidence < MINIMUM_CONFIDENCE:
                return False, (
                    f"❌ Modification Rejected\n\n"
                    f"New item: {new_subcategory}\n"
                    f"Confidence: {confidence*100:.1f}% (need ≥ {MINIMUM_CONFIDENCE*100:.0f}%)\n"
                    f"Reason: {match_type}"
                ), None

            modified["category"] = category
            modified["subcategory"] = matched
            changes.append(f"Item: '{original['subcategory']}' → '{matched}'")

        # 2) Payment type
        if new_payment_type is not None:
            modified["payment"] = new_payment_type
            changes.append(f"Payment: '{original['payment']}' → '{new_payment_type}'")

        # 3) Currency amounts
        if new_currency_amounts:
            # Clear first
            modified["usd"] = None
            modified["lbp"] = None
            modified["euro"] = None

            for curr, amt in new_currency_amounts.items():
                if curr == "USD":
                    modified["usd"] = float(amt) if amt is not None else None
                elif curr == "LBP":
                    modified["lbp"] = float(amt) if amt is not None else None
                elif curr == "EURO":
                    modified["euro"] = float(amt) if amt is not None else None

            # change summary
            old_parts, new_parts = [], []
            for curr in ["USD", "LBP", "EURO"]:
                o = original.get(curr.lower())
                n = modified.get(curr.lower())
                if o is not None or n is not None:
                    old_parts.append(format_currency_amount(o, curr) if o is not None else "None")
                    new_parts.append(format_currency_amount(n, curr) if n is not None else "None")
            changes.append(f"Amount: {' + '.join(old_parts) if old_parts else 'None'} → {' + '.join(new_parts) if new_parts else 'None'}")

        # 4) Details/Notes column (K)
        # If user provided notes, replace the base notes; if they provided empty, clear; if None, keep existing.
        details_value = original.get("notes") or ""
        if new_notes is not None:
            if new_notes.strip():
                details_value = new_notes.strip()
            else:
                details_value = ""

        # 5) Append a LIMITED modification marker
        if changes:
            marker = f"[Modified {datetime.now().strftime('%Y-%m-%d %H:%M')}]"
            # Only add marker once (don’t keep stacking)
            if "[Modified" not in details_value:
                details_value = (details_value + ("\n\n" if details_value else "") + marker).strip()

        modified["notes"] = details_value

        # If nothing changed, return early
        if not changes and new_notes is None:
            return True, "ℹ️ Nothing changed.", original

        # Apply update via Graph (this preserves shapes + formulas)
        graph_ok, graph_msg = graph_update_transaction_at_row(
            row,
            date_value=modified["date"],
            payment=modified["payment"],
            tx_type=modified["type"],
            category=modified["category"],
            subcategory=modified["subcategory"],
            usd=modified.get("usd"),
            lbp=modified.get("lbp"),
            euro=modified.get("euro"),
            details=modified.get("notes"),
        )

        if not graph_ok:
            return False, graph_msg, None

        success_msg = "✅ Transaction modified successfully!\n\n" + "\n".join([f"• {c}" for c in changes])
        return True, success_msg, modified

    except Exception as e:
        logger.error("Error modifying transaction", exc_info=True)
        return False, f"❌ Error modifying transaction: {str(e)[:200]}", None

# ========== NEW: DOWNLOAD/EXPORT FUNCTIONS ==========

def setup_export_directory():
    """Create export directory if it doesn't exist"""
    EXPORT_DIR.mkdir(exist_ok=True)

def export_to_csv(time_range: str = "month") -> Tuple[bool, str]:
    """
    Export transactions to CSV format using pandas
    time_range: "day", "week", "month", "year", "all"
    """
    try:
        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl not installed"

        # Copy fresh file
        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}"

        # Use pandas to read Excel
        df = pd.read_excel(
            LOCAL_COPY_PATH,
            sheet_name=TRACKING_SHEET_NAME,
            header=10,  # Skip first 10 rows if they're headers
            usecols='C:K'  # Columns C to K
        )

        # Rename columns
        df.columns = ['Date', 'Payment', 'Type', 'Category', 'Subcategory', 'USD', 'LBP', 'EURO', 'Notes']

        # Filter by date if needed
        if time_range != "all":
            today = datetime.now()
            if time_range == "day":
                start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
            elif time_range == "week":
                start_date = today - timedelta(days=today.weekday())
                start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            elif time_range == "month":
                start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            elif time_range == "year":
                start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

            df = df[pd.to_datetime(df['Date']) >= start_date]

        if df.empty:
            return False, "❌ No transactions found"

        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transactions_{time_range}_{timestamp}.csv"
        filepath = EXPORT_DIR / filename

        # Save to CSV
        df.to_csv(filepath, index=False, encoding='utf-8')

        logger.info(f"Exported {len(df)} transactions to {filepath}")
        return True, str(filepath)

    except Exception as e:
        logger.error(f"Export error: {str(e)}", exc_info=True)
        return False, f"❌ Export error: {str(e)[:200]}"

def export_summary() -> Tuple[bool, str]:
    """Export summary statistics"""
    try:
        if not EXCEL_AVAILABLE:
            return False, "❌ openpyxl not installed"

        success, msg = copy_excel_from_onedrive()
        if not success:
            return False, f"❌ {msg}"

        # Use pandas to read Excel
        df = pd.read_excel(
            LOCAL_COPY_PATH,
            sheet_name=TRACKING_SHEET_NAME,
            header=10,
            usecols='C:K'
        )

        df.columns = ['Date', 'Payment', 'Type', 'Category', 'Subcategory', 'USD', 'LBP', 'EURO', 'Notes']

        # Calculate totals
        usd_total = df['USD'].sum() if 'USD' in df.columns else 0
        lbp_total = df['LBP'].sum() if 'LBP' in df.columns else 0
        euro_total = df['EURO'].sum() if 'EURO' in df.columns else 0

        # Category breakdown
        categories = {}
        if 'Category' in df.columns and 'USD' in df.columns:
            for category in df['Category'].unique():
                if pd.notna(category):
                    cat_data = df[df['Category'] == category]
                    categories[category] = {
                        'usd': cat_data['USD'].sum(),
                        'lbp': cat_data['LBP'].sum() if 'LBP' in df.columns else 0,
                        'euro': cat_data['EURO'].sum() if 'EURO' in df.columns else 0,
                        'count': len(cat_data)
                    }

        # Create summary JSON
        summary = {
            'export_date': datetime.now().isoformat(),
            'total_transactions': len(df),
            'totals': {
                'USD': float(usd_total),
                'LBP': float(lbp_total),
                'EURO': float(euro_total)
            },
            'categories': categories,
            'file_info': {
                'source': str(LOCAL_COPY_PATH),
                'last_modified': datetime.fromtimestamp(LOCAL_COPY_PATH.stat().st_mtime).isoformat() if LOCAL_COPY_PATH.exists() else None
            }
        }

        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"summary_{timestamp}.json"
        filepath = EXPORT_DIR / filename

        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, default=str)

        logger.info(f"Exported summary to {filepath}")
        return True, str(filepath)

    except Exception as e:
        logger.error(f"Summary export error: {str(e)}", exc_info=True)
        return False, f"❌ Summary error: {str(e)[:200]}"

def create_backup_copy() -> Tuple[bool, str]:
    """Create a backup copy of the Excel file for download"""
    try:
        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local copy not found"

        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"budget_backup_{timestamp}.xlsm"
        backup_path = BACKUP_DIR / backup_filename

        shutil.copy2(LOCAL_COPY_PATH, backup_path)

        logger.info(f"Created downloadable backup: {backup_path}")
        return True, str(backup_path)

    except Exception as e:
        logger.error(f"Backup copy error: {str(e)}")
        return False, f"❌ Backup copy error: {str(e)[:200]}"

# ========== UTILITY FUNCTIONS ==========

def upload_to_different_name_then_move() -> Tuple[bool, str]:
    """Upload to a temporary filename, then move to correct location"""
    try:
        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated with OneDrive"

        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local file not found"

        # Create unique temporary filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_filename = f"temp_upload_{timestamp}.xlsm"

        # Get the directory path
        original_path = Path(ONEDRIVE_FILE_PATH)
        original_dir = str(original_path.parent) if original_path.parent != Path('.') else ""

        # Create temp path
        temp_path = f"{original_dir}/{temp_filename}" if original_dir else temp_filename

        logger.info(f"Uploading to temporary file: {temp_path}")

        # Upload to temp location
        with open(LOCAL_COPY_PATH, 'rb') as f:
            data_bytes = f.read()
        file_size = len(data_bytes)

        # Create upload session for temp file
        session_url = f"https://graph.microsoft.com/v1.0/me/drive/root:{temp_path}:/createUploadSession"
        session_headers = {"Authorization": f"Bearer {get_onedrive_token()}", "Content-Type": "application/json"}
        session_body = {"item": {"@microsoft.graph.conflictBehavior": "replace", "name": temp_filename}}

        resp = requests.post(session_url, headers=session_headers, json=session_body, timeout=30)
        if resp.status_code not in (200, 201):
            return False, f"❌ Temp session error: {resp.text}"

        upload_url = resp.json().get("uploadUrl")
        if not upload_url:
            return False, "❌ Temp session missing uploadUrl"

        # Upload the file
        content_range = f"bytes 0-{file_size-1}/{file_size}"
        put_headers = {"Content-Length": str(file_size), "Content-Range": content_range}

        put_resp = requests.put(upload_url, headers=put_headers, data=data_bytes, timeout=60)
        if put_resp.status_code not in (200, 201):
            return False, f"❌ Temp upload error: {put_resp.text}"

        # Now move/rename the temp file to the original location
        time.sleep(1)  # Wait for temp file to settle

        # Move file using Graph API
        move_url = f"https://graph.microsoft.com/v1.0/me/drive/root:{temp_path}"
        move_headers = {"Authorization": f"Bearer {get_onedrive_token()}", "Content-Type": "application/json"}
        move_body = {
            "name": original_path.name,
            "@microsoft.graph.conflictBehavior": "replace"
        }

        move_resp = requests.patch(move_url, headers=move_headers, json=move_body, timeout=30)
        if move_resp.status_code in (200, 201):
            logger.info(f"Successfully moved {temp_path} to {ONEDRIVE_FILE_PATH}")
            return True, "✅ File uploaded and moved successfully"
        else:
            # At least we have the temp file
            logger.warning(f"Could not move temp file, but it's saved as {temp_path}")
            return True, f"⚠️ File uploaded as temporary copy ({temp_filename})"

    except Exception as e:
        logger.error(f"Error in upload_to_different_name_then_move: {e}")
        return False, f"❌ Error: {str(e)[:200]}"

def force_upload_with_unique_name() -> Tuple[bool, str]:
    """Force upload with guaranteed unique name to avoid conflicts"""
    try:
        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated with OneDrive"

        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local file not found"

        # Create unique name with timestamp and random component
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        random_suffix = os.urandom(4).hex()
        original_name = Path(ONEDRIVE_FILE_PATH).name
        unique_name = f"{original_name.rsplit('.', 1)[0]}_{timestamp}_{random_suffix}.xlsm"

        # Upload to unique name
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{unique_name}:/content"
        headers = {"Authorization": f"Bearer {token}"}

        with open(LOCAL_COPY_PATH, 'rb') as f:
            response = requests.put(url, headers=headers, data=f, timeout=60)

        if response.status_code in (200, 201):
            logger.info(f"Uploaded to unique file: {unique_name}")
            return True, f"✅ File uploaded as {unique_name}"
        else:
            return False, f"❌ Upload error: {response.text}"

    except Exception as e:
        logger.error(f"Error in force_upload_with_unique_name: {e}")
        return False, f"❌ Error: {str(e)[:200]}"

def wait_and_retry_upload(max_wait_seconds: int = 30) -> Tuple[bool, str]:
    """Wait for any ongoing upload to complete, then retry"""
    import time

    logger.info("Waiting for any ongoing OneDrive uploads to complete...")

    # First, check if we can list files to see if OneDrive is responsive
    try:
        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated"

        # List recent files in the folder to check activity
        original_path = Path(ONEDRIVE_FILE_PATH)
        folder_path = str(original_path.parent) if original_path.parent != Path('.') else ""

        list_url = f"https://graph.microsoft.com/v1.0/me/drive/root:{folder_path}:/children"
        headers = {"Authorization": f"Bearer {token}"}

        response = requests.get(list_url, headers=headers, timeout=30)
        if response.status_code == 200:
            logger.info("OneDrive is responsive")
    except:
        pass

    # Wait with progressive backoff
    wait_intervals = [2, 3, 5, 7, 10]  # 27 seconds total
    total_wait = 0

    for interval in wait_intervals:
        if total_wait >= max_wait_seconds:
            break

        logger.info(f"Waiting {interval} seconds... (total: {total_wait}s)")
        time.sleep(interval)
        total_wait += interval

        # Try a simple upload after waiting
        success, msg = simple_direct_upload()
        if success:
            return True, msg

    return False, "❌ Could not upload after waiting"

def simple_direct_upload() -> Tuple[bool, str]:
    """Simple PUT upload without session - works better for small files"""
    try:
        token = get_onedrive_token()
        if not token:
            return False, "❌ Not authenticated with OneDrive"

        if not LOCAL_COPY_PATH.exists():
            return False, "❌ Local file not found"

        # Format path
        file_path = ONEDRIVE_FILE_PATH.replace(" ", "%20")

        # Simple PUT request
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:/content"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }

        logger.info(f"Simple direct upload to: {ONEDRIVE_FILE_PATH}")

        with open(LOCAL_COPY_PATH, 'rb') as f:
            response = requests.put(url, headers=headers, data=f, timeout=60)

        if response.status_code in (200, 201):
            logger.info("Simple upload successful!")
            return True, "✅ File uploaded successfully"
        else:
            try:
                error_json = response.json()
                error_msg = error_json.get('error', {}).get('message', response.text)
            except Exception:
                error_msg = response.text
            return False, f"❌ Upload error: {error_msg}"

    except Exception as e:
        logger.error(f"Error in simple_direct_upload: {e}")
        return False, f"❌ Error: {str(e)[:200]}"

def wait_for_onedrive_unlock(max_wait_seconds: int = 30) -> bool:
    """Wait for OneDrive to unlock the file"""
    import time
    start_time = time.time()

    while time.time() - start_time < max_wait_seconds:
        # Try a simple read operation to see if file is accessible
        try:
            with open(LOCAL_COPY_PATH, 'rb') as f:
                f.read(100)  # Just read a small portion
            logger.info("File appears to be unlocked")
            return True
        except (IOError, PermissionError) as e:
            logger.warning(f"File still locked, waiting... Error: {e}")
            time.sleep(2)

    logger.error("File remained locked after waiting")
    return False

def create_local_backup_before_operation() -> Optional[Path]:
    """Create a local backup before any destructive operation"""
    try:
        if not LOCAL_COPY_PATH.exists():
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_path = BACKUP_DIR / f"pre_delete_backup_{timestamp}.xlsm"

        shutil.copy2(LOCAL_COPY_PATH, backup_path)
        logger.info(f"Created pre-operation backup: {backup_path}")
        return backup_path
    except Exception as e:
        logger.error(f"Failed to create backup: {e}")
        return None

def force_close_excel_handles():
    """Force close any Excel file handles on Linux"""
    try:
        # Kill any processes that might have the file open
        commands = [
            ['lsof', '-t', str(LOCAL_COPY_PATH)],  # Get PIDs
            ['fuser', '-k', str(LOCAL_COPY_PATH)],  # Kill processes
        ]

        for cmd in commands:
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=5)
                if result.returncode == 0 and result.stdout:
                    logger.info(f"Closed handles using {cmd[0]}: {result.stdout}")
            except:
                pass

        time.sleep(1)  # Give system time to release handles
        return True
    except Exception as e:
        logger.error(f"Error closing handles: {e}")
        return False

def is_authorized(user_id: int) -> bool:
    """Check if user is authorized to use the bot"""
    return user_id in ALLOWED_USER_IDS

def format_currency_amount(amount: float, currency: str) -> str:
    """Format amount according to currency rules - supports negative numbers"""
    if amount is None:
        return "No amount"

    # If no currency, just return the number
    if not currency:
        return f"{amount:,.2f}"

    is_negative = amount < 0
    abs_amount = abs(amount)

    if currency not in CURRENCIES:
        currency = DEFAULT_CURRENCY

    config = CURRENCIES[currency]

    if config.thousands_separator:
        formatted = f"{abs_amount:,.{config.decimal_places}f}"
    else:
        formatted = f"{abs_amount:.{config.decimal_places}f}"

    symbol = config.symbols[0]

    if is_negative:
        if currency == "USD":
            return f"-{symbol}{formatted}"
        elif currency == "EURO":
            return f"-{symbol}{formatted}"
        elif currency == "LBP":
            return f"-{formatted} LBP"
        else:
            return f"-{formatted}"
    else:
        if currency == "USD":
            return f"{symbol}{formatted}"
        elif currency == "EURO":
            return f"{symbol}{formatted}"
        elif currency == "LBP":
            return f"{formatted} LBP"
        return formatted

def format_transaction_response(transaction: Dict, action: str = "Added") -> str:
    """Format a transaction for display in messages"""
    date_val = transaction['date']
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)

    # Format amounts
    amounts = []
    if transaction.get('usd') is not None:
        prefix = "🔄 " if transaction['usd'] < 0 else ""
        amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
    if transaction.get('lbp') is not None:
        prefix = "🔄 " if transaction['lbp'] < 0 else ""
        amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
    if transaction.get('euro') is not None:
        prefix = "🔄 " if transaction['euro'] < 0 else ""
        amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")

    amount_str = " + ".join(amounts) if amounts else "No amount"

    payment_emoji = "💳" if str(transaction.get('payment', '')).lower() == "card" else "💵"

    import html
    safe_item = html.escape(transaction.get('subcategory') or transaction.get('category', ''))
    safe_category = html.escape(transaction.get('category', 'Unknown'))

    response = (
        f"✅ <b>Transaction {action}:</b>\n\n"
        f"{payment_emoji} {date_str}: {safe_item}\n"
        f"Amount: {amount_str}\n"
        f"Category: {safe_category}\n"
        f"Payment: {html.escape(transaction.get('payment', 'Cash'))}"
    )

    notes = transaction.get('notes')
    if notes:
        response += f"\nNotes: {html.escape(str(notes)[:100])}{'...' if len(str(notes)) > 100 else ''}"

    return response

def create_backup():
    """Create backup of the Excel file"""
    try:
        BACKUP_DIR.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"backup_{timestamp}_{LOCAL_COPY_PATH.name}"
        shutil.copy2(LOCAL_COPY_PATH, backup_path)

        cleanup_old_backups()
        logger.info(f"Backup created: {backup_path}")
        return True
    except Exception as e:
        logger.error(f"Backup creation failed: {e}")
        return False

def cleanup_old_backups():
    """Remove backups older than retention period"""
    try:
        cutoff_date = datetime.now().timestamp() - (BACKUP_RETENTION_DAYS * 24 * 3600)

        for backup_file in BACKUP_DIR.glob("backup_*"):
            if backup_file.stat().st_mtime < cutoff_date:
                backup_file.unlink()
                logger.info(f"Removed old backup: {backup_file}")
    except Exception as e:
        logger.error(f"Backup cleanup failed: {e}")

def unlock_excel_file():
    """Try to unlock the Excel file if it's locked - Linux version"""
    try:
        if LOCAL_COPY_PATH.exists():
            # Try to remove lock files
            lock_files = [
                LOCAL_COPY_PATH.with_suffix('.lock'),
                Path(f"{LOCAL_COPY_PATH}.lock"),
                Path(f"~{LOCAL_COPY_PATH}.lock").expanduser()
            ]

            for lock_file in lock_files:
                if lock_file.exists():
                    try:
                        lock_file.unlink()
                        logger.info(f"Removed lock file: {lock_file}")
                    except:
                        pass

            # Kill any LibreOffice processes that might have the file open
            subprocess.run(['pkill', '-f', 'libreoffice'], capture_output=True)
            subprocess.run(['pkill', '-f', 'soffice'], capture_output=True)

            logger.info("Attempted to unlock Excel file")
            return True
        else:
            return False

    except Exception as e:
        logger.error(f"Unlock error: {e}")
        return False

def fix_excel_file_format(input_path: Path, output_path: Path) -> bool:
    """
    Fix Excel file format issues - specifically for .xlsm files on Linux.
    This ensures the file maintains proper .xlsm format.
    """
    try:
        if not input_path.exists():
            return False

        # First, make a backup
        backup_path = input_path.with_suffix('.original.xlsm')
        shutil.copy2(input_path, backup_path)

        # Method 1: Try to read with pandas and write with engine='openpyxl'
        try:
            import pandas as pd

            # Read all sheets
            excel_file = pd.ExcelFile(input_path, engine='openpyxl')
            sheet_names = excel_file.sheet_names

            # Write back with engine that supports .xlsm
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet in sheet_names:
                    df = pd.read_excel(input_path, sheet_name=sheet, engine='openpyxl')
                    df.to_excel(writer, sheet_name=sheet, index=False)

            # Check if output is valid
            if output_path.stat().st_size > 1000:
                logger.info(f"Fixed Excel format using pandas: {output_path.stat().st_size} bytes")
                return True
        except Exception as e:
            logger.warning(f"Pandas method failed: {e}")

        # Method 2: Direct copy (if file is already good)
        shutil.copy2(input_path, output_path)

        # Verify
        if output_path.stat().st_size > 1000:
            return True

        # Method 3: Restore from backup
        shutil.copy2(backup_path, output_path)
        return output_path.exists()

    except Exception as e:
        logger.error(f"Error fixing Excel format: {e}")
        return False

def save_excel_safely(filepath: Path, sheet_name: str, row: int, values: dict) -> bool:
    """
    Save to Excel safely without corrupting .xlsm format.

    Improvements:
    - Writes workbook to a NamedTemporaryFile in the same directory and fsyncs it.
    - Uses os.replace for atomic rename.
    - Preserves original file permissions (mode) and attempts to preserve ownership.
    - Uses numeric coordinate handling for data validation removal (existing logic preserved).
    """
    import tempfile
    try:
        # Ensure directory exists
        parent_dir = filepath.parent
        parent_dir.mkdir(parents=True, exist_ok=True)

        # Stat original (if exists) so we can preserve mode/ownership
        orig_mode = None
        orig_uid = None
        orig_gid = None
        if filepath.exists():
            try:
                st = filepath.stat()
                orig_mode = st.st_mode
                orig_uid = st.st_uid
                orig_gid = st.st_gid
            except Exception:
                orig_mode = None

        # Create unique temp file in same directory (important for atomic replace)
        tf = tempfile.NamedTemporaryFile(prefix=f".{filepath.name}.tmp-", dir=str(parent_dir), delete=False)
        temp_path = Path(tf.name)
        tf.close()  # we'll let openpyxl write to the path

        # Create a backup copy (best-effort)
        backup_path = filepath.with_suffix('.backup.xlsm')
        try:
            if filepath.exists():
                shutil.copy2(filepath, backup_path)
        except Exception:
            # Non-fatal; we still proceed
            logger.warning("Could not create backup prior to safe save")

        # Load workbook and apply changes (keep_vba to preserve macros)
        wb = load_workbook(str(filepath), keep_vba=True)
        ws = wb[sheet_name]

        # Helper: parse a single cell coordinate "C12" -> (col_index:int, row:int)
        def _coord_to_col_row(coord: str):
            m = re.match(r'^([A-Za-z]+)(\d+)$', coord)
            if not m:
                return None, None
            col_letters, row_str = m.group(1), m.group(2)
            try:
                col_idx = column_index_from_string(col_letters)
                return col_idx, int(row_str)
            except Exception:
                return None, None

        # Remove/adjust data validations safely (numeric comparisons)
        if ws.data_validations:
            dv_list = list(ws.data_validations.dataValidation)
            for cell_coord in list(values.keys()):
                target_col, target_row = _coord_to_col_row(cell_coord)
                if target_col is None:
                    continue

                for dv in dv_list:
                    if not getattr(dv, "sqref", None):
                        continue

                    raw_ranges = str(dv.sqref).replace('$', '').split()
                    new_ranges = []
                    modified = False

                    for r in raw_ranges:
                        if ':' in r:
                            try:
                                min_col, min_row, max_col, max_row = range_boundaries(r)
                            except Exception:
                                new_ranges.append(r)
                                continue

                            if not (min_col <= target_col <= max_col and min_row <= target_row <= max_row):
                                new_ranges.append(r)
                                continue

                            modified = True

                            # handle vertical single-column
                            if min_col == max_col:
                                if target_row == min_row and target_row == max_row:
                                    continue
                                elif target_row == min_row:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row+1}:{get_column_letter(max_col)}{max_row}")
                                elif target_row == max_row:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row-1}")
                                else:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{target_row-1}")
                                    new_ranges.append(f"{get_column_letter(min_col)}{target_row+1}:{get_column_letter(max_col)}{max_row}")
                            # handle single-row ranges
                            elif min_row == max_row:
                                if target_col == min_col and target_col == max_col:
                                    continue
                                elif target_col == min_col:
                                    new_ranges.append(f"{get_column_letter(min_col+1)}{min_row}:{get_column_letter(max_col)}{max_row}")
                                elif target_col == max_col:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col-1)}{max_row}")
                                else:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(target_col-1)}{max_row}")
                                    new_ranges.append(f"{get_column_letter(target_col+1)}{min_row}:{get_column_letter(max_col)}{max_row}")
                            else:
                                # general rectangle: split vertically into two blocks (simple and safe)
                                if target_row == min_row:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row+1}:{get_column_letter(max_col)}{max_row}")
                                elif target_row == max_row:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row-1}")
                                else:
                                    new_ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{target_row-1}")
                                    new_ranges.append(f"{get_column_letter(min_col)}{target_row+1}:{get_column_letter(max_col)}{max_row}")
                        else:
                            try:
                                col_idx, row_idx = _coord_to_col_row(r)
                                if col_idx is None:
                                    new_ranges.append(r)
                                    continue
                                if col_idx == target_col and row_idx == target_row:
                                    modified = True
                                    continue
                                else:
                                    new_ranges.append(r)
                            except Exception:
                                new_ranges.append(r)

                    if modified:
                        if new_ranges:
                            dv.sqref = ' '.join(new_ranges)
                        else:
                            try:
                                ws.data_validations.dataValidation.remove(dv)
                            except Exception:
                                logger.debug("Unable to remove empty data validation, continuing.")

        # Apply requested cell updates
        for cell, value in values.items():
            ws[cell] = value

        # Save to the temp path
        try:
            wb.save(str(temp_path))
            wb.close()
        except Exception as e:
            logger.error("Failed to save workbook to temp file", exc_info=True)
            # cleanup temp, restore and return failure
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception:
                pass
            return False

        # fsync the temporary file to ensure durability
        try:
            with open(str(temp_path), 'rb') as fh:
                fh.flush()
                os.fsync(fh.fileno())
        except Exception:
            # not fatal — best effort
            logger.debug("fsync of temp file failed or not supported")

        # Verify size (sanity check)
        try:
            if temp_path.stat().st_size <= 1000:
                logger.error("Temporary saved file too small, aborting safe replace")
                # restore from backup if possible
                if backup_path.exists():
                    try:
                        shutil.copy2(backup_path, filepath)
                    except Exception:
                        pass
                try:
                    temp_path.unlink()
                except Exception:
                    pass
                return False
        except Exception:
            pass

        # Attempt atomic replace and preserve mode/ownership
        try:
            os.replace(str(temp_path), str(filepath))  # atomic on same filesystem
            if orig_mode is not None:
                try:
                    os.chmod(str(filepath), orig_mode)
                except Exception:
                    logger.debug("Failed to restore file mode")
            # Try to preserve ownership (best-effort)
            try:
                if orig_uid is not None and orig_gid is not None:
                    os.chown(str(filepath), orig_uid, orig_gid)
            except PermissionError:
                # we may not have privileges to chown; ignore
                logger.debug("Insufficient permissions to chown file; skipping")
            except Exception:
                logger.debug("Failed to chown file (non-fatal)")
        except Exception as e:
            logger.error(f"Atomic replace failed: {e}", exc_info=True)
            # attempt fallback: copy2 and cleanup
            try:
                shutil.copy2(str(temp_path), str(filepath))
                if temp_path.exists():
                    temp_path.unlink()
            except Exception as e2:
                logger.error(f"Fallback copy failed: {e2}", exc_info=True)
                # try to restore original from backup
                try:
                    if backup_path.exists():
                        shutil.copy2(backup_path, filepath)
                except Exception:
                    pass
                return False

        logger.info(f"Excel saved safely: {filepath.stat().st_size} bytes")
        return True

    except Exception as e:
        logger.error(f"Error in save_excel_safely: {e}", exc_info=True)
        # Best-effort restore original
        try:
            backup_path = filepath.with_suffix('.backup.xlsm')
            if backup_path.exists():
                shutil.copy2(backup_path, filepath)
        except Exception:
            pass
        return False


# ========== TEXT PROCESSING ==========

def normalize_text_for_matching(text: str) -> str:
    """Normalize text for matching"""
    if not text:
        return ""

    text = text.lower()
    text_without_parentheses = re.sub(r'\([^)]*\)', ' ', text)
    content_in_parentheses = re.findall(r'\(([^)]+)\)', text)
    all_text = text_without_parentheses + ' ' + ' '.join(content_in_parentheses)
    all_text = re.sub(r'[^\w\s]', ' ', all_text)
    all_text = ' '.join(all_text.split())

    return all_text.strip()

def extract_payment_amount_currency(text: str) -> Tuple[str, Dict[str, Optional[float]], str]:
    """
    Extract payment type, multiple amounts with currencies, and clean subcategory
    Now supports multiple currencies in one input

    Returns: (subcategory, currency_amounts_dict, payment_type)
    where currency_amounts_dict is {'USD': amount, 'LBP': amount, 'EURO': amount}
    """
    if not text:
        return "", {}, "Cash"

    original_text = text
    text_lower = text.lower()

    # 1. Extract payment type
    payment_type = "Cash"

    for keyword in CARD_KEYWORDS:
        if keyword in text_lower:
            payment_type = "Card"
            break

    if payment_type == "Cash":
        for keyword in DIGITAL_KEYWORDS:
            if keyword in text_lower:
                payment_type = "Digital Wallet"
                break

    # 2. RESTORED ORIGINAL MULTI-CURRENCY EXTRACTION WITH FIXES
    currency_amounts = {}  # {'USD': 10.0, 'LBP': -150000.0}

    # Patterns for all currency types - RESTORED FROM ORIGINAL
    patterns = [
        # LBP patterns - UPDATED for better matching
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(lbp|ليرة|lira|ل\.ل)', 'LBP'),
        (r'(lbp|ليرة|lira|ل\.ل)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'LBP'),
        # Simple LBP pattern without commas requirement
        (r'(-?\d+)\s*(lbp|ليرة|lira|ل\.ل)', 'LBP'),

        # USD patterns
        (r'[$\$](-?\d+(?:,\d{3})*(?:\.\d+)?)', 'USD'),
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(usd|dollar)', 'USD'),
        (r'(usd|dollar)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'USD'),
        # Simple USD pattern
        (r'(-?\d+(?:\.\d+)?)\s*\$', 'USD'),

        # EURO patterns
        (r'€(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'EURO'),
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)\s*(eur|euro)', 'EURO'),
        (r'(eur|euro)\s*(-?\d+(?:,\d{3})*(?:\.\d+)?)', 'EURO'),
        # Simple EURO pattern
        (r'(-?\d+(?:\.\d+)?)\s*€', 'EURO'),

        # Symbol suffix patterns
        (r'(-?\d+(?:,\d{3})*(?:\.\d+)?)[€$]', None),
    ]

    # Process each pattern - IMPORTANT: Don't break, collect ALL matches
    for pattern, currency_code in patterns:
        try:
            matches = list(re.finditer(pattern, text_lower, re.IGNORECASE))
            for match in matches:
                try:
                    # Determine amount and currency code robustly
                    if currency_code is None:
                        amount_str = match.group(1).replace(',', '')
                        symbol = match.group(0)[-1]
                        currency_code_detected = 'USD' if symbol == '$' else 'EURO'
                        amount = float(amount_str)
                        currency_code_used = currency_code_detected
                    elif currency_code == 'LBP':
                        amount_str = match.group(1).replace(',', '')
                        amount = float(amount_str)
                        currency_code_used = 'LBP'
                    else:  # USD or EURO
                        amount_str = match.group(1).replace(',', '')
                        amount = float(amount_str)
                        currency_code_used = currency_code

                    # Check if amount starts with - in matched text
                    match_text = match.group(0)
                    if match_text.strip().startswith('-') and amount > 0:
                        amount = -amount

                    # Store the amount for this currency (last occurrence wins)
                    currency_amounts[currency_code_used] = amount

                    logger.info(f"Found amount: {amount} {currency_code_used}")

                except (ValueError, IndexError) as e:
                    logger.debug(f"Error parsing amount: {e}")
                    continue

        except re.error:
            continue

    # 3. Also look for standalone negative numbers near currency words
    if 'LBP' not in currency_amounts:
        lbp_pattern = r'-\s*(\d+(?:,\d{3})*)\s*(lbp|ليرة|lira|ل\.ل)'
        matches = re.findall(lbp_pattern, text_lower, re.IGNORECASE)
        for match in matches:
            try:
                amount = -float(match[0].replace(',', ''))
                currency_amounts['LBP'] = amount
                logger.info(f"Found negative LBP amount: {amount}")
            except ValueError:
                pass

    # 4. If we found some but not all currencies, look for standalone amounts
    if currency_amounts:
        # Look for standalone numbers that aren't already part of currency patterns
        standalone_pattern = r'(?<!\S)(-?\d+(?:,\d{3})*(?:\.\d+)?)(?![\$\€\w])'
        matches = re.finditer(standalone_pattern, text_lower)

        for match in matches:
            try:
                amount = float(match.group(1).replace(',', ''))
                # If it's a large whole number and no LBP yet, assume LBP
                if amount >= 1000 and float(amount).is_integer() and 'LBP' not in currency_amounts:
                    currency_amounts['LBP'] = amount
                    logger.info(f"Assumed LBP for standalone: {amount}")
                # If it has decimals and no USD yet, assume USD
                elif not float(amount).is_integer() and 'USD' not in currency_amounts:
                    currency_amounts['USD'] = amount
                    logger.info(f"Assumed USD for standalone: {amount}")
                # Small whole number and we need USD
                elif amount < 1000 and float(amount).is_integer() and 'USD' not in currency_amounts:
                    currency_amounts['USD'] = amount
                    logger.info(f"Assumed USD for small whole number: {amount}")
            except ValueError:
                continue

    # 5. Clean subcategory - remove all amount patterns
    subcategory = original_text

    # Remove all detected amounts
    for currency_code, amount in currency_amounts.items():
        abs_amount = abs(amount)

        if currency_code == 'LBP':
            # Remove LBP formats
            formats_to_remove = [
                f"{int(abs_amount):,}LBP", f"{int(abs_amount)}LBP",
                f"LBP{int(abs_amount):,}", f"LBP{int(abs_amount)}",
                f"{int(abs_amount):,} lbp", f"{int(abs_amount)} lbp",
                f"lbp {int(abs_amount):,}", f"lbp {int(abs_amount)}",
                f"{int(abs_amount):,}ليرة", f"{int(abs_amount)}ليرة",
                f"ليرة{int(abs_amount):,}", f"ليرة{int(abs_amount)}",
                f"{int(abs_amount):,} ليرة", f"{int(abs_amount)} ليرة",
                f"ليرة {int(abs_amount):,}", f"ليرة {int(abs_amount)}",
            ]
        elif currency_code == 'USD':
            # Remove USD formats
            formats_to_remove = [
                f"${abs_amount:.2f}", f"${abs_amount:,.2f}",
                f"{abs_amount:.2f}$", f"{abs_amount:,.2f}$",
                f"{abs_amount:.2f}usd", f"{abs_amount:.2f} dollar",
                f"{abs_amount}usd", f"{abs_amount} dollar",
            ]
        elif currency_code == 'EURO':
            # Remove EURO formats
            formats_to_remove = [
                f"€{abs_amount:.2f}", f"€{abs_amount:,.2f}",
                f"{abs_amount:.2f}€", f"{abs_amount:,.2f}€",
                f"{abs_amount:.2f}eur", f"{abs_amount:.2f} euro",
                f"{abs_amount}eur", f"{abs_amount} euro",
            ]
        else:
            formats_to_remove = []

        # Add negative versions
        negative_formats = [f"-{fmt}" for fmt in formats_to_remove]
        formats_to_remove.extend(negative_formats)

        # Remove all formats (case insensitive)
        for fmt in formats_to_remove:
            try:
                subcategory = re.sub(re.escape(fmt), ' ', subcategory, flags=re.IGNORECASE)
            except Exception:
                subcategory = subcategory.replace(fmt, ' ')
            # Also try with spaces removed
            subcategory = subcategory.replace(fmt.replace(' ', ''), ' ')

        # Also remove just the number (with and without commas)
        num_str = str(abs_amount)
        if '.' in num_str:
            # Has decimals
            subcategory = subcategory.replace(num_str, ' ')
        else:
            # Whole number
            subcategory = subcategory.replace(num_str, ' ')
            if abs_amount >= 1000:
                # Remove with commas
                subcategory = subcategory.replace(f"{int(abs_amount):,}", ' ')
                # Remove without commas
                subcategory = subcategory.replace(f"{int(abs_amount)}", ' ')

    # Remove standalone currency symbols and words
    subcategory = subcategory.replace('$', ' ').replace('€', ' ')

    # Remove payment keywords
    for keyword in CARD_KEYWORDS + DIGITAL_KEYWORDS:
        subcategory = re.sub(re.escape(keyword), ' ', subcategory, flags=re.IGNORECASE)

    # Remove currency words
    currency_words = ['usd', 'dollar', 'euro', 'eur', 'lbp', 'lira', 'ليرة', 'ل.ل']
    for word in currency_words:
        subcategory = re.sub(re.escape(word), ' ', subcategory, flags=re.IGNORECASE)

    # Clean up
    subcategory = ' '.join(subcategory.split())
    subcategory = subcategory.title()

    if not subcategory:
        subcategory = original_text

    logger.info(f"Parsed: subcategory='{subcategory}', amounts={currency_amounts}, payment={payment_type}")
    return subcategory, currency_amounts, payment_type

def find_best_match_for_input(input_text: str, tables_dict: Dict) -> Tuple[str, str, str, float]:
    """Find the best match for input text with confidence score"""
    if not input_text:
        return "", "Unknown", "Empty input", 0.0

    normalized_input = normalize_text_for_matching(input_text)

    if not normalized_input:
        return "", "Unknown", "Empty after normalization", 0.0

    # Reject very short inputs
    if len(normalized_input) < 3:
        return "", "Unknown", "Input too short", 0.0

    # Reject common non-category words appearing anywhere in the input (token-level)
    input_tokens = set(normalized_input.split())
    common_set = set(COMMON_NON_CATEGORIES)
    if input_tokens & common_set:
        return "", "Unknown", "Common non-category word", 0.0

    best_match = None
    best_category = None
    best_score = 0.0
    match_type = "No match"

    # SIMPLIFY: Just split by spaces, no complex regex
    input_words = set(normalized_input.split())

    for category, data in tables_dict.items():
        for variation, original in data['variations'].items():
            # SIMPLIFY: Split variation by spaces
            variation_words = set(variation.split())

            # Calculate intersection
            intersection = len(input_words.intersection(variation_words))
            union = len(input_words.union(variation_words))

            if union > 0:
                score = intersection / union

                # Boost for exact matches
                if normalized_input == variation:
                    score = 1.0
                elif normalized_input in variation:
                    score = max(score, 0.8)
                elif variation in normalized_input:
                    score = max(score, 0.8)

                if score > best_score:
                    best_score = score
                    best_match = original
                    best_category = category

    # Check if we have a good enough match
    if best_score >= MINIMUM_CONFIDENCE:
        if best_score >= 0.95:
            match_type = "Exact match"
        elif best_score >= 0.85:
            match_type = "Strong match"
        elif best_score >= MINIMUM_CONFIDENCE:
            match_type = "Good match"

        logger.info(f"Good match found: '{input_text}' → '{best_match}' (score: {best_score:.2f})")
        return best_match, best_category, match_type, best_score

    # Try fuzzy matching with simplified approach
    if best_score < FUZZY_MATCH_THRESHOLD:
        # Build simple lists for fuzzy matching
        all_variations = []
        all_originals = []
        all_categories = []

        for category, data in tables_dict.items():
            for variation, original in data['variations'].items():
                all_variations.append(variation)
                all_originals.append(original)
                all_categories.append(category)

        # Use simple fuzzy matching
        matches = get_close_matches(normalized_input, all_variations, n=1, cutoff=0.5)

        if matches:
            best_variation = matches[0]
            idx = all_variations.index(best_variation)
            fuzzy_score = 0.5  # Fixed score for fuzzy matches
            if fuzzy_score >= MINIMUM_CONFIDENCE:
                return all_originals[idx], all_categories[idx], "Fuzzy match", fuzzy_score

    logger.info(f"No good match found for '{input_text}'. Best score: {best_score:.2f}")
    return "", "Unknown", "No good match found", best_score

# ========== TELEGRAM BOT FUNCTIONS ==========

async def start_command(update, context):
    """Handle /start - UPDATED with enhanced delete feature"""
    logger.info(f"Received /start from user {update.effective_user.id}")

    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    excel_status = "✅ INSTALLED" if EXCEL_AVAILABLE else "❌ NOT INSTALLED"
    onedrive_status = "✅ INSTALLED" if ONEDRIVE_AVAILABLE else "❌ NOT INSTALLED"

    await update.message.reply_text(
        f"💰 <b>Smart Budget Tracker Bot</b>\n\n"
        f"📊 <b>Status:</b>\n"
        f"• Excel: {excel_status}\n"
        f"• OneDrive: {onedrive_status}\n"
        f"🎯 <b>Minimum Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}&#37;\n\n"

        "✅ <b>Smart Matching Examples:</b>\n"
        "• <code>Chamsin</code> → Bakery Products (Chamsin)\n"
        "• <code>Mazda</code> → Fuel (Mazda)\n"
        "• <code>KSC</code> → KSC (exact match)\n"
        "• <code>Fuel Opel</code> → Fuel (Opel)\n\n"

        "🔄 <b>Negative Amounts (Refunds/Corrections):</b>\n"
        "• <code>Chamsin -10$</code> → Refund from Chamsin\n"
        "• <code>Refund -200000 LBP</code> → Refund (no commas needed)\n"
        "• <code>Correction -50€</code> → Negative adjustment\n\n"

        "💱 <b>LBP Amounts:</b>\n"
        "• <code>200000 LBP</code> → Works without commas\n"
        "• <code>200,000 LBP</code> → Also works with commas\n\n"

        "❌ <b>Will Reject:</b>\n"
        f"• <code>OKAY</code>, <code>Test</code>, <code>Hello</code>\n"
        f"• Poor matches (&lt;{MINIMUM_CONFIDENCE*100:.0f}&#37; confidence)\n\n"

        "<b>📝 How to Use:</b>\n"
        "Just send: <code>ITEM AMOUNT CURRENCY PAYMENT</code>\n\n"
        "<b>Examples:</b>\n"
        "• <code>Chamsin 10</code> (Expense)\n"
        "• <code>Chamsin -10$</code> (Refund)\n"
        "• <code>Fuel Mazda 200000 LBP</code> (No commas needed)\n"
        "• <code>Refund -500$ card</code> (Negative amount)\n"
        "• <code>KSC 15.50 Card</code>\n"
        "• <code>Daouk Sweets 20$ Card</code>\n\n"

        "<b>🔐 OneDrive Setup:</b>\n"
        "1. <code>/onedrive_auth</code> - Get authentication URL\n"
        "2. <code>/onedrive_test</code> - Test connection\n\n"

        "<b>📥 DOWNLOAD FEATURES:</b>\n"
        "• /download csv [day/week/month/year/all] - Export to CSV\n"
        "• /download summary - Export summary JSON\n"
        "• /download backup - Download Excel backup\n\n"

        "<b>⚙️ Commands:</b>\n"
        "/start - This message\n"
        "/help - Detailed help\n"
        "/testparse [text] - Test parsing & matching\n"
        "/recent - Last 10 transactions\n"
        "/stats - Statistics\n"
        "/save - Force save to OneDrive\n"
        "/clearcache - Clear matching cache\n"
        "/download - Download data files\n"
        "/onedrive_auth - Setup OneDrive authentication\n"
        "/onedrive_test - Test OneDrive connection\n\n"

        "💡 <b>Tip:</b> Tap the menu icon (/) to see all commands",
        parse_mode='HTML'
    )

async def help_command(update, context):
    """Extended help command"""
    logger.info(f"Received /help from user {update.effective_user.id}")

    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    await update.message.reply_text(
        f"💰 <b>Smart Budget Tracker Bot - Help Guide</b>\n\n"

        f"<b>🎯 Smart Matching System:</b>\n"
        f"• Minimum confidence required: {MINIMUM_CONFIDENCE*100:.0f}%\n"
        f"• Rejects poor matches with suggestions\n"
        f"• Uses fuzzy matching for close matches\n\n"

        "<b>🔄 Negative Amount Support:</b>\n"
        "Use negative amounts for:\n"
        "• Refunds\n"
        "• Corrections\n"
        "• Negative adjustments\n"
        "• Transaction type remains 'Expenses'\n\n"

        "<b>💱 LBP Amount Format:</b>\n"
        "• <code>200000 LBP</code> - No commas needed\n"
        "• <code>200,000 LBP</code> - Commas also work\n"
        "• <code>-150000 lbp</code> - Negative also works\n\n"

        "<b>🌐 Multi-Currency Support:</b>\n"
        "• <code>KSC 10$, -150000 LBP</code> - Multiple currencies in one transaction\n"
        "• <code>Grocery 20€ 150000 LBP</code> - EURO + LBP\n"
        "• <code>Refund -5$, -75000 LBP</code> - Multiple refunds\n\n"

        "<b>📝 Basic Usage:</b>\n"
        "Simply send a message in this format:\n"
        "<code>ITEM [AMOUNT] [CURRENCY] [PAYMENT]</code>\n\n"

        "<b>✅ Will Accept (Examples):</b>\n"
        "• <code>Chamsin 10</code> - Expense\n"
        "• <code>Chamsin -10$</code> - Refund (still 'Expenses' type)\n"
        "• <code>Fuel Mazda 200000 LBP</code> - No commas needed\n"
        "• <code>Refund -500$ card</code> - Negative amount\n"
        "• <code>KSC 15.50 Card</code> - Expense\n"
        "• <code>Correction -200€</code> - Negative adjustment\n"
        "• <code>KSC 10$, -150000 LBP</code> - Multi-currency\n\n"

        "<b>❌ Will Reject:</b>\n"
        "• <code>OKAY 10$</code> - No matching category\n"
        "• <code>Test -5</code> - Common word\n"
        "• <code>Hello 20</code> - Greeting\n\n"

        "<b>🔐 OneDrive Setup:</b>\n"
        "1. Register app at: https://portal.azure.com\n"
        "2. Get Client ID and Secret\n"
        "3. Run <code>/onedrive_auth</code>\n"
        "4. Follow authentication steps\n\n"

        "<b>📥 DOWNLOAD COMMANDS:</b>\n"
        "• <code>/download csv day</code> - Today's transactions\n"
        "• <code>/download csv week</code> - This week's transactions\n"
        "• <code>/download csv month</code> - This month's transactions\n"
        "• <code>/download csv year</code> - This year's transactions\n"
        "• <code>/download csv all</code> - All transactions\n"
        "• <code>/download summary</code> - Summary statistics (JSON)\n"
        "• <code>/download backup</code> - Excel backup file\n\n"

        "<b>⚙️ Available Commands:</b>\n"
        "/start - Welcome message\n"
        "/help - This help message\n"
        "/testparse [text] - Test parsing & matching\n"
        "/recent - Show last 10 transactions\n"
        "/stats - Show statistics\n"
        "/save - Force save to OneDrive\n"
        "/clearcache - Clear matching cache\n"
        "/download - Download data files\n"
        "/onedrive_auth - OneDrive authentication\n"
        "/onedrive_test - Test OneDrive connection\n\n"

        "<b>💡 Tips:</b>\n"
        "• LBP amounts work with or without commas\n"
        "• Negative amounts keep 'Expenses' type\n"
        "• Be specific with item names\n"
        "• Check spelling if not matching\n"
        "• Use /testparse to test before adding\n"
        "• Download data regularly for backup",
        parse_mode='HTML'
    )

async def testparse_command(update, context):
    """Test input parsing and matching - UPDATED for multi-currency"""
    logger.info(f"Received /testparse from user {update.effective_user.id}")

    if not is_authorized(update.effective_user.id):
        return

    if not context.args:
        await update.message.reply_text(
            "Usage: /testparse [text]\n\n"
            "Examples (Multi-Currency):\n"
            "• /testparse KSC 10$, -150000 LBP\n"
            "• /testparse Grocery 20€ 150000 LBP\n"
            "• /testparse Refund -5$, -75000 LBP\n\n"
            "Examples (Single Currency):\n"
            "• /testparse Chamsin 10$\n"
            "• /testparse Chamsin -10$\n"
            "• /testparse Fuel Mazda 200000 LBP"
        )
        return

    test_input = ' '.join(context.args)

    # Parse the input with multi-currency support
    subcategory, currency_amounts, payment_type = extract_payment_amount_currency(test_input)

    # Load tables and find match
    tables_dict = load_all_tables_with_details()
    matched, category, match_type, confidence = find_best_match_for_input(subcategory, tables_dict)

    # FIXED: Convert confidence to percentage
    confidence_percentage = confidence * 100

    # Determine if it would be accepted
    would_accept = confidence >= MINIMUM_CONFIDENCE
    accept_status = "✅ WOULD BE ACCEPTED" if would_accept else "❌ WOULD BE REJECTED"

    # Format currency amounts
    amounts_str = ""
    if currency_amounts:
        amount_parts = []
        for currency, amount in currency_amounts.items():
            if amount is not None:
                amount_parts.append(f"{format_currency_amount(amount, currency)}")
        amounts_str = ", ".join(amount_parts)
    else:
        amounts_str = "None"

    # Format the results
    result = [
        f"🔍 <b>Parsing & Matching Test (Multi-Currency)</b>",
        f"",
        f"<b>📝 Input:</b> <code>{test_input}</code>",
        f"",
        f"<b>📊 Parsing Results:</b>",
        f"• Subcategory: <code>{subcategory}</code>",
        f"• Amount{'s' if len(currency_amounts) > 1 else ''}: {amounts_str}",
        f"• Currencies Detected: {len(currency_amounts)}",
        f"• Payment Type: {payment_type}",
        f"",
        f"<b>🎯 Matching Results:</b>",
        f"• Matched to: <code>{matched if matched else 'No match'}</code>",
        f"• Category: {category}",
        f"• Match Type: {match_type}",
        f"• Confidence: {confidence_percentage:.1f}%",
        f"• Minimum Required: {MINIMUM_CONFIDENCE*100:.0f}%",
        f"",
        f"<b>📋 Final Result:</b> {accept_status}"
    ]

    await update.message.reply_text("\n".join(result), parse_mode='HTML')

# Remaining handlers (recent, handle_message, stats, save, clear_cache, unlock, repair, download,
# delete/modify flows, etc.) remain unchanged except for removing duplicate registrations.
# For brevity in this response, the rest of the file (handlers and main) is kept the same logic
# as previously provided but with duplicates removed and the fixes applied above (locks & consolidations).

# ========== TELEGRAM HANDLERS (rest of implementations follow as before) ==========
# Due to message size, the rest of the functions are unchanged in behavior from the original file,
# except:
# - Duplicate ONEDRIVE_* definitions removed
# - _onedrive_upload_lock/_upload_in_progress moved earlier
# - Stray global time.sleep removed
# - Duplicate command registration removed in main()


async def recent_command(update, context):
    """Show recent transactions - non-blocking"""
    logger.info(f"Received /recent from user {update.effective_user.id}")

    if not is_authorized(update.effective_user.id):
        return

    if not EXCEL_AVAILABLE:
        await update.message.reply_text("❌ openpyxl not installed. Cannot read Excel file.")
        return

    await update.message.reply_text("📋 Loading last 10 transactions...")

    try:
        # get_recent_transactions does copy_excel_from_onedrive and reads the workbook — run it in a thread
        recent_transactions = await run_blocking(get_recent_transactions, 10)

        if not recent_transactions:
            await update.message.reply_text("No recent transactions found.")
            return

        response_lines = [f"📋 <b>Recent Transactions (Last {len(recent_transactions)}):</b>", ""]
        for t in recent_transactions:
            date_val = t['date']
            date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)

            amounts = []
            if t.get('usd') is not None:
                prefix = "🔄 " if t['usd'] < 0 else ""
                amounts.append(f"{prefix}${abs(t['usd']):.2f}")
            if t.get('lbp') is not None:
                prefix = "🔄 " if t['lbp'] < 0 else ""
                amounts.append(f"{prefix}{abs(t['lbp']):,.0f} LBP")
            if t.get('euro') is not None:
                prefix = "🔄 " if t['euro'] < 0 else ""
                amounts.append(f"{prefix}€{abs(t['euro']):.2f}")

            amount_str = " + ".join(amounts) if amounts else "No amount"

            payment_emoji = "💳" if str(t.get('payment','')).lower() == "card" else "💵"
            item = t.get('subcategory') if t.get('subcategory') else t.get('category', '')
            import html
            safe_item = html.escape(item) if item else "Unknown"

            response_lines.append(f"• {payment_emoji} {date_str}: <code>{safe_item}</code> - {amount_str}")

        await update.message.reply_text("\n".join(response_lines), parse_mode='HTML')

    except Exception as e:
        logger.exception("Error in recent_command")
        await update.message.reply_text(f"❌ Error loading transactions: {str(e)[:200]}")

async def run_blocking(func, *args, use_lock: bool = False, **kwargs):
    """
    Run a blocking function in a background thread and return its result.
    If use_lock=True the wrapper will acquire excel_lock inside the worker thread
    before calling the function.

    - func: blocking callable
    - use_lock: if True acquire excel_lock inside the worker thread
    """
    if use_lock:
        def _wrapped():
            with excel_lock:
                return func(*args, **kwargs)
        return await asyncio.to_thread(_wrapped)
    else:
        return await asyncio.to_thread(func, *args, **kwargs)

async def handle_message(update, context):
    """Handle regular messages - non-blocking by offloading heavy work to thread"""
    user_id = update.effective_user.id
    text = (update.message.text or "").strip()

    logger.info(f"Received message from user {user_id}: {text}")

    if not is_authorized(user_id):
        logger.warning(f"User {user_id} not authorized")
        await update.message.reply_text("⛔ Unauthorized.")
        return

    if text.startswith('/'):
        return

    if not text:
        await update.message.reply_text(
            "💰 <b>How to Add a Transaction (Multi-Currency):</b>\n\n"
            "Send: <code>ITEM [AMOUNTS WITH CURRENCIES] [PAYMENT]</code>\n\n"
            "<b>📝 Examples (Single Currency):</b>\n"
            "• <code>Chamsin 10</code> (Expense)\n"
            "• <code>Chamsin -10$</code> (Refund)\n"
            "• <code>Fuel 200000 lbp</code> (No commas needed)\n\n"
            "<b>🌐 Examples (Multi-Currency):</b>\n"
            "• <code>KSC 10$, -150000 LBP</code>\n"
            "• <code>Grocery 20€ 150000 LBP</code>\n"
            "• <code>Refund -5$, -75000 LBP</code>\n\n"
            "<b>📝 Optional Notes:</b>\n"
            "Add a second line starting with 'DETAILS : ' for notes\n"
            "• <code>ITEM AMOUNTS\nDETAILS : your notes</code>\n\n"
            f"<b>🎯 Minimum Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}%\n"
            "Poor matches will be rejected with suggestions.\n\n"
            "Need help? Use /help for detailed instructions.",
            parse_mode='HTML'
        )
        return

    # Extract optional notes on a new line (DETAILS :)
    transaction_text = text
    optional_notes = ""
    details_pattern = r'\n\s*DETAILS\s*:\s*(.+)'
    details_match = re.search(details_pattern, text, re.IGNORECASE | re.DOTALL)
    if details_match:
        transaction_text = text[:details_match.start()].strip()
        optional_notes = details_match.group(1).strip()
        logger.info(f"Found optional notes: '{optional_notes}'")

    # Parse (fast, non-blocking)
    subcategory, currency_amounts, payment_type = extract_payment_amount_currency(transaction_text)

    if not currency_amounts:
        await update.message.reply_text(
            f"⚠️ <b>Amount Not Detected</b>\n\n"
            "Could not detect any amounts with currencies.\n\n"
            "Examples:\n"
            "• <code>KSC 10$, -150000 LBP</code>\n"
            "• <code>Grocery 20€ 150000 LBP</code>\n\n"
            "<b>Optional notes:</b>\n"
            "<code>ITEM AMOUNTS\nDETAILS : your notes here</code>",
            parse_mode='HTML'
        )
        return

    if not subcategory:
        await update.message.reply_text(
            "❌ <b>Parsing Error</b>\n\n"
            "Could not extract item from message.\n"
            "Try: <code>Item [Amounts with Currencies] [Payment]</code>\nUse /help for examples.",
            parse_mode='HTML'
        )
        return

    # Quick preview to user (non-blocking formatting)
    amount_display_parts = []
    has_negative = False
    for currency, amount in currency_amounts.items():
        if amount is not None:
            amount_str = format_currency_amount(amount, currency)
            if amount < 0:
                amount_display_parts.append(f"🔄 {amount_str}")
                has_negative = True
            else:
                amount_display_parts.append(amount_str)

    amount_display = " + ".join(amount_display_parts) if amount_display_parts else "No amount specified"
    transaction_desc = "Expense (with refunds/corrections)" if has_negative else "Expense"
    multi_currency_indicator = "🌐 " if len(currency_amounts) > 1 else ""
    import html
    notes_preview = f"\n📝 <b>Notes:</b> {html.escape(optional_notes[:50])}{'...' if len(optional_notes) > 50 else ''}" if optional_notes else ""

    await update.message.reply_text(
        f"{multi_currency_indicator}💰 <b>Processing {transaction_desc}:</b>\n"
        f"Item: <code>{html.escape(subcategory)}</code>\n"
        f"Amount{'s' if len(currency_amounts) > 1 else ''}: {amount_display}\n"
        f"Payment: {payment_type}"
        f"{notes_preview}\n\n"
        f"⏳ Matching and adding...",
        parse_mode='HTML'
    )

    # Offload blocking add_transaction_smart (it uses excel_lock internally) to a worker thread
    try:
        success, message = await run_blocking(add_transaction_smart, subcategory, currency_amounts, payment_type, optional_notes, use_lock=False)
        await update.message.reply_text(message, parse_mode='HTML')
    except Exception as e:
        logger.exception("Error adding transaction")
        await update.message.reply_text(f"❌ Error adding transaction: {str(e)[:200]}")

async def stats_command(update, context):
    """Show statistics"""
    logger.info(f"Received /stats from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    try:
        success, msg = copy_excel_from_onedrive()
        if not success:
            await update.message.reply_text(msg)
            return
        
        total_transactions = 0
        if EXCEL_AVAILABLE:
            with excel_lock:
                wb = load_workbook(str(LOCAL_COPY_PATH), data_only=True)
                sheet = wb[TRACKING_SHEET_NAME]
                
                # Count transactions
                row = 12
                while sheet[f'C{row}'].value not in [None, ""]:
                    total_transactions += 1
                    row += 1
                    if row > 1000:
                        break
                
                wb.close()
        else:
            total_transactions = "N/A (openpyxl not installed)"
        
        # Get table count from cache
        table_count = len(_table_cache) if _table_cache else 0
        total_items = sum(data['count'] for data in _table_cache.values()) if _table_cache else 0
        
        # Get last backup time
        def get_last_backup_time() -> str:
            try:
                backups = list(BACKUP_DIR.glob("backup_*"))
                if not backups:
                    return "No backups yet"
                
                latest = max(backups, key=lambda x: x.stat().st_mtime)
                mod_time = datetime.fromtimestamp(latest.stat().st_mtime)
                return mod_time.strftime("%Y-%m-%d %H:%M")
            except:
                return "Error checking backups"
        
        # Check export directory
        export_files = list(EXPORT_DIR.glob("*")) if EXPORT_DIR.exists() else []
        
        await update.message.reply_text(
            f"📊 <b>Budget Tracker Statistics</b>\n\n"
            f"• <b>Excel Integration:</b> {'✅ Installed' if EXCEL_AVAILABLE else '❌ Not installed'}\n"
            f"• <b>OneDrive Integration:</b> {'✅ Installed' if ONEDRIVE_AVAILABLE else '❌ Not installed'}\n"
            f"• <b>Total Transactions:</b> {total_transactions}\n"
            f"• <b>Categories Loaded:</b> {table_count}\n"
            f"• <b>Items in Database:</b> {total_items}\n"
            f"• <b>Cache Status:</b> {'✅ Loaded' if _table_cache else '⚠️ Empty'}\n"
            f"• <b>Last Backup:</b> {get_last_backup_time()}\n"
            f"• <b>Exports Available:</b> {len(export_files)} files\n"
            f"• <b>Min Confidence:</b> {MINIMUM_CONFIDENCE*100:.0f}%\n"
            f"• <b>Negative Amounts:</b> ✅ Supported\n"
            f"• <b>LBP without commas:</b> ✅ Supported\n"
            f"• <b>Bot Status:</b> ✅ Running",
            parse_mode='HTML'
        )
        
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)[:200]}")

async def save_command(update, context):
    """Force save to OneDrive (non-blocking)"""
    logger.info(f"Received /save from user {update.effective_user.id}")

    if not is_authorized(update.effective_user.id):
        return

    await update.message.reply_text("💾 Saving to OneDrive...")

    try:
        success, message = await run_blocking(save_excel_to_onedrive)
        await update.message.reply_text(message)
    except Exception as e:
        logger.exception("Error in save_command")
        await update.message.reply_text(f"❌ Save error: {str(e)[:200]}")


async def force_sync_command(update, context):
    """Force immediate sync with OneDrive (non-blocking)"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    await update.message.reply_text("🔄 Forcing immediate sync with OneDrive...")

    try:
        success, msg = await run_blocking(save_excel_to_onedrive)
        if success:
            await update.message.reply_text(f"✅ {msg}")
            return

        # Try a second attempt after a short non-blocking wait
        await asyncio.sleep(2)
        success2, msg2 = await run_blocking(save_excel_to_onedrive)
        if success2:
            await update.message.reply_text("✅ Sync succeeded on second attempt!")
        else:
            await update.message.reply_text(
                f"❌ Sync failed twice:\n\n"
                f"First attempt: {msg}\n\n"
                f"Second attempt: {msg2}\n\n"
                f"Background sync will continue trying."
            )

    except Exception as e:
        logger.exception("Error in force_sync_command")
        await update.message.reply_text(f"❌ Sync error: {str(e)[:200]}")


async def manual_upload_command(update, context):
    """Manual upload command (non-blocking)"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    await update.message.reply_text("🔄 Manual upload started...")

    strategies = [
        ("Direct upload", simple_direct_upload),
        ("Temp name strategy", upload_to_different_name_then_move),
        ("Unique name", force_upload_with_unique_name)
    ]

    results = []
    try:
        for name, func in strategies:
            await update.message.reply_text(f"Trying {name}...")
            # run each blocking upload strategy in a worker thread
            success, msg = await run_blocking(func)
            results.append(f"{'✅' if success else '❌'} {name}: {msg}")
            if success:
                break
            # small non-blocking pause between attempts
            await asyncio.sleep(2)

        result_text = "\n".join(results)
        await update.message.reply_text(f"<b>Manual Upload Results:</b>\n\n{result_text}", parse_mode='HTML')

    except Exception as e:
        logger.exception("Error in manual_upload_command")
        await update.message.reply_text(f"❌ Manual upload error: {str(e)[:200]}")

async def clear_cache_command(update, context):
    """Clear the table cache"""
    logger.info(f"Received /clearcache from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    global _table_cache, _table_cache_timestamp
    _table_cache = {}
    _table_cache_timestamp = None
    
    await update.message.reply_text("✅ Cache cleared! Will reload on next transaction.")

async def unlock_command(update, context):
    """Emergency unlock Excel file - Linux version"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    await update.message.reply_text("🔓 Attempting to unlock Excel file...")
    
    try:
        # Kill any LibreOffice processes
        subprocess.run(['pkill', '-f', 'libreoffice'], capture_output=True, text=True)
        subprocess.run(['pkill', '-f', 'soffice'], capture_output=True, text=True)
        
        # Try to delete locked files
        locked = False
        if os.path.exists(LOCAL_COPY_PATH):
            try:
                os.remove(LOCAL_COPY_PATH)
                msg = "✅ Locked file removed. Bot will create new copy."
            except PermissionError:
                locked = True
                msg = "⚠️ File still locked. Try: sudo lsof | grep temp_budget.xlsm"
        
        if locked:
            msg += "\n\n💡 **Solutions:**"
            msg += "\n1. Check for open processes: <code>sudo lsof | grep xlsm</code>"
            msg += "\n2. Kill process: <code>sudo kill -9 PID</code>"
            msg += "\n3. Wait 1 minute and try again"
        
        await update.message.reply_text(msg)
        
    except Exception as e:
        await update.message.reply_text(f"❌ Unlock error: {str(e)[:200]}")

async def repair_excel_command(update, context):
    """Repair a corrupted Excel file"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    await update.message.reply_text("🔧 Repairing Excel file...")
    
    # Download fresh from OneDrive
    success, msg = copy_excel_from_onedrive()
    if not success:
        await update.message.reply_text(f"❌ {msg}")
        return
    
    # Create backup
    backup_path = LOCAL_COPY_PATH.with_suffix('.backup.xlsm')
    shutil.copy2(LOCAL_COPY_PATH, backup_path)
    
    try:
        # Try to fix the file
        fixed = fix_excel_file_format(LOCAL_COPY_PATH, LOCAL_COPY_PATH)
        
        if fixed:
            # Test if it works
            try:
                wb = load_workbook(str(LOCAL_COPY_PATH), read_only=True, data_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                
                await update.message.reply_text(
                    f"✅ <b>Excel file repaired!</b>\n\n"
                    f"• Sheets: {sheet_count}\n"
                    f"• Size: {LOCAL_COPY_PATH.stat().st_size:,} bytes\n"
                    f"• Backup saved: {backup_path.name}\n\n"
                    f"Try /sync to upload to OneDrive.",
                    parse_mode='HTML'
                )
            except Exception as e:
                await update.message.reply_text(
                    f"⚠️ <b>Partial repair</b>\n\n"
                    f"File was modified but may still have issues.\n"
                    f"Error: {str(e)[:200]}",
                    parse_mode='HTML'
                )
        else:
            await update.message.reply_text("❌ Could not repair the file")
            
    except Exception as e:
        await update.message.reply_text(f"❌ Repair error: {str(e)[:200]}")

# ========== DOWNLOAD COMMAND ==========

async def download_command(update, context):
    """Handle download requests"""
    logger.info(f"Received /download from user {update.effective_user.id}")
    
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "📥 <b>Download Options:</b>\n\n"
            "<b>CSV Exports:</b>\n"
            "• <code>/download csv day</code> - Today's transactions\n"
            "• <code>/download csv week</code> - This week's transactions\n"
            "• <code>/download csv month</code> - This month's transactions\n"
            "• <code>/download csv year</code> - This year's transactions\n"
            "• <code>/download csv all</code> - All transactions\n\n"
            "<b>Other Downloads:</b>\n"
            "• <code>/download summary</code> - Summary statistics (JSON)\n"
            "• <code>/download backup</code> - Excel backup file\n\n"
            "<b>Examples:</b>\n"
            "• <code>/download csv month</code>\n"
            "• <code>/download summary</code>\n"
            "• <code>/download backup</code>",
            parse_mode='HTML'
        )
        return
    
    command = context.args[0].lower()
    
    if command == "csv":
        if len(context.args) < 2:
            await update.message.reply_text(
                "Usage: /download csv [day/week/month/year/all]\n\n"
                "Examples:\n"
                "• /download csv day\n"
                "• /download csv month\n"
                "• /download csv all"
            )
            return
        
        time_range = context.args[1].lower()
        if time_range not in ["day", "week", "month", "year", "all"]:
            await update.message.reply_text(
                "❌ Invalid time range. Use: day, week, month, year, or all"
            )
            return
        
        await update.message.reply_text(f"📊 Exporting {time_range} transactions to CSV...")
        
        # Export to CSV
        success, result = export_to_csv(time_range)
        
        if success:
            try:
                # Send the file
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"transactions_{time_range}_{datetime.now().strftime('%Y%m%d')}.csv",
                        caption=f"📊 Transactions ({time_range}) - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Clean up the file after sending
                os.remove(result)
            except Exception as e:
                logger.error(f"Error sending CSV file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    elif command == "summary":
        await update.message.reply_text("📈 Exporting summary statistics...")
        
        # Export summary
        success, result = export_summary()
        
        if success:
            try:
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"budget_summary_{datetime.now().strftime('%Y%m%d')}.json",
                        caption=f"📈 Budget Summary - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Clean up
                os.remove(result)
            except Exception as e:
                logger.error(f"Error sending summary file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    elif command == "backup":
        await update.message.reply_text("💾 Creating Excel backup...")
        
        # Create backup copy
        success, result = create_backup_copy()
        
        if success:
            try:
                with open(result, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=file,
                        filename=f"budget_backup_{datetime.now().strftime('%Y%m%d')}.xlsm",
                        caption=f"💾 Excel Backup - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                # Note: Don't delete backup files from backup directory
            except Exception as e:
                logger.error(f"Error sending backup file: {e}")
                await update.message.reply_text(f"❌ Error sending file: {str(e)[:200]}")
        else:
            await update.message.reply_text(f"❌ {result}")
    
    else:
        await update.message.reply_text(
            "❌ Unknown download type.\n"
            "Use: csv, summary, or backup\n\n"
            "Example: /download csv month"
        )

async def error_handler(update, context):
    """Handle errors"""
    logger.error(f"Update {update} caused error {context.error}", exc_info=True)
    
    try:
        import html
        error_msg = html.escape(str(context.error)[:200])
        await update.message.reply_text(
            f"❌ <b>Bot Error</b>\n\n"
            f"{error_msg}\n\n"
            f"Please try again or use /help for assistance.",
            parse_mode='HTML'
        )
    except:
        pass


# === Helpers for enhanced delete/modify flows (ported from Windows, adapted to Linux) ===

async def show_transaction_selection_for_delete(update, context):
    """Show recent transactions for deletion selection menu"""
    await update.message.reply_text("📋 Loading last 10 transactions...")
    recent_transactions = get_recent_transactions(10)
    if not recent_transactions:
        await update.message.reply_text("❌ No transactions found.")
        return

    response_lines = ["🗑️ <b>Recent Transactions (Select one to delete):</b>\n"]
    for i, t in enumerate(recent_transactions, 1):
        date_val = t['date']
        date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)
        amounts = []
        if t['usd'] is not None:
            prefix = "🔄 " if t['usd'] < 0 else ""
            amounts.append(f"{prefix}${abs(t['usd']):.2f}")
        if t['lbp'] is not None:
            prefix = "🔄 " if t['lbp'] < 0 else ""
            amounts.append(f"{prefix}{abs(t['lbp']):,.0f} LBP")
        if t['euro'] is not None:
            prefix = "🔄 " if t['euro'] < 0 else ""
            amounts.append(f"{prefix}€{abs(t['euro']):.2f}")
        amount_str = " + ".join(amounts) if amounts else "No amount"

        item = t.get("subcategory") or t.get("category", "")
        import html
        safe_item = html.escape(item) if item else "Unknown"
        row_num = t.get("row", "N/A")

        response_lines.append(
            f"\n<b>{i}.</b> {'💳' if str(t['payment']).lower()=='card' else '💵'} {date_str}: "
            f"<code>{safe_item}</code>\n Amount: {amount_str}\n Category: {html.escape(t.get('category','Unknown'))}"
            f"\n Row: {row_num}"
        )

    response_lines.append(
        "\n\n<b>To delete, reply with:</b>\n"
        "<code>/delete NUMBER</code>\n\n"
        "<b>Examples:</b>\n"
        "<code>/delete 1</code> to select the first transaction\n"
        "<code>/delete 83 confirm</code> to delete row 83 immediately\n\n"
        "⚠️ <b>Warning:</b> Deletion cannot be undone!"
    )
    await update.message.reply_text("\n".join(response_lines), parse_mode='HTML')


async def show_deletion_confirmation(update, transaction):
    """Show confirmation message before deleting a transaction"""
    date_val = transaction['date']
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)

    amounts = []
    if transaction['usd'] is not None:
        prefix = "🔄 " if transaction['usd'] < 0 else ""
        amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
    if transaction['lbp'] is not None:
        prefix = "🔄 " if transaction['lbp'] < 0 else ""
        amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
    if transaction['euro'] is not None:
        prefix = "🔄 " if transaction['euro'] < 0 else ""
        amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
    amount_str = " + ".join(amounts) if amounts else "No amount"

    payment_emoji = "💳" if str(transaction['payment']).lower() == "card" else "💵"
    item = transaction['subcategory'] if transaction['subcategory'] else transaction['category']
    import html
    safe_item = html.escape(item) if item else "Unknown"
    row_number = transaction.get('row', 'N/A')

    response = (
        f"⚠️ <b>Confirm Deletion:</b>\n\n"
        f"{payment_emoji} {date_str}: <code>{safe_item}</code>\n"
        f"Amount: {amount_str}\n"
        f"Category: {html.escape(transaction.get('category','Unknown'))}\n"
        f"Payment: {html.escape(transaction.get('payment','Cash'))}\n"
    )
    if transaction.get('notes'):
        response += f"\nNotes: {html.escape(transaction['notes'][:100])}{'...' if len(transaction['notes'])>100 else ''}"

    response += (
        f"\n\n<b>Are you sure you want to delete this transaction?</b>\n\n"
        f"✅ To confirm deletion, type:\n"
        f"<code>/delete {row_number} confirm</code>\n\n"
        f"❌ To cancel, just ignore this message.\n"
        f"<i>Note: This action cannot be undone!</i>"
    )
    await update.message.reply_text(response, parse_mode='HTML')


async def process_deletion(update, transaction):
    """Process the deletion of a transaction"""
    row_number = transaction.get('row')
    if not row_number:
        await update.message.reply_text("❌ Error: Could not determine row number.")
        return
    await update.message.reply_text(f"🗑️ Deleting transaction at row {row_number}...")
    success, message, deleted_transaction = delete_transaction_at_row(row_number)
    await send_deletion_result(update, deleted_transaction, success, message)


async def send_deletion_result(update, deleted_transaction, success, message):
    """Send the deletion result to user"""
    if success and deleted_transaction:
        date_val = deleted_transaction['date']
        date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)

        amounts = []
        if deleted_transaction.get('usd') is not None:
            prefix = "🔄 " if deleted_transaction['usd'] < 0 else ""
            amounts.append(f"{prefix}${abs(deleted_transaction['usd']):.2f}")
        if deleted_transaction.get('lbp') is not None:
            prefix = "🔄 " if deleted_transaction['lbp'] < 0 else ""
            amounts.append(f"{prefix}{abs(deleted_transaction['lbp']):,.0f} LBP")
        if deleted_transaction.get('euro') is not None:
            prefix = "🔄 " if deleted_transaction['euro'] < 0 else ""
            amounts.append(f"{prefix}€{abs(deleted_transaction['euro']):.2f}")
        amount_str = " + ".join(amounts) if amounts else "No amount"

        payment_emoji = "💳" if str(deleted_transaction.get('payment','')).lower() == "card" else "💵"
        item = deleted_transaction.get('subcategory') or deleted_transaction.get('category', 'Unknown')
        import html
        safe_item = html.escape(item) if item else "Unknown"

        details = (
            f"🗑️ <b>Transaction Deleted:</b>\n\n"
            f"{payment_emoji} {date_str}: {safe_item}\n"
            f"Amount: {amount_str}\n"
            f"Category: {html.escape(deleted_transaction.get('category','Unknown'))}\n"
            f"Payment: {html.escape(deleted_transaction.get('payment','Cash'))}\n"
        )
        if deleted_transaction.get('notes'):
            details += f"\nNotes: {html.escape(str(deleted_transaction['notes'])[:100])}{'...' if len(str(deleted_transaction['notes']))>100 else ''}"

        await update.message.reply_text(f"{details}\n\n{message}", parse_mode='HTML')
    else:
        await update.message.reply_text(message, parse_mode='HTML')


async def show_transaction_selection(update, context):
    """Show recent transactions for selection menu (modify)"""
    await update.message.reply_text("📋 Loading last 10 transactions...")
    recent_transactions = get_recent_transactions(10)
    if not recent_transactions:
        await update.message.reply_text("❌ No transactions found.")
        return

    response_lines = ["📋 <b>Recent Transactions (Select one to modify):</b>\n"]
    for i, t in enumerate(recent_transactions, 1):
        date_val = t['date']
        date_str = date_val.strftime("%d/%m") if hasattr(date_val, 'strftime') else str(date_val)

        amounts = []
        if t['usd'] is not None:
            prefix = "🔄 " if t['usd'] < 0 else ""
            amounts.append(f"{prefix}${abs(t['usd']):.2f}")
        if t['lbp'] is not None:
            prefix = "🔄 " if t['lbp'] < 0 else ""
            amounts.append(f"{prefix}{abs(t['lbp']):,.0f} LBP")
        if t['euro'] is not None:
            prefix = "🔄 " if t['euro'] < 0 else ""
            amounts.append(f"{prefix}€{abs(t['euro']):.2f}")
        amount_str = " + ".join(amounts) if amounts else "No amount"

        payment_emoji = "💳" if str(t['payment']).lower() == "card" else "💵"
        item = t['subcategory'] if t['subcategory'] else t['category']
        import html
        safe_item = html.escape(item) if item else "Unknown"
        response_lines.append(
            f"\n<b>{i}.</b> {payment_emoji} {date_str}: <code>{safe_item}</code>\n"
            f" Amount: {amount_str}\n Category: {html.escape(t.get('category','Unknown'))}"
        )

    response_lines.append(
        "\n\n<b>To modify, reply with:</b>\n"
        "<code>/modify NUMBER</code>\n\n"
        "<b>Example:</b> <code>/modify 1</code> to modify the first transaction\n"
        f"<i>Showing last {len(recent_transactions)} transactions</i>"
    )
    await update.message.reply_text("\n".join(response_lines), parse_mode='HTML')


async def show_selected_transaction_flexible(update, transaction):
    """Show selected transaction with flexible modification options"""
    date_val = transaction['date']
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, 'strftime') else str(date_val)

    amounts = []
    if transaction['usd'] is not None:
        prefix = "🔄 " if transaction['usd'] < 0 else ""
        amounts.append(f"{prefix}${abs(transaction['usd']):.2f}")
    if transaction['lbp'] is not None:
        prefix = "🔄 " if transaction['lbp'] < 0 else ""
        amounts.append(f"{prefix}{abs(transaction['lbp']):,.0f} LBP")
    if transaction['euro'] is not None:
        prefix = "🔄 " if transaction['euro'] < 0 else ""
        amounts.append(f"{prefix}€{abs(transaction['euro']):.2f}")
    amount_str = " + ".join(amounts) if amounts else "No amount"

    payment_emoji = "💳" if str(transaction['payment']).lower() == "card" else "💵"
    item = transaction['subcategory'] if transaction['subcategory'] else transaction['category']
    import html
    safe_item = html.escape(item) if item else "Unknown"

    response = (
        f"🛠️ <b>Selected Transaction:</b>\n\n"
        f"{payment_emoji} {date_str}: <code>{safe_item}</code>\n"
        f"Amount: {amount_str}\n"
        f"Category: {html.escape(transaction.get('category','Unknown'))}\n"
        f"Payment: {html.escape(transaction.get('payment','Cash'))}\n"
    )
    if transaction.get('notes'):
        response += f"Notes: {html.escape(transaction['notes'][:100])}{'...' if len(transaction['notes'])>100 else ''}\n"

    response += (
        "\n<b>Now modify what you want:</b>\n\n"
        "<b>Change everything:</b>\n"
        "<code>/modify NewItem 15$ card</code>\n\n"
        "<b>Change amount only:</b>\n"
        "<code>/modify 20$</code> or <code>/modify 150000 LBP</code>\n\n"
        "<b>Change payment only:</b>\n"
        "<code>/modify to card</code> or <code>/modify payment: cash</code>\n\n"
        "<b>Change item only:</b>\n"
        "<code>/modify Chamsin</code>\n\n"
        "<b>Add notes:</b>\n"
        "<code>/modify NOTES : Your notes here</code>\n\n"
        "<b>Keep original:</b> Use <code>same</code> or <code>keep</code>\n"
        "Example: <code>/modify same 20$</code> (keeps item, changes amount)"
    )
    await update.message.reply_text(response, parse_mode='HTML')


async def process_modification(update, row_number, original_transaction, modify_text):
    """Process a modification request"""
    new_subcategory = None
    new_currency_amounts = None
    new_payment_type = None
    new_notes = None

    text_lower = modify_text.lower()

    # payment-only quick path
    payment_keywords = {
        'card': 'Card',
        'cash': 'Cash',
        'bank': 'Bank Transfer',
        'digital': 'Digital Wallet',
        'transfer': 'Bank Transfer',
        'wallet': 'Digital Wallet'
    }
    payment_only = any(kw in text_lower for kw in ['to card', 'to cash', 'to bank', 'to digital', 'payment:', 'pay:'])
    if payment_only:
        for kw, val in payment_keywords.items():
            if kw in text_lower:
                new_payment_type = val
                break
        if new_payment_type:
            await update.message.reply_text(f"🛠️ Changing payment to {new_payment_type}...")
            success, message, modified_transaction = modify_transaction_at_row(
                row_number, None, None, new_payment_type, None
            )
            await send_modification_result(update, original_transaction, modified_transaction, success, message)
            return

    # notes extraction
    notes_pattern = r'\s+NOTES\s*:\s*(.+)'
    notes_match = re.search(notes_pattern, modify_text, re.IGNORECASE)
    if notes_match:
        modify_text = modify_text[:notes_match.start()].strip()
        new_notes = notes_match.group(1).strip()

    # amount-only quick path
    amount_only_pattern = r'^[\$\€\-]?\d'
    is_amount_only = re.match(amount_only_pattern, modify_text.strip()) is not None
    if is_amount_only:
        _, parsed_amounts, _ = extract_payment_amount_currency(modify_text)
        if parsed_amounts:
            new_currency_amounts = parsed_amounts
            await update.message.reply_text(f"🛠️ Changing amount to {modify_text}...")
            success, message, modified_transaction = modify_transaction_at_row(
                row_number, None, new_currency_amounts, None, new_notes
            )
            await send_modification_result(update, original_transaction, modified_transaction, success, message)
            return

    # full parsing
    parsed_subcategory, parsed_amounts, parsed_payment = extract_payment_amount_currency(modify_text)
    if parsed_subcategory and parsed_subcategory.lower() not in ['same', 'keep', 'unchanged', '']:
        new_subcategory = parsed_subcategory
    if parsed_amounts:
        new_currency_amounts = parsed_amounts
    if parsed_payment and parsed_payment != "Cash":
        new_payment_type = parsed_payment

    if new_subcategory and new_subcategory.lower() in ['same', 'keep', 'unchanged']:
        new_subcategory = None

    await update.message.reply_text(f"🛠️ Modifying transaction at row {row_number}...")
    success, message, modified_transaction = modify_transaction_at_row(
        row_number, new_subcategory, new_currency_amounts, new_payment_type, new_notes
    )
    await send_modification_result(update, original_transaction, modified_transaction, success, message)


async def send_modification_result(update, original_transaction, modified_transaction, success, message):
    """Send the modification result to user"""
    if success and modified_transaction:
        response = format_transaction_response(modified_transaction, "Modified")

        changes = []
        if original_transaction.get('subcategory') != modified_transaction.get('subcategory'):
            old_item = original_transaction.get('subcategory') or original_transaction.get('category', 'Unknown')
            new_item = modified_transaction.get('subcategory') or modified_transaction.get('category', 'Unknown')
            if old_item != new_item:
                changes.append(f"Item: {old_item} → {new_item}")

        for currency in ['USD', 'LBP', 'EURO']:
            old_val = original_transaction.get(currency.lower())
            new_val = modified_transaction.get(currency.lower())
            if old_val != new_val:
                old_str = format_currency_amount(old_val, currency) if old_val is not None else "None"
                new_str = format_currency_amount(new_val, currency) if new_val is not None else "None"
                changes.append(f"{currency}: {old_str} → {new_str}")

        if original_transaction.get('payment') != modified_transaction.get('payment'):
            changes.append(f"Payment: {original_transaction.get('payment','Cash')} → {modified_transaction.get('payment','Cash')}")

        if original_transaction.get('notes') != modified_transaction.get('notes'):
            old_notes = original_transaction.get('notes', 'None') or 'None'
            new_notes = modified_transaction.get('notes', 'None') or 'None'
            old_preview = str(old_notes)[:50] + "..." if len(str(old_notes)) > 50 else str(old_notes)
            new_preview = str(new_notes)[:50] + "..." if len(str(new_notes)) > 50 else str(new_notes)
            changes.append(f"Notes: {old_preview} → {new_preview}")

        if changes:
            changes_text = "\n".join([f"• {c}" for c in changes])
            full_response = f"{response}\n\n<b>Changes:</b>\n{changes_text}\n\n{message}"
        else:
            full_response = f"{response}\n\n{message}"

        await update.message.reply_text(full_response, parse_mode='HTML')
    else:
        await update.message.reply_text(message, parse_mode='HTML')


# ========== TELEGRAM COMMANDS: /delete and /modify ==========


async def delete_command(update, context):
    """
    Usage:
    /delete → show last 10 with row numbers and instructions
    /delete last → ask to confirm deleting last transaction
    /delete last confirm → delete last transaction
    /delete 125 → ask to confirm deleting row 125 (direct row)
    /delete 125 confirm → delete row 125 (direct row)
    /delete 3 → select the 3rd transaction from the recent list
    /delete 3 confirm → delete the 3rd transaction from the recent list
    """
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    args = [a.strip().lower() for a in context.args] if context.args else []

    # No args → show selector of recent 10
    if not args:
        await show_transaction_selection_for_delete(update, context)
        return

    # /delete last confirm
    if args[0] == "last":
        if len(args) >= 2 and args[1] == "confirm":
            ok, msg, deleted = delete_last_transaction()
            await update.message.reply_text(msg)
        else:
            await update.message.reply_text("Type: /delete last confirm to delete the last transaction.")
        return

    # If first arg is a plain integer, we need to decide:
    # 1) if a direct row number (confirm → delete that row), OR
    # 2) if a selection index among the last 10.
    if args[0].isdigit():
        num = int(args[0])

        # Direct-row + confirm? (like /delete 125 confirm)
        if len(args) >= 2 and args[1] == "confirm":
            ok, msg, deleted = delete_transaction_at_row(num)
            await update.message.reply_text(msg)
            return

        # Otherwise treat it as selection among recent 10
        recent_transactions = get_recent_transactions(10)
        if not recent_transactions:
            await update.message.reply_text("❌ No transactions found.")
            return
        if num < 1 or num > len(recent_transactions):
            await update.message.reply_text(f"❌ Invalid selection. Choose 1-{len(recent_transactions)}")
            return

        # Ask for confirmation using selection
        selected = recent_transactions[num - 1]
        await show_deletion_confirmation(update, selected)
        return

    # Fallback help
    await update.message.reply_text(
        "🗑️ <b>Delete Transaction Command</b>\n\n"
        "<b>To delete a transaction:</b>\n"
        "1. <code>/delete</code> — See recent transactions list\n"
        "2. <code>/delete 3</code> — Select item #3 from the recent list\n"
        "3. <code>/delete 83 confirm</code> — Delete row 83 directly\n\n"
        "<b>Old method:</b>\n"
        "<code>/delete last confirm</code> — Delete last transaction",
        parse_mode='HTML'
    )


def _parse_modify_args(arg_list: List[str]) -> Tuple[Optional[int], Optional[str], Dict[str, float], Optional[str], Optional[str]]:
    """
    Parse /modify arguments.
    Supported:
      /modify <row> item="New Name" usd=12.5 lbp=100000 euro=-3 payment=Card notes="any text"
    Quotes are optional for single-word values.
    """
    if not arg_list:
        return None, None, {}, None, None

    # Row
    try:
        row = int(arg_list[0])
    except ValueError:
        return None, None, {}, None, None

    item = None
    amounts: Dict[str, float] = {}
    payment = None
    notes = None

    def _take_value(token: str) -> str:
        # token is key=value or key="multi word"
        if "=" not in token:
            return ""
        val = token.split("=", 1)[1].strip()
        if val.startswith('"') and val.endswith('"') and len(val) >= 2:
            return val[1:-1]
        if val.startswith("'") and val.endswith("'") and len(val) >= 2:
            return val[1:-1]
        return val

    for raw in arg_list[1:]:
        t = raw.strip()
        t_low = t.lower()

        if t_low.startswith("item="):
            item = _take_value(t)
        elif t_low.startswith("usd="):
            try:
                amounts["USD"] = float(_take_value(t))
            except ValueError:
                pass
        elif t_low.startswith("lbp="):
            try:
                amounts["LBP"] = float(_take_value(t))
            except ValueError:
                pass
        elif t_low.startswith("euro=") or t_low.startswith("eur="):
            try:
                amounts["EURO"] = float(_take_value(t))
            except ValueError:
                pass
        elif t_low.startswith("payment="):
            payment = _take_value(t)
        elif t_low.startswith("notes="):
            notes = _take_value(t)

    return row, item, amounts, payment, notes


async def modify_command(update, context):
    """
    Supported:
    /modify (shows last 10, selection UI)
    /modify 3 (selects 3rd recent)
    /modify <row> item="New" usd=12.5 lbp=100000 euro=-3 payment=Card notes="..."
    /modify 3 55$   or   /modify 3 to card   or   /modify 3 NOTES : text
    Then you can send: /modify 55$  or  /modify to cash  (after selecting)
    """
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return

    # If explicit key=value syntax is used, keep your Linux parser path
    # Examples: item=..., usd=..., lbp=..., euro=..., payment=..., notes=...
    if context.args and (context.args[0].isdigit() and any(("=" in tok) for tok in context.args[1:])):
        row, item, amts, payment, notes = _parse_modify_args(context.args)
        if row is None:
            await update.message.reply_text('❌ Invalid syntax. Example:\n/modify 125 item="Chamsin" usd=10.5 payment=Card')
            return
        amts = amts or None
        ok, msg, modified = modify_transaction_at_row(
            row=row,
            new_subcategory=item,
            new_currency_amounts=amts,
            new_payment_type=payment,
            new_notes=notes,
        )
        await update.message.reply_text(msg)
        return

    # Enhanced flexible flow (ported)
    if not context.args:
        await show_transaction_selection(update, context)
        return

    args = context.args

    # Case: /modify 3 <something>
    if args[0].isdigit() and len(args) > 1:
        selection_num = int(args[0])
        recent_transactions = get_recent_transactions(10)
        if not recent_transactions:
            await update.message.reply_text("❌ No transactions found.")
            return
        if selection_num < 1 or selection_num > len(recent_transactions):
            await update.message.reply_text(f"❌ Invalid selection. Choose 1-{len(recent_transactions)}")
            return
        selected = recent_transactions[selection_num - 1]
        row_number = selected['row']
        modify_text = ' '.join(args[1:])
        await process_modification(update, row_number, selected, modify_text)
        return

    # Case: /modify 3    (just select and show options)
    if args[0].isdigit() and len(args) == 1:
        selection_num = int(args[0])
        recent_transactions = get_recent_transactions(10)
        if not recent_transactions:
            await update.message.reply_text("❌ No transactions found.")
            return
        if selection_num < 1 or selection_num > len(recent_transactions):
            await update.message.reply_text(f"❌ Invalid selection. Choose 1-{len(recent_transactions)}")
            return
        selected = recent_transactions[selection_num - 1]
        context.user_data['modify_row'] = selected['row']
        context.user_data['original_transaction'] = selected
        await show_selected_transaction_flexible(update, selected)
        return

    # Case: already selected earlier, now user sends the change text
    if 'modify_row' in context.user_data:
        row_number = context.user_data['modify_row']
        original_transaction = context.user_data.get('original_transaction', {})
        modify_text = ' '.join(args)
        await process_modification(update, row_number, original_transaction, modify_text)
        return

    # Fallback help
    await update.message.reply_text(
        "❌ <b>No transaction selected.</b>\n\n"
        "<b>Select from last 10:</b>\n"
        "1. <code>/modify</code> — See last 10 transactions\n"
        "2. <code>/modify 1</code> — Select item #1\n"
        "3. <code>/modify 1 55$</code> — Change amount to $55\n"
        "4. <code>/modify 1 to card</code> — Change payment to card\n\n"
        "<b>Direct explicit form:</b>\n"
        "<code>/modify 125 item=\"New Name\" usd=12.5 payment=Card notes=\"text\"</code>",
        parse_mode='HTML'
    )

async def check_sync_status_command(update, context):
    """Check the status of OneDrive sync"""
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Unauthorized.")
        return
    
    # Check if background sync is running
    if _upload_in_progress.is_set():
        await update.message.reply_text(
            "🔄 <b>OneDrive Sync Status</b>\n\n"
            "• Status: <b>Sync in progress</b>\n"
            "• Background task: <b>Running</b>\n"
            "• Please wait for completion...",
            parse_mode='HTML'
        )
    elif _bg_sync_future and not _bg_sync_future.done():
        await update.message.reply_text(
            "🔄 <b>OneDrive Sync Status</b>\n\n"
            "• Status: <b>Background sync queued</b>\n"
            "• Background task: <b>Pending</b>\n"
            "• Your changes will sync automatically.",
            parse_mode='HTML'
        )
    else:
        # Try a test upload
        await update.message.reply_text("🔄 Testing OneDrive connection...")
        success, msg = save_excel_to_onedrive()
        
        if success:
            await update.message.reply_text(
                "✅ <b>OneDrive Sync Status</b>\n\n"
                "• Status: <b>Fully synced</b>\n"
                "• Connection: <b>Working</b>\n"
                "• All changes are up to date.",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                f"⚠️ <b>OneDrive Sync Status</b>\n\n"
                f"• Status: <b>Sync issues detected</b>\n"
                f"• Error: {msg}\n\n"
                f"Try: /sync to force a manual sync",
                parse_mode='HTML'
            )

# ========== MAIN FUNCTION ==========

def main():
    """Start the bot"""
    print("\n" + "="*70)
    print("💰 SMART BUDGET TRACKER BOT - LINUX VERSION")
    print("Optimized for Ubuntu 22.04 aarch64")
    print("="*70)
    
    # Check BOT_TOKEN
    if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        print("❌ ERROR: BOT_TOKEN not set in .env file!", file=sys.stderr)
        logger.error("BOT_TOKEN not configured")
        sys.exit(1)
    
    # Check ALLOWED_USER_IDS
    if not ALLOWED_USER_IDS:
        print("❌ ERROR: No ALLOWED_USER_IDS or ALLOWED_USER_ID set in .env file!", file=sys.stderr)
        logger.error("ALLOWED_USER_IDS not configured")
        print("Add your user ID to .env file:")
        print("Example: ALLOWED_USER_ID=1663164223")
        sys.exit(1)
    
    print(f"✅ BOT_TOKEN: {BOT_TOKEN[:10]}...")
    print(f"✅ ALLOWED_USER_IDS: {ALLOWED_USER_IDS}")
    print(f"✅ Excel library: {'✅ openpyxl/pandas' if EXCEL_AVAILABLE else '❌ NOT INSTALLED'}")
    print(f"✅ OneDrive API: {'✅ onedrivesdk' if ONEDRIVE_AVAILABLE else '❌ NOT INSTALLED'}")
    print(f"✅ Minimum Confidence: {MINIMUM_CONFIDENCE*100:.0f}%")
    print(f"✅ Negative Amounts: ✅ SUPPORTED")
    print(f"✅ LBP without commas: ✅ SUPPORTED")
    print(f"✅ Download Feature: ✅ ADDED")
    
    # Setup directories
    BACKUP_DIR.mkdir(exist_ok=True, parents=True)
    EXPORT_DIR.mkdir(exist_ok=True, parents=True)
    LOCAL_COPY_PATH.parent.mkdir(exist_ok=True, parents=True)

    # Check for OneDrive configuration
    print("\n🔐 Checking OneDrive configuration...")
    if not ONEDRIVE_CLIENT_ID or ONEDRIVE_CLIENT_ID == "YOUR_CLIENT_ID":
        print("⚠️  ONEDRIVE_CLIENT_ID not set. OneDrive sync will not work.")
        print("Run /onedrive_auth in Telegram after setting up your Azure app.")
    
    # Test OneDrive connection if configured
    if ONEDRIVE_CLIENT_ID and ONEDRIVE_CLIENT_ID != "YOUR_CLIENT_ID":
        print("Testing OneDrive connection...")
        success, msg = download_from_onedrive()
        if success:
            print(f"✅ OneDrive: {msg}")
        else:
            print(f"⚠️  OneDrive: {msg}")
            print("Run /onedrive_auth in Telegram to authenticate.")
    else:
        print("⚠️  OneDrive not configured. Files will only be saved locally.")
    
    # Load tables (requires openpyxl)
    print("\n🔧 Loading data from Excel...")
    if EXCEL_AVAILABLE and LOCAL_COPY_PATH.exists():
        tables_dict = load_all_tables_with_details()
        if tables_dict:
            total_items = sum(data['count'] for data in tables_dict.values())
            print(f"✅ Loaded {len(tables_dict)} categories with {total_items} items")
        else:
            print("❌ Could not load data from Excel")
    else:
        print("⚠️  openpyxl not installed or Excel file not found.")
        print("Install with: pip install openpyxl pandas")
    
    print("\n" + "="*70)
    print("🤖 BOT IS STARTING...")
    print("="*70)
    
    # Create bot application
    try:
        app = Application.builder().token(BOT_TOKEN).build()
        logger.info("Bot application created successfully")
    except Exception as e:
        print(f"❌ Failed to create bot: {e}", file=sys.stderr)
        logger.exception("Failed to create bot application")
        sys.exit(1)
    
    # Add command handlers (each command registered only once)
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("testparse", testparse_command))
    app.add_handler(CommandHandler("recent", recent_command))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CommandHandler("save", save_command))
    app.add_handler(CommandHandler("clearcache", clear_cache_command))
    app.add_handler(CommandHandler("download", download_command))
    app.add_handler(CommandHandler("unlock", unlock_command))
    app.add_handler(CommandHandler("onedrive_auth", onedrive_auth_command))
    app.add_handler(CommandHandler("onedrive_test", onedrive_test_command))
    app.add_handler(CommandHandler("onedrive_code", onedrive_complete_auth_command))
    app.add_handler(CommandHandler("onedrive_code_url", onedrive_complete_auth_from_url_command))
    app.add_handler(CommandHandler("direct_auth", direct_auth_command))
    app.add_handler(CommandHandler("delete", delete_command))
    app.add_handler(CommandHandler("modify", modify_command))
    app.add_handler(CommandHandler("debug_onedrive", debug_onedrive_command))
    app.add_handler(CommandHandler("manual_upload", manual_upload_command))
    app.add_handler(CommandHandler("check_sync", check_sync_status_command))
    app.add_handler(CommandHandler("sync", force_sync_command))

    
    # Add message handler
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Add error handler
    app.add_error_handler(error_handler)
    
    print("\n✅ Bot is ready and running!")
    print("📱 Open Telegram and send /start to your bot")
    print("🔐 New: Use /onedrive_auth to setup OneDrive")
    print("📥 Use /download to export data")
    print("⏳ Waiting for messages...")
    print("="*70)
    
    try:
        # Start polling
        app.run_polling(
            poll_interval=1.0,
            timeout=10,
            drop_pending_updates=True
        )
    except KeyboardInterrupt:
        print("\n🛑 Bot stopped by user")
    except Exception as e:
        print(f"\n❌ Bot crashed: {e}", file=sys.stderr)
        logger.error(f"Bot crashed: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # Make sure startup exceptions are visible
        print(f"Fatal error starting bot: {e}", file=sys.stderr)
        logger.exception("Fatal error in main()")
        raise